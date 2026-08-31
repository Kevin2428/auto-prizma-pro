from playwright.sync_api import (
    sync_playwright,
    TimeoutError as PlaywrightTimeoutError,
)

from openpyxl import load_workbook

import csv
import os
import re
import zipfile
import io
import json
import unicodedata
import shutil
import traceback
import time
import threading
import html as html_lib
from urllib.parse import quote
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError


# ============================================================
# CONFIGURACIÓN
# ============================================================

URL_PRIZMA = "https://admin.prizma.site/inicio-sesion"

VERSION_SCRIPT = "PRUEBA_H5P_NUEVO_GESTION_FINANCIERA_V2"

BUILD_INTERNO = "ESTABILIDAD_COLA_CANCELACION_V10"


TEXTOS_DESCRIPCION_A_BORRAR = [
    "NO_DISPONIBLE",
    "NO_DISPOINBLE",
]


# ============================================================
# ESTADO DEL TRABAJO
# ============================================================

def actualizar_estado(
    estado,
    etapa=None,
    mensaje=None,
    total=None,
    procesadas=None,
    exitosas=None,
    errores=None,
    terminado=None,
):

    if etapa is not None:
        estado["etapa"] = etapa

    if mensaje is not None:
        estado["mensaje"] = mensaje

    if total is not None:
        estado["total"] = total

    if procesadas is not None:
        estado["procesadas"] = procesadas

    if exitosas is not None:
        estado["exitosas"] = exitosas

    if errores is not None:
        estado["errores"] = errores

    if terminado is not None:
        estado["terminado"] = terminado


# ============================================================
# DETALLE DE ACTIVIDADES PARA LA INTERFAZ
# ============================================================

def preparar_detalle_actividades(
    estado,
    actividades,
):

    estado["detalle_actividades"] = [
        {
            "numero": numero,
            "nombre": actividad["nombre"],
            "estado": "pendiente",
            "error": "",
        }
        for numero, actividad in enumerate(
            actividades,
            start=1,
        )
    ]


def actualizar_detalle_actividad(
    estado,
    numero,
    nuevo_estado,
    error="",
):

    detalle = estado.get(
        "detalle_actividades",
        [],
    )

    indice = numero - 1

    if not (
        0 <= indice < len(detalle)
    ):
        return

    detalle[indice][
        "estado"
    ] = nuevo_estado

    detalle[indice][
        "error"
    ] = error or ""


# ============================================================
# NORMALIZACIÓN
# ============================================================

def normalizar_texto(texto):

    if texto is None:
        return ""

    texto = str(texto)

    texto = unicodedata.normalize(
        "NFD",
        texto,
    )

    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(caracter) != "Mn"
    )

    texto = texto.lower()

    texto = re.sub(
        r"[^a-z0-9]+",
        " ",
        texto,
    )

    return " ".join(
        texto.split()
    )


def normalizar_categoria(categoria):

    valor = normalizar_texto(
        categoria
    )

    if valor == "ovi":
        return "OVI"

    if valor == "ova":
        return "OVA"

    if valor == "challenge":
        return "CHALLENGE"

    return None


# ============================================================
# VARIANTES SEGURAS DE PROGRAMA
# ============================================================

def obtener_variantes_programa(
    programa,
):

    programa_n = normalizar_texto(
        programa
    )

    if not programa_n:
        return set()

    variantes = {
        programa_n
    }

    partes = programa_n.split()

    # PRIZMA y las matrices no siempre usan exactamente el mismo
    # prefijo para las especializaciones. Ejemplos reales:
    #   Matriz:  "Esp en Crisis, Conflicto..."
    #   PRIZMA:  "Especialización Crisis, Conflicto... - Virtual"
    #
    # Generamos únicamente variantes equivalentes del prefijo, sin
    # relajar la comparación del nombre propio del programa.
    if partes and partes[0] in {"esp", "especializacion"}:

        resto_partes = partes[1:]

        if resto_partes:
            resto = " ".join(resto_partes)
            variantes.add(
                "especializacion " + resto
            )

            if resto_partes[0] == "en":
                resto_sin_en = " ".join(
                    resto_partes[1:]
                )

                if resto_sin_en:
                    variantes.add(
                        "especializacion "
                        + resto_sin_en
                    )
                    variantes.add(
                        "esp " + resto_sin_en
                    )
            else:
                variantes.add(
                    "especializacion en " + resto
                )
                variantes.add(
                    "esp en " + resto
                )

    return variantes


# ============================================================
# NOMBRES DE RECURSOS CORTOS / TRUNCADOS
# ============================================================

def limpiar_prefijo_tecnico_recurso(
    nombre_archivo,
):

    nombre_sin_extension = os.path.splitext(
        os.path.basename(
            nombre_archivo
        )
    )[0]

    nombre_n = normalizar_texto(
        nombre_sin_extension
    )

    # Prefijos técnicos observados en paquetes de recursos.
    # Ejemplos:
    # ID003-Tema_4-Ruta_para_analizar.h5p
    # U1-T1-Principios jurídicos.pdf
    # U2-T4-Caso aplicado.pdf
    #
    # normalizar_texto() convierte guiones y guiones bajos en espacios,
    # por eso los patrones se aplican sobre la versión normalizada.
    nombre_n = re.sub(
        r"^id\d+\s+tema\s+\d+\s+",
        "",
        nombre_n,
    )

    nombre_n = re.sub(
        r"^u\d+\s+t\d+\s+",
        "",
        nombre_n,
    )

    return nombre_n


def recurso_es_prefijo_de_actividad(
    nombre_archivo,
    actividad,
):

    archivo_n = limpiar_prefijo_tecnico_recurso(
        nombre_archivo
    )

    actividad_n = normalizar_texto(
        actividad["nombre"]
    )

    if not archivo_n or not actividad_n:
        return False

    palabras_archivo = archivo_n.split()
    palabras_actividad = actividad_n.split()

    if not palabras_archivo:
        return False

    if len(palabras_archivo) > len(palabras_actividad):
        return False

    # Permitimos incluso una sola palabra cuando es suficientemente
    # descriptiva. La seguridad real se valida después comprobando
    # que ese prefijo identifique una única actividad del curso.
    if len(palabras_archivo[0]) < 4:
        return False

    for indice, palabra_archivo in enumerate(
        palabras_archivo
    ):

        palabra_actividad = palabras_actividad[indice]

        # Todas las palabras anteriores a la última deben coincidir.
        if indice < len(palabras_archivo) - 1:

            if palabra_archivo != palabra_actividad:
                return False

            continue

        # La última palabra puede estar completa o truncada.
        if palabra_archivo == palabra_actividad:
            continue

        if len(palabra_archivo) < 4:
            return False

        if not palabra_actividad.startswith(
            palabra_archivo
        ):
            return False

    return True


def recurso_corto_es_unico_para_actividad(
    nombre_archivo,
    actividad,
    actividades_curso,
):

    if not actividades_curso:
        return False

    tipo_objetivo = actividad[
        "tipo_archivo"
    ]

    compatibles = []

    for candidata in actividades_curso:

        if candidata[
            "tipo_archivo"
        ] != tipo_objetivo:
            continue

        if recurso_es_prefijo_de_actividad(
            nombre_archivo,
            candidata,
        ):

            compatibles.append(
                candidata
            )

    if len(compatibles) != 1:
        return False

    unica = compatibles[0]

    return (
        unica["fila_excel"]
        == actividad["fila_excel"]
    )


def obtener_titulo_interno_h5p(
    recurso,
):
    """
    Lee únicamente h5p.json de un H5P concreto.

    IMPORTANTE:
    - No se usa durante el indexado general.
    - Solo se invoca como fallback cuando el nombre externo quedó
      ambiguo y el flujo normal no pudo resolverlo de forma segura.
    - No modifica la regla de ERROR_RECURSO_DUPLICADO.
    """

    if recurso.get("extension") != ".h5p":
        return ""

    try:
        with zipfile.ZipFile(
            recurso["zip"],
            "r",
        ) as zip_ref:

            datos_h5p = zip_ref.read(
                recurso["miembro"]
            )

        with zipfile.ZipFile(
            io.BytesIO(datos_h5p),
            "r",
        ) as h5p_ref:

            if "h5p.json" not in h5p_ref.namelist():
                return ""

            datos_json = h5p_ref.read(
                "h5p.json"
            )

        manifiesto = json.loads(
            datos_json.decode(
                "utf-8-sig"
            )
        )

        titulo = manifiesto.get(
            "title",
            "",
        )

        return str(titulo).strip()

    except Exception as error:
        print(
            "No se pudo leer h5p.json de",
            recurso.get("nombre", ""),
            "|",
            type(error).__name__,
            str(error),
        )
        return ""


def resolver_h5p_ambiguo_por_titulo_interno(
    actividad,
    recursos_ambiguos,
):
    """
    Intenta resolver un ERROR_RECURSO_AMBIGUO leyendo el título interno
    de los pocos H5P implicados. Exige coincidencia EXACTA normalizada.

    Si encuentra 0 o más de 1 coincidencias, no adivina.
    """

    if not recursos_ambiguos:
        return None

    actividad_n = normalizar_texto(
        actividad["nombre"]
    )

    coincidencias = []

    # Evitar abrir dos veces el mismo miembro si llegó repetido a la lista.
    vistos = set()

    for recurso in recursos_ambiguos:
        clave = (
            recurso.get("zip"),
            recurso.get("miembro"),
        )

        if clave in vistos:
            continue

        vistos.add(clave)

        titulo = obtener_titulo_interno_h5p(
            recurso
        )

        if not titulo:
            continue

        titulo_n = normalizar_texto(
            titulo
        )

        print(
            "Título H5P interno:",
            recurso.get("nombre", ""),
            "->",
            titulo,
        )

        if titulo_n == actividad_n:
            coincidencias.append(
                recurso.copy()
            )

    if len(coincidencias) != 1:
        return None

    elegido = coincidencias[0]
    elegido["puntuacion"] = 1000
    elegido["metodo"] = "TITULO_INTERNO_H5P_EXACTO"

    return elegido


# ============================================================
# TIPO DE ARCHIVO
# ============================================================

def determinar_tipo_archivo(
    categoria,
    tipo_recurso,
    enlace,
):

    categoria_n = normalizar_texto(
        categoria
    )

    tipo_n = normalizar_texto(
        tipo_recurso
    )

    enlace_n = normalizar_texto(
        enlace
    )

    # OVI / OVA / RETO EVALUATIVO
    if categoria_n not in [
        "ovi",
        "ova",
        "challenge",
    ]:
        return None

    # RETO EVALUATIVO SIEMPRE PDF
    if (
        categoria_n == "challenge"
        and tipo_n == "reto evaluativo"
    ):
        return "PDF"

    # VIDEOS EXCLUIDOS
    if tipo_n in [
        "video intro",
        "video cierre",
        "video a camara",
    ]:
        return None

    # OVA SIEMPRE H5P
    if categoria_n == "ova":
        return "H5P"

    # OVI PDF
    if "pdf" in tipo_n:
        return "PDF"

    if "pdf" in enlace_n:
        return "PDF"

    # OVI H5P
    if "h5p" in tipo_n:
        return "H5P"

    if "h5p" in enlace_n:
        return "H5P"

    # INFOGRAFÍA OVI
    if "infografia" in tipo_n:
        return "H5P"

    return None


# ============================================================
# VALIDACIÓN DE CARPETAS DE RECURSOS EN GOOGLE DRIVE
# ============================================================

_CACHE_CARPETAS_DRIVE = {}
_CACHE_CARPETAS_DRIVE_LOCK = threading.Lock()


def extraer_url_celda_recurso(celda_valor, celda_formula):
    """Obtiene el hipervínculo real conservado en el XLSX exportado por Sheets."""
    for celda in (celda_formula, celda_valor):
        try:
            if celda is not None and celda.hyperlink and celda.hyperlink.target:
                return str(celda.hyperlink.target).strip()
        except Exception:
            pass

    valor_formula = ""
    try:
        valor_formula = str(celda_formula.value or "").strip()
    except Exception:
        pass

    coincidencia = re.search(
        r'^\s*=\s*HYPERLINK\(\s*["\']([^"\']+)["\']',
        valor_formula,
        flags=re.IGNORECASE,
    )
    if coincidencia:
        return coincidencia.group(1).strip()

    for celda in (celda_valor, celda_formula):
        try:
            valor = str(celda.value or "").strip()
        except Exception:
            valor = ""
        if valor.lower().startswith(("https://", "http://")):
            return valor

    return ""


def extraer_id_carpeta_drive(url):
    url = str(url or "").strip()
    coincidencia = re.search(
        r"drive\.google\.com/(?:drive/(?:u/\d+/)?folders|folders)/([A-Za-z0-9_-]+)",
        url,
        flags=re.IGNORECASE,
    )
    if not coincidencia:
        return None
    return coincidencia.group(1)


def _limpiar_html_drive(texto):
    texto = html_lib.unescape(str(texto or ""))
    reemplazos = {
        r"\u0026": "&",
        r"\u003d": "=",
        r"\u003c": "<",
        r"\u003e": ">",
        r"\u0027": "'",
        r"\u0022": '"',
    }
    for origen, destino in reemplazos.items():
        texto = texto.replace(origen, destino)
    return texto


def _extraer_nombres_archivos_drive(html_drive):
    texto = _limpiar_html_drive(html_drive)
    nombres = []
    vistos = set()

    # Vista embebida: los nombres suelen aparecer como texto de enlaces.
    for contenido in re.findall(
        r"<a[^>]*>(.*?)</a>",
        texto,
        flags=re.IGNORECASE | re.DOTALL,
    ):
        limpio = re.sub(r"<[^>]+>", " ", contenido)
        limpio = html_lib.unescape(limpio)
        limpio = " ".join(limpio.split()).strip()
        if not limpio or "." not in limpio:
            continue
        if len(limpio) > 260:
            continue
        clave = limpio.casefold()
        if clave not in vistos:
            vistos.add(clave)
            nombres.append(limpio)

    # Fallback para nombres serializados en JSON/HTML.
    for patron in [
        r'["\']([^"\'<>\\]{1,250}\.(?:h5p|pdf|zip|docx?|xlsx?|pptx?|webm|mp4))["\']',
        r'>([^<>]{1,250}\.(?:h5p|pdf|zip|docx?|xlsx?|pptx?|webm|mp4))<',
    ]:
        for nombre in re.findall(patron, texto, flags=re.IGNORECASE):
            nombre = html_lib.unescape(" ".join(nombre.split()).strip())
            clave = nombre.casefold()
            if nombre and clave not in vistos:
                vistos.add(clave)
                nombres.append(nombre)

    return nombres


def listar_archivos_carpeta_drive(url_carpeta):
    carpeta_id = extraer_id_carpeta_drive(url_carpeta)
    if not carpeta_id:
        return {
            "ok": False,
            "error": "ERROR_CARPETA_RECURSO_NO_ACCESIBLE",
            "archivos": [],
        }

    with _CACHE_CARPETAS_DRIVE_LOCK:
        cache = _CACHE_CARPETAS_DRIVE.get(carpeta_id)
    if cache is not None:
        return cache

    urls = [
        "https://drive.google.com/embeddedfolderview?id="
        + quote(carpeta_id)
        + "#list",
        "https://drive.google.com/drive/folders/"
        + quote(carpeta_id),
    ]

    ultimo_error = None
    for url in urls:
        solicitud = Request(
            url,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 Chrome/124 Safari/537.36"
                )
            },
        )
        try:
            with urlopen(solicitud, timeout=30) as respuesta:
                url_final = str(respuesta.geturl() or "")
                datos = respuesta.read(8 * 1024 * 1024)
        except (HTTPError, URLError, TimeoutError, OSError) as error:
            ultimo_error = error
            continue

        if "accounts.google.com" in url_final.lower():
            ultimo_error = RuntimeError("La carpeta requiere autenticación de Google")
            continue

        texto = datos.decode("utf-8", errors="ignore")
        archivos = _extraer_nombres_archivos_drive(texto)

        resultado = {
            "ok": True,
            "error": None,
            "archivos": archivos,
        }
        with _CACHE_CARPETAS_DRIVE_LOCK:
            if len(_CACHE_CARPETAS_DRIVE) > 500:
                _CACHE_CARPETAS_DRIVE.clear()
            _CACHE_CARPETAS_DRIVE[carpeta_id] = resultado
        return resultado

    print(
        "No se pudo consultar carpeta Drive:",
        carpeta_id,
        type(ultimo_error).__name__ if ultimo_error else "Error",
        str(ultimo_error or ""),
    )
    return {
        "ok": False,
        "error": "ERROR_CARPETA_RECURSO_NO_ACCESIBLE",
        "archivos": [],
    }


def validar_carpeta_drive_para_actividad(actividad):
    """
    Si la matriz contiene un hipervínculo a una carpeta Drive, comprueba:
    - que sea accesible;
    - que no esté vacía;
    - que contenga un H5P/PDF del tipo esperado que corresponda a la actividad.

    Si la celda no contiene una carpeta Drive, no altera el flujo histórico del ZIP.
    """
    url_recurso = str(actividad.get("url_recurso") or "").strip()
    if not url_recurso:
        return None

    if not extraer_id_carpeta_drive(url_recurso):
        return None

    resultado = listar_archivos_carpeta_drive(url_recurso)
    if not resultado.get("ok"):
        return resultado.get("error") or "ERROR_CARPETA_RECURSO_NO_ACCESIBLE"

    archivos = resultado.get("archivos") or []
    if not archivos:
        return "ERROR_CARPETA_RECURSO_VACIA"

    extension_objetivo = (
        ".h5p"
        if actividad.get("tipo_archivo") == "H5P"
        else ".pdf"
        if actividad.get("tipo_archivo") == "PDF"
        else ""
    )

    compatibles = [
        nombre
        for nombre in archivos
        if os.path.splitext(nombre)[1].lower() == extension_objetivo
    ]

    if not compatibles:
        return "ERROR_RECURSO_NO_COINCIDE_CON_ACTIVIDAD"

    coincidencias = []
    for nombre in compatibles:
        puntuacion = puntuar_recurso(nombre, actividad)
        if puntuacion >= 100 or recurso_es_prefijo_de_actividad(nombre, actividad):
            coincidencias.append(nombre)

    if not coincidencias:
        return "ERROR_RECURSO_NO_COINCIDE_CON_ACTIVIDAD"

    if len(coincidencias) != 1:
        return "ERROR_RECURSO_AMBIGUO_EN_CARPETA_DRIVE"

    nombre_drive = coincidencias[0]
    actividad["nombre_recurso_drive"] = nombre_drive

    print(
        "✅ Carpeta Drive validada:",
        actividad["nombre"],
        "->",
        nombre_drive,
    )
    return None


# ============================================================
# LEER ACTIVIDADES DEL EXCEL
# ============================================================

def leer_actividades_excel(
    ruta_excel,
    procesar_ovi=True,
    procesar_ova=True,
    procesar_retos=True,
):

    libro = load_workbook(
        ruta_excel,
        data_only=True,
    )
    libro_formulas = load_workbook(
        ruta_excel,
        data_only=False,
    )

    actividades = []

    for nombre_hoja in libro.sheetnames:

        hoja = libro[
            nombre_hoja
        ]
        hoja_formulas = libro_formulas[
            nombre_hoja
        ]

        fila_cabecera = None

        # ----------------------------------------------------
        # BUSCAR CABECERA
        # ----------------------------------------------------

        for fila in range(
            1,
            min(
                hoja.max_row,
                30,
            ) + 1,
        ):

            valor = hoja.cell(
                row=fila,
                column=1,
            ).value

            if normalizar_texto(
                valor
            ) == normalizar_texto(
                "Semana correspondiente"
            ):

                fila_cabecera = fila
                break

        if fila_cabecera is None:
            continue

        programa = hoja.cell(
            row=1,
            column=1,
        ).value

        curso = hoja.cell(
            row=2,
            column=1,
        ).value

        programa = (
            str(programa).strip()
            if programa
            else ""
        )

        curso = (
            str(curso).strip()
            if curso
            else nombre_hoja
        )

        # ----------------------------------------------------
        # ACTIVIDADES
        # ----------------------------------------------------

        for fila in range(
            fila_cabecera + 1,
            hoja.max_row + 1,
        ):

            semana = hoja.cell(
                row=fila,
                column=1,
            ).value

            unidad = hoja.cell(
                row=fila,
                column=2,
            ).value

            nombre = hoja.cell(
                row=fila,
                column=4,
            ).value

            categoria = hoja.cell(
                row=fila,
                column=5,
            ).value

            tipo_recurso = hoja.cell(
                row=fila,
                column=6,
            ).value

            celda_enlace = hoja.cell(
                row=fila,
                column=7,
            )
            celda_enlace_formula = hoja_formulas.cell(
                row=fila,
                column=7,
            )
            enlace = celda_enlace.value
            url_recurso = extraer_url_celda_recurso(
                celda_enlace,
                celda_enlace_formula,
            )

            if not nombre:
                continue

            if not categoria:
                continue

            categoria_prizma = normalizar_categoria(
                categoria
            )

            if categoria_prizma not in [
                "OVI",
                "OVA",
                "CHALLENGE",
            ]:
                continue

            if (
                categoria_prizma == "OVI"
                and not procesar_ovi
            ):
                continue

            if (
                categoria_prizma == "OVA"
                and not procesar_ova
            ):
                continue

            if (
                categoria_prizma == "CHALLENGE"
                and not procesar_retos
            ):
                continue

            tipo_archivo = determinar_tipo_archivo(
                categoria,
                tipo_recurso,
                enlace,
            )

            if tipo_archivo is None:
                continue

            actividades.append(
                {
                    "fila_excel": fila,
                    "hoja": nombre_hoja,
                    "programa": programa,
                    "curso": curso,

                    "semana": (
                        str(semana).strip()
                        if semana
                        else ""
                    ),

                    "unidad": (
                        str(unidad).strip()
                        if unidad
                        else ""
                    ),

                    "nombre": str(
                        nombre
                    ).strip(),

                    "categoria_prizma":
                        categoria_prizma,

                    "tipo_recurso": (
                        str(tipo_recurso).strip()
                        if tipo_recurso
                        else ""
                    ),

                    "tipo_archivo":
                        tipo_archivo,

                    "enlace": (
                        str(enlace).strip()
                        if enlace
                        else ""
                    ),

                    "url_recurso": url_recurso,
                }
            )

    return actividades


# ============================================================
# ÍNDICE DE RECURSOS
# ============================================================

def crear_indice_recursos(
    ruta_zip,
):

    indice = []

    with zipfile.ZipFile(
        ruta_zip,
        "r",
    ) as zip_ref:

        for miembro in zip_ref.namelist():

            if miembro.endswith("/"):
                continue

            nombre = os.path.basename(
                miembro
            )

            extension = os.path.splitext(
                nombre
            )[1].lower()

            if extension not in [
                ".h5p",
                ".pdf",
            ]:
                continue

            indice.append(
                {
                    "nombre": nombre,
                    "extension": extension,
                    "zip": ruta_zip,
                    "miembro": miembro,
                }
            )

    return indice


# ============================================================
# PUNTUAR RECURSO
# ============================================================

def puntuar_recurso(
    nombre_archivo,
    actividad,
):

    nombre_sin_extension = os.path.splitext(
        os.path.basename(
            nombre_archivo
        )
    )[0]

    archivo_n = normalizar_texto(
        nombre_sin_extension
    )

    actividad_n = normalizar_texto(
        actividad["nombre"]
    )

    referencia_n = normalizar_texto(
        actividad["enlace"]
    )

    puntuacion = 0

    if archivo_n == actividad_n:
        puntuacion += 300

    if (
        actividad_n
        and actividad_n in archivo_n
    ):
        puntuacion += 180

    if (
        archivo_n
        and archivo_n in referencia_n
    ):
        puntuacion += 80

    if (
        referencia_n
        and referencia_n in archivo_n
    ):
        puntuacion += 120

    palabras_actividad = {
        palabra
        for palabra in actividad_n.split()
        if len(palabra) >= 4
    }

    palabras_archivo = set(
        archivo_n.split()
    )

    if palabras_actividad:

        comunes = (
            palabras_actividad
            & palabras_archivo
        )

        porcentaje = (
            len(comunes)
            / len(palabras_actividad)
        )

        puntuacion += int(
            porcentaje * 100
        )

    return puntuacion


# ============================================================
# GOOGLE DRIVE -> ZIP: VALIDACIÓN EXACTA ANTES DE PRIZMA
# ============================================================

def clave_nombre_archivo_exacto(nombre):
    return unicodedata.normalize(
        "NFC",
        os.path.basename(str(nombre or "")).strip(),
    ).casefold()


def prevalidar_recursos_drive_zip(
    ruta_excel,
    ruta_zip,
    procesar_ovi=True,
    procesar_ova=True,
    procesar_retos=True,
):
    """Valida todos los enlaces Drive y exige el mismo archivo dentro del ZIP."""

    actividades = leer_actividades_excel(
        ruta_excel,
        procesar_ovi,
        procesar_ova,
        procesar_retos,
    )
    indice_recursos = crear_indice_recursos(ruta_zip)

    errores = []
    recursos = []

    if not actividades:
        return {
            "ok": False,
            "total_actividades": 0,
            "validadas_drive": 0,
            "errores": ["No se encontraron actividades compatibles en la matriz."],
            "recursos": [],
        }

    if not indice_recursos:
        return {
            "ok": False,
            "total_actividades": len(actividades),
            "validadas_drive": 0,
            "errores": ["El ZIP no contiene archivos H5P o PDF."],
            "recursos": [],
        }

    for actividad in actividades:
        descripcion = (
            f'Hoja {actividad.get("hoja", "")} · '
            f'fila {actividad.get("fila_excel", "")} · '
            f'{actividad.get("nombre", "")}'
        )

        url_recurso = str(actividad.get("url_recurso") or "").strip()

        if not url_recurso:
            errores.append(descripcion + ": no tiene enlace de Google Drive.")
            continue

        if not extraer_id_carpeta_drive(url_recurso):
            errores.append(
                descripcion
                + ": el enlace no corresponde a una carpeta válida de Google Drive."
            )
            continue

        error_drive = validar_carpeta_drive_para_actividad(actividad)
        if error_drive:
            errores.append(descripcion + ": " + str(error_drive))
            continue

        nombre_drive = str(actividad.get("nombre_recurso_drive") or "").strip()
        if not nombre_drive:
            errores.append(
                descripcion + ": Google Drive no permitió determinar un archivo único."
            )
            continue

        clave_drive = clave_nombre_archivo_exacto(nombre_drive)
        exactos = [
            recurso
            for recurso in indice_recursos
            if clave_nombre_archivo_exacto(recurso.get("nombre")) == clave_drive
        ]

        if len(exactos) == 0:
            errores.append(
                descripcion
                + f': Drive indica "{nombre_drive}", pero ese archivo no está en el ZIP.'
            )
            continue

        if len(exactos) != 1:
            errores.append(
                descripcion
                + f': el archivo "{nombre_drive}" aparece más de una vez en el ZIP.'
            )
            continue

        recursos.append({
            "hoja": actividad.get("hoja", ""),
            "fila": actividad.get("fila_excel", 0),
            "actividad": actividad.get("nombre", ""),
            "archivo": nombre_drive,
        })

    return {
        "ok": len(errores) == 0 and len(recursos) == len(actividades),
        "total_actividades": len(actividades),
        "validadas_drive": len(recursos),
        "errores": errores,
        "recursos": recursos,
    }


# ============================================================
# RESOLVER RECURSO
# ============================================================

def resolver_recurso(
    actividad,
    indice_recursos,
    carpeta_temp,
    actividades_curso,
):

    print()
    print("--------------------------------------")
    print("BUSCANDO RECURSO")
    print("--------------------------------------")

    print(
        "Actividad:",
        actividad["nombre"],
    )

    print(
        "Tipo esperado:",
        actividad["tipo_archivo"],
    )

    print(
        "Referencia:",
        actividad["enlace"],
    )

    if actividad[
        "tipo_archivo"
    ] == "H5P":

        extension = ".h5p"

    elif actividad[
        "tipo_archivo"
    ] == "PDF":

        extension = ".pdf"

    else:

        return (
            None,
            "ERROR_TIPO_RECURSO_NO_SOPORTADO",
        )

    candidatos = []

    nombre_drive = str(
        actividad.get("nombre_recurso_drive")
        or ""
    ).strip()

    if nombre_drive:
        clave_drive = clave_nombre_archivo_exacto(nombre_drive)

        for recurso in indice_recursos:
            if recurso["extension"] != extension:
                continue

            if clave_nombre_archivo_exacto(recurso["nombre"]) != clave_drive:
                continue

            candidato = recurso.copy()
            candidato["puntuacion"] = 10000
            candidato["metodo"] = "DRIVE_EXACTO"
            candidatos.append(candidato)

        if len(candidatos) == 0:
            return (
                None,
                "ERROR_RECURSO_DRIVE_NO_ESTA_EN_ZIP",
            )

        if len(candidatos) != 1:
            return (
                None,
                "ERROR_RECURSO_DUPLICADO",
            )

        print(
            "Coincidencia exacta Drive -> ZIP:",
            nombre_drive,
        )

    else:
        for recurso in indice_recursos:

            if recurso[
                "extension"
            ] != extension:
                continue

            puntuacion = puntuar_recurso(
                recurso["nombre"],
                actividad,
            )

            if puntuacion < 100:
                continue

            candidato = recurso.copy()

            candidato[
                "puntuacion"
            ] = puntuacion

            candidatos.append(
                candidato
            )

    print(
        "Candidatos:",
        len(candidatos),
    )

    # --------------------------------------------------------
    # SEGUNDO MÉTODO SEGURO PARA NOMBRES CORTOS / TRUNCADOS
    # --------------------------------------------------------
    # Se usa únicamente cuando el método normal no encontró
    # ningún candidato con puntuación >= 100.
    #
    # Ejemplos admitidos:
    # "Dimensiones.h5p"
    # "Dimensiones e indicadores.h5p"
    # "... principios de legalid.h5p"
    #
    # No adivina: el prefijo del archivo debe identificar una
    # única actividad del curso para ese tipo de archivo.
    # --------------------------------------------------------

    if not candidatos:

        candidatos_cortos = []
        hubo_ambiguos = False
        recursos_ambiguos = []

        for recurso in indice_recursos:

            if recurso[
                "extension"
            ] != extension:
                continue

            if not recurso_es_prefijo_de_actividad(
                recurso["nombre"],
                actividad,
            ):
                continue

            if not recurso_corto_es_unico_para_actividad(
                recurso["nombre"],
                actividad,
                actividades_curso,
            ):

                hubo_ambiguos = True
                recursos_ambiguos.append(
                    recurso
                )
                continue

            candidato = recurso.copy()

            candidato[
                "puntuacion"
            ] = puntuar_recurso(
                recurso["nombre"],
                actividad,
            )

            candidato[
                "metodo"
            ] = "PREFIJO_UNICO"

            candidatos_cortos.append(
                candidato
            )

        print(
            "Candidatos por prefijo único:",
            len(candidatos_cortos),
        )

        for candidato in candidatos_cortos:

            print(
                "  →",
                candidato["nombre"],
                "| PREFIJO_UNICO",
            )

        if len(candidatos_cortos) == 0:

            if hubo_ambiguos:

                # Fallback excepcional SOLO para H5P ambiguos.
                # No se ejecuta para duplicados y no recorre todos los H5P
                # del curso: únicamente abre los recursos que ya quedaron
                # implicados en la ambigüedad del nombre externo.
                if extension == ".h5p":

                    candidato_interno = (
                        resolver_h5p_ambiguo_por_titulo_interno(
                            actividad,
                            recursos_ambiguos,
                        )
                    )

                    if candidato_interno is not None:
                        candidatos = [
                            candidato_interno
                        ]
                    else:
                        return (
                            None,
                            "ERROR_RECURSO_AMBIGUO",
                        )

                else:
                    return (
                        None,
                        "ERROR_RECURSO_AMBIGUO",
                    )

            else:
                return (
                    None,
                    "ERROR_RECURSO_NO_ENCONTRADO",
                )

        if candidatos:
            pass
        elif len(candidatos_cortos) > 1:

            return (
                None,
                "ERROR_RECURSO_DUPLICADO",
            )
        else:
            candidatos = candidatos_cortos

        # Si el fallback interno ya resolvió la ambigüedad, no reemplazarlo.
        if candidatos and candidatos[0].get("metodo") == "TITULO_INTERNO_H5P_EXACTO":
            pass
        elif len(candidatos_cortos) > 1:
            return (
                None,
                "ERROR_RECURSO_DUPLICADO",
            )
        elif candidatos_cortos:
            candidatos = candidatos_cortos

    mejor_puntuacion = max(
        candidato["puntuacion"]
        for candidato in candidatos
    )

    mejores = [
        candidato
        for candidato in candidatos
        if candidato["puntuacion"]
        == mejor_puntuacion
    ]

    print(
        "Mejor puntuación:",
        mejor_puntuacion,
    )

    for candidato in mejores:

        print(
            "  →",
            candidato["nombre"],
            "|",
            candidato["puntuacion"],
        )

    # SEGURIDAD: NO ADIVINAR
    if len(mejores) != 1:

        return (
            None,
            "ERROR_RECURSO_DUPLICADO",
        )

    elegido = mejores[0]

    carpeta_fila = os.path.join(
        carpeta_temp,
        "fila_" + str(
            actividad["fila_excel"]
        ),
    )

    os.makedirs(
        carpeta_fila,
        exist_ok=True,
    )

    ruta_destino = os.path.join(
        carpeta_fila,
        elegido["nombre"],
    )

    try:

        with zipfile.ZipFile(
            elegido["zip"],
            "r",
        ) as zip_ref:

            with zip_ref.open(
                elegido["miembro"]
            ) as origen:

                with open(
                    ruta_destino,
                    "wb",
                ) as destino:

                    shutil.copyfileobj(
                        origen,
                        destino,
                    )

    except Exception:

        return (
            None,
            "ERROR_EXTRAYENDO_RECURSO",
        )

    if not os.path.isfile(
        ruta_destino
    ):

        return (
            None,
            "ERROR_EXTRAYENDO_RECURSO",
        )

    print()
    print(
        "✅ Recurso seleccionado:",
        elegido["nombre"],
    )

    return (
        {
            "ruta":
                ruta_destino,

            "nombre_original":
                elegido["nombre"],

            "puntuacion":
                mejor_puntuacion,
        },
        None,
    )


# ============================================================
# REPORTE
# ============================================================

def guardar_resultado(
    ruta_reporte,
    actividad,
    resultado,
    observacion,
    recurso="",
):

    existe = os.path.isfile(
        ruta_reporte
    )

    with open(
        ruta_reporte,
        "a",
        newline="",
        encoding="utf-8-sig",
    ) as archivo:

        escritor = csv.writer(
            archivo
        )

        if not existe:

            escritor.writerow(
                [
                    "Fila Excel",
                    "Programa",
                    "Curso",
                    "Semana",
                    "Unidad",
                    "Actividad",
                    "Categoría",
                    "Tipo",
                    "Recurso",
                    "Resultado",
                    "Observación",
                ]
            )

        escritor.writerow(
            [
                actividad["fila_excel"],
                actividad["programa"],
                actividad["curso"],
                actividad["semana"],
                actividad["unidad"],
                actividad["nombre"],
                actividad["categoria_prizma"],
                actividad["tipo_archivo"],
                recurso,
                resultado,
                observacion,
            ]
        )


# ============================================================
# OVERLAY POST-GUARDADO
# ============================================================

def desbloquear_interfaz_post_guardado(
    pagina,
):
    """Cierra el overlay solo si realmente existe y esta visible.

    Antes esta funcion agregaba 1.1 s de espera y un clic ciego incluso
    cuando no habia overlay. En el camino normal ahora retorna de inmediato.
    """
    overlay = pagina.locator(
        "div.MuiBox-root.css-15m6u24"
    )

    try:
        for i in range(overlay.count()):
            elemento = overlay.nth(i)
            if not elemento.is_visible():
                continue

            print("✅ Overlay detectado.")

            try:
                elemento.click(
                    position={"x": 5, "y": 5},
                    force=True,
                )
            except Exception:
                pagina.mouse.click(720, 450)

            # Pequena pausa solo despues de haber cerrado un overlay real.
            pagina.wait_for_timeout(120)
            return True
    except Exception:
        pass

    return True


# ============================================================
# CANCELAR EDICIÓN
# ============================================================

def cancelar_edicion_segura(
    pagina,
):

    try:

        boton = pagina.get_by_role(
            "button",
            name="Cancelar",
            exact=True,
        )

        if boton.count() == 0:
            return False

        if not boton.first.is_visible():
            return False

        boton.first.click(
            timeout=10000
        )

        pagina.wait_for_timeout(
            900
        )

        desbloquear_interfaz_post_guardado(
            pagina
        )

        return True

    except Exception:

        return False


# ============================================================
# ASEGURAR LISTADO
# ============================================================

def asegurar_listado(
    pagina,
):

    desbloquear_interfaz_post_guardado(
        pagina
    )

    buscador = pagina.locator(
        'input[placeholder="Buscar..."]'
    )

    try:

        if (
            buscador.count() > 0
            and buscador.first.is_visible()
        ):
            return True

    except Exception:
        pass

    try:

        cancelar = pagina.get_by_role(
            "button",
            name="Cancelar",
            exact=True,
        )

        if (
            cancelar.count() > 0
            and cancelar.first.is_visible()
        ):

            cancelar.first.click(
                timeout=10000
            )

            pagina.wait_for_timeout(
                1000
            )

            desbloquear_interfaz_post_guardado(
                pagina
            )

    except Exception:
        pass

    try:

        buscador.wait_for(
            state="visible",
            timeout=10000,
        )

        return True

    except Exception:

        return False


# ============================================================
# ENTRAR A CATEGORÍA
# ============================================================

def entrar_categoria(
    pagina,
    categoria,
    estado_navegacion=None,
):

    nombres_pestana = {
        "OVI": "OVI",
        "OVA": "OVA",
        "CHALLENGE": "Retos Evaluativos",
    }

    texto_pestana = nombres_pestana.get(
        categoria
    )

    if not texto_pestana:
        return False

    if (
        isinstance(estado_navegacion, dict)
        and estado_navegacion.get("categoria_actual") == categoria
    ):
        return True

    try:

        desbloquear_interfaz_post_guardado(
            pagina
        )

        pestana = pagina.locator(
            "label.tab",
            has_text=texto_pestana,
        ).first

        pestana.wait_for(
            state="visible",
            timeout=30000,
        )

        pestana.click(
            timeout=10000
        )

        pagina.wait_for_timeout(
            1200
        )

        if isinstance(estado_navegacion, dict):
            estado_navegacion["categoria_actual"] = categoria

        return True

    except Exception:

        return False


# ============================================================
# ENCONTRAR FILA
# ============================================================

def encontrar_fila_desde_elemento(
    elemento,
):

    for nivel in range(
        1,
        10,
    ):

        try:

            padre = elemento.locator(
                "xpath=" + "/.." * nivel
            )

            if padre.locator(
                "button"
            ).count() == 3:

                return padre

        except Exception:
            pass

    return None


def obtener_id_temporal_fila(
    fila,
):

    try:

        return fila.evaluate(
            """
            (element) => {

                if (!element.dataset.prizmaScanId) {

                    element.dataset.prizmaScanId =
                        "scan-" +
                        Math.random()
                        .toString(36)
                        .slice(2);
                }

                return element.dataset.prizmaScanId;
            }
            """
        )

    except Exception:

        return None


def obtener_nombre_fila(
    texto,
):

    lineas = [
        linea.strip()
        for linea in texto.splitlines()
        if linea.strip()
    ]

    if len(lineas) < 2:
        return ""

    return lineas[1]


# ============================================================
# ANALIZAR RESULTADOS
# ============================================================

def analizar_resultados_pagina(
    pagina,
    actividad,
):

    elementos = pagina.get_by_text(
        actividad["nombre"],
        exact=False,
    )

    filas = {}

    for i in range(
        elementos.count()
    ):

        try:

            elemento = elementos.nth(i)

            if not elemento.is_visible():
                continue

            fila = encontrar_fila_desde_elemento(
                elemento
            )

            if fila is None:
                continue

            identificador = obtener_id_temporal_fila(
                fila
            )

            if identificador:

                filas[
                    identificador
                ] = fila

        except Exception:
            pass

    nombre_objetivo = normalizar_texto(
        actividad["nombre"]
    )

    semana_objetivo = normalizar_texto(
        actividad["semana"]
    )

    unidad_objetivo = normalizar_texto(
        actividad["unidad"]
    )

    variantes_programa = obtener_variantes_programa(
        actividad["programa"]
    )

    resultados = []

    for fila in filas.values():

        try:

            texto = fila.inner_text().strip()

        except Exception:
            continue

        texto_n = normalizar_texto(
            texto
        )

        nombre_fila = normalizar_texto(
            obtener_nombre_fila(
                texto
            )
        )

        cumple_nombre = (
            nombre_fila
            == nombre_objetivo
        )

        cumple_semana = (
            semana_objetivo
            in texto_n
        )

        cumple_unidad = (
            unidad_objetivo
            in texto_n
        )

        cumple_programa = any(
            variante
            and variante in texto_n
            for variante in variantes_programa
        )

        if actividad[
            "categoria_prizma"
        ] == "OVI":

            cumple_categoria = (
                "ovi" in texto_n
            )

        elif actividad[
            "categoria_prizma"
        ] == "OVA":

            cumple_categoria = (
                "ova" in texto_n
            )

        elif actividad[
            "categoria_prizma"
        ] == "CHALLENGE":

            # En Retos Evaluativos la fila no muestra la categoría
            # dentro de su texto. La categoría ya quedó validada al
            # entrar explícitamente a la pestaña Retos Evaluativos.
            cumple_categoria = True

        else:

            cumple_categoria = False

        coincide = (
            cumple_nombre
            and cumple_semana
            and cumple_unidad
            and cumple_programa
            and cumple_categoria
        )

        resultados.append(
            {
                "fila": fila,
                "texto": texto,
                "coincide": coincide,
            }
        )

    return resultados


def obtener_firma_pagina(
    pagina,
    actividad,
):

    resultados = analizar_resultados_pagina(
        pagina,
        actividad,
    )

    return "||".join(
        sorted(
            normalizar_texto(
                resultado["texto"]
            )
            for resultado in resultados
        )
    )


# ============================================================
# PAGINACIÓN
# ============================================================

def obtener_paginas_numericas(
    pagina,
):

    numeros = set()

    botones = pagina.locator(
        "button"
    )

    for i in range(
        botones.count()
    ):

        try:

            boton = botones.nth(i)

            if not boton.is_visible():
                continue

            texto = boton.inner_text().strip()

            if texto.isdigit():

                numero = int(
                    texto
                )

                if 1 <= numero <= 200:

                    numeros.add(
                        numero
                    )

            aria = (
                boton.get_attribute(
                    "aria-label"
                )
                or ""
            )

            coincidencia = re.search(
                r"(?:page|página|pagina)\s*(\d+)",
                aria,
                flags=re.IGNORECASE,
            )

            if coincidencia:

                numero = int(
                    coincidencia.group(1)
                )

                if 1 <= numero <= 200:

                    numeros.add(
                        numero
                    )

        except Exception:
            pass

    return sorted(
        numeros
    )


def ir_a_pagina_numero(
    pagina,
    numero,
):

    botones = pagina.locator(
        "button"
    )

    # POR TEXTO
    for i in range(
        botones.count()
    ):

        try:

            boton = botones.nth(i)

            if not boton.is_visible():
                continue

            texto = boton.inner_text().strip()

            if texto != str(
                numero
            ):
                continue

            boton.click(
                timeout=5000
            )

            pagina.wait_for_timeout(
                1200
            )

            return True

        except Exception:
            pass

    # POR ARIA
    for i in range(
        botones.count()
    ):

        try:

            boton = botones.nth(i)

            if not boton.is_visible():
                continue

            aria = normalizar_texto(
                boton.get_attribute(
                    "aria-label"
                )
            )

            if (
                f"page {numero}" not in aria
                and
                f"pagina {numero}" not in aria
            ):
                continue

            boton.click(
                timeout=5000
            )

            pagina.wait_for_timeout(
                1200
            )

            return True

        except Exception:
            pass

    return False


def encontrar_boton_siguiente(
    pagina,
):

    botones = pagina.locator(
        "button"
    )

    for i in range(
        botones.count()
    ):

        try:

            boton = botones.nth(i)

            if not boton.is_visible():
                continue

            if not boton.is_enabled():
                continue

            texto = normalizar_texto(
                boton.inner_text()
            )

            aria = normalizar_texto(
                boton.get_attribute(
                    "aria-label"
                )
            )

            title = normalizar_texto(
                boton.get_attribute(
                    "title"
                )
            )

            combinado = (
                texto
                + " "
                + aria
                + " "
                + title
            )

            if any(
                valor in combinado
                for valor in [
                    "next page",
                    "pagina siguiente",
                    "siguiente",
                    "go to next page",
                ]
            ):

                return boton

        except Exception:
            pass

    return None


def asegurar_pagina_1(
    pagina,
):

    paginas = obtener_paginas_numericas(
        pagina
    )

    if 1 not in paginas:
        return True

    return ir_a_pagina_numero(
        pagina,
        1,
    )


# ============================================================
# TÉRMINOS DE BÚSQUEDA
# ============================================================

def obtener_terminos_busqueda_general(
    nombre,
):

    nombre_original = str(
        nombre or ""
    ).strip()

    if not nombre_original:
        return []

    terminos = [
        nombre_original
    ]

    nombre_n = normalizar_texto(
        nombre_original
    )

    palabras = [
        palabra
        for palabra in nombre_n.split()
        if len(palabra) >= 3
        and palabra not in {
            "del", "las", "los", "una", "uno",
            "para", "por", "con", "sin", "desde",
            "entre", "sobre", "sus", "que", "como",
        }
    ]

    # PRIZMA a veces no devuelve registros con el título completo,
    # pero sí con una frase corta del mismo nombre. La fila SOLO se
    # acepta después si nombre + semana + unidad + programa + categoría
    # coinciden exactamente, así que estos términos amplían la búsqueda
    # sin volverla ambigua.
    candidatos = []

    if len(palabras) >= 2:
        candidatos.append(
            " ".join(palabras[-2:])
        )

    if len(palabras) >= 4:
        medio = max(0, len(palabras) // 2 - 1)
        candidatos.append(
            " ".join(palabras[medio:medio + 2])
        )

    if len(palabras) >= 2:
        candidatos.append(
            " ".join(palabras[:2])
        )

    for candidato in candidatos:
        if (
            candidato
            and normalizar_texto(candidato) != nombre_n
            and candidato not in terminos
        ):
            terminos.append(candidato)

    return terminos


def obtener_terminos_busqueda_reto(
    nombre,
):

    nombre_original = str(
        nombre or ""
    ).strip()

    if not nombre_original:
        return []

    terminos = [
        nombre_original
    ]

    # Fallback específico para Retos Evaluativos. PRIZMA puede mostrar
    # la fila al repetir la búsqueda quitando una o dos letras finales.
    # La fila solo se acepta después con validación exacta de nombre,
    # semana, unidad, programa y categoría.
    for recorte in (1, 2):
        if len(nombre_original) > recorte + 4:
            truncado = nombre_original[:-recorte].rstrip()
            if truncado and truncado not in terminos:
                terminos.append(truncado)

    nombre_n = normalizar_texto(
        nombre_original
    )

    # Algunos Retos Evaluativos no aparecen cuando el buscador de
    # PRIZMA recibe el título completo. Como fallback usamos frases
    # más cortas, pero la aceptación de la fila sigue siendo EXACTA
    # en analizar_resultados_pagina().
    palabras = [
        palabra
        for palabra in nombre_n.split()
        if len(palabra) >= 4
        and palabra not in {
            "analisis",
            "modelo",
            "modelos",
            "sobre",
            "para",
            "entre",
            "desde",
        }
    ]

    if palabras:
        # Preferimos las últimas palabras porque suelen ser las más
        # distintivas del reto (p. ej. "modelo de Piaget").
        frase_final = " ".join(
            palabras[-3:]
        )

        if (
            frase_final
            and frase_final not in terminos
        ):
            terminos.append(
                frase_final
            )

        ultima = palabras[-1]

        if (
            len(ultima) >= 5
            and ultima not in terminos
        ):
            terminos.append(
                ultima
            )

    return terminos


# ============================================================
# ESPERAS ADAPTATIVAS DE RENDIMIENTO
# ============================================================

def _esperar_resultados_busqueda(
    pagina,
    actividad,
    timeout_ms=2600,
    intervalo_ms=100,
    firma_previa=None,
):
    """Espera hasta que aparezcan resultados y la firma quede estable.

    firma_previa es la pantalla que habia ANTES de aplicar el filtro.
    Mientras la firma siga siendo esa, el filtro todavia no se aplico y
    seguir adelante significaria leer datos viejos. Ese era el origen de
    ERROR_RECUPERANDO_FILA cuando PRIZMA respondia lento.

    Si no aparece ningun resultado conserva el timeout completo, evitando
    falsos negativos por una respuesta lenta de PRIZMA.
    """
    transcurrido = 0
    firma_anterior = None
    estables = 0

    while True:
        firma = obtener_firma_pagina(pagina, actividad)

        if firma and firma != firma_previa:
            if firma == firma_anterior:
                estables += 1
            else:
                estables = 0

            # Dos lecturas iguales seguidas: la tabla ya dejo de moverse.
            if estables >= 2:
                return firma

        firma_anterior = firma

        if transcurrido >= timeout_ms:
            return firma or ""

        espera = min(intervalo_ms, timeout_ms - transcurrido)
        if espera <= 0:
            return firma or ""

        pagina.wait_for_timeout(espera)
        transcurrido += espera


def _esperar_cambio_firma(
    pagina,
    actividad,
    firma_anterior,
    timeout_ms=1200,
    intervalo_ms=100,
):
    """Espera un cambio real de pagina sin pagar 1.2 s fijos."""
    transcurrido = 0

    while True:
        firma = obtener_firma_pagina(pagina, actividad)
        if firma != firma_anterior:
            return True

        if transcurrido >= timeout_ms:
            return False

        espera = min(intervalo_ms, timeout_ms - transcurrido)
        if espera <= 0:
            return False

        pagina.wait_for_timeout(espera)
        transcurrido += espera


def _recuperar_fila_con_filtro(
    pagina,
    actividad,
    buscador,
    termino,
    timeout_ms,
):
    """Vuelve a filtrar y recorre las paginas hasta ver la fila exacta.

    Devuelve el locator de la fila, la cadena "DUPLICADA" si aparece mas
    de una vez, o None si en este intento no se pudo encontrar.

    A diferencia de la version anterior, un fallo al cambiar de pagina no
    aborta todo el proceso: simplemente termina este intento y el que
    llama vuelve a probar.
    """
    try:
        buscador.fill("")
        pagina.wait_for_timeout(400)
        asegurar_pagina_1(pagina)

        # Pantalla SIN filtro. Sirve para saber cuando el filtro nuevo
        # ya se aplico de verdad y no estamos leyendo lo anterior.
        firma_sin_filtro = obtener_firma_pagina(
            pagina,
            actividad,
        )

        buscador.fill(termino)

        _esperar_resultados_busqueda(
            pagina,
            actividad,
            timeout_ms=timeout_ms,
            firma_previa=firma_sin_filtro,
        )

        asegurar_pagina_1(pagina)
    except Exception:
        return None

    firmas_vistas = set()
    numero_pagina = 1

    while numero_pagina <= 50:

        try:
            firma_actual = obtener_firma_pagina(
                pagina,
                actividad,
            )
        except Exception:
            return None

        if firma_actual in firmas_vistas:
            break

        firmas_vistas.add(firma_actual)

        try:
            resultados = analizar_resultados_pagina(
                pagina,
                actividad,
            )
        except Exception:
            return None

        finales = [
            resultado
            for resultado in resultados
            if resultado["coincide"]
        ]

        if len(finales) > 1:
            return "DUPLICADA"

        if len(finales) == 1:
            print("✅ Fila recuperada por escaneo.")
            return finales[0]["fila"]

        siguiente_numero = numero_pagina + 1

        try:
            paginas = obtener_paginas_numericas(pagina)
        except Exception:
            break

        avanzo = False

        if siguiente_numero in paginas:
            try:
                avanzo = ir_a_pagina_numero(
                    pagina,
                    siguiente_numero,
                )
            except Exception:
                avanzo = False

        if not avanzo:
            try:
                siguiente = encontrar_boton_siguiente(pagina)
            except Exception:
                siguiente = None

            if siguiente is None:
                break

            try:
                siguiente.click(timeout=5000)
                _esperar_cambio_firma(
                    pagina,
                    actividad,
                    firma_actual,
                    timeout_ms=1500,
                )
                avanzo = True
            except Exception:
                # No se pudo avanzar. No es un error definitivo:
                # este intento termina y el que llama reintenta.
                break

        if not avanzo:
            break

        try:
            firma_nueva = obtener_firma_pagina(
                pagina,
                actividad,
            )
        except Exception:
            break

        if firma_nueva == firma_actual:
            break

        numero_pagina += 1

    return None


# ============================================================
# BUSCAR ACTIVIDAD
# ============================================================

def buscar_actividad_correcta(
    pagina,
    actividad,
):

    print()
    print(
        "Buscando actividad:",
        actividad["nombre"],
    )

    print(
        "Programa:",
        actividad["programa"],
    )

    print(
        "Semana:",
        actividad["semana"],
    )

    print(
        "Unidad:",
        actividad["unidad"],
    )

    print(
        "Categoría:",
        actividad["categoria_prizma"],
    )

    buscador = pagina.locator(
        'input[placeholder="Buscar..."]'
    )

    buscador.wait_for(
        state="visible",
        timeout=30000,
    )

    if actividad[
        "categoria_prizma"
    ] == "CHALLENGE":

        terminos_busqueda = (
            obtener_terminos_busqueda_reto(
                actividad["nombre"]
            )
        )

    else:

        terminos_busqueda = (
            obtener_terminos_busqueda_general(
                actividad["nombre"]
            )
        )

    coincidencias = []
    claves_coincidencias = set()
    termino_encontrado = None

    for termino_busqueda in terminos_busqueda:

        print(
            "Término de búsqueda:",
            termino_busqueda,
        )

        buscador.fill("")

        pagina.wait_for_timeout(
            500
        )

        asegurar_pagina_1(
            pagina
        )

        buscador.fill(
            termino_busqueda
        )

        espera_busqueda = (
            4200
            if actividad["categoria_prizma"] == "CHALLENGE"
            else 2600
        )

        _esperar_resultados_busqueda(
            pagina,
            actividad,
            timeout_ms=espera_busqueda,
        )

        asegurar_pagina_1(
            pagina
        )

        coincidencias_intento = []
        firmas_visitadas = set()
        numero_logico = 1

        while numero_logico <= 50:

            firma_actual = obtener_firma_pagina(
                pagina,
                actividad,
            )

            if firma_actual in firmas_visitadas:
                break

            firmas_visitadas.add(
                firma_actual
            )

            resultados = analizar_resultados_pagina(
                pagina,
                actividad,
            )

            print(
                "Página",
                numero_logico,
                "- candidatos:",
                len(resultados),
            )

            for resultado in resultados:

                if resultado[
                    "coincide"
                ]:

                    clave = (
                        numero_logico,
                        normalizar_texto(
                            resultado["texto"]
                        ),
                    )

                    if clave not in claves_coincidencias:
                        claves_coincidencias.add(
                            clave
                        )
                        coincidencia = {
                            "pagina":
                                numero_logico,

                            "texto":
                                resultado["texto"],

                            "termino":
                                termino_busqueda,
                        }
                        coincidencias.append(
                            coincidencia
                        )
                        coincidencias_intento.append(
                            coincidencia
                        )

            siguiente_numero = (
                numero_logico + 1
            )

            paginas = obtener_paginas_numericas(
                pagina
            )

            if siguiente_numero in paginas:

                if ir_a_pagina_numero(
                    pagina,
                    siguiente_numero,
                ):

                    firma_nueva = obtener_firma_pagina(
                        pagina,
                        actividad,
                    )

                    if firma_nueva != firma_actual:

                        numero_logico += 1
                        continue

            siguiente = encontrar_boton_siguiente(
                pagina
            )

            if siguiente is None:
                break

            try:

                siguiente.click(
                    timeout=5000
                )

                _esperar_cambio_firma(
                    pagina,
                    actividad,
                    firma_actual,
                    timeout_ms=1200,
                )

            except Exception:
                break

            firma_nueva = obtener_firma_pagina(
                pagina,
                actividad,
            )

            if firma_nueva == firma_actual:
                break

            numero_logico += 1

        # Si el término actual ya encontró la coincidencia exacta,
        # no necesitamos ampliar la búsqueda del reto.
        if coincidencias_intento:
            termino_encontrado = termino_busqueda
            break

    print(
        "Coincidencias exactas:",
        len(coincidencias),
    )

    if len(coincidencias) == 0:

        return (
            None,
            "ERROR_ACTIVIDAD_NO_ENCONTRADA",
        )

    if len(coincidencias) > 1:

        return (
            None,
            "ERROR_ACTIVIDAD_DUPLICADA",
        )

    # Para recuperar la misma vista debemos conservar el término que
    # produjo la coincidencia, especialmente en el fallback de Retos.
    if termino_encontrado is None:
        termino_encontrado = coincidencias[
            0
        ].get(
            "termino",
            actividad["nombre"],
        )

    # --------------------------------------------------------
    # RECUPERAR FILA POR ESCANEO REAL
    # --------------------------------------------------------
    # Antes intentábamos regresar a un número lógico de página. En
    # Retos Evaluativos la paginación puede reordenarse o cambiar al
    # refrescar el filtro, provocando ERROR_RECUPERANDO_PAGINA/FILA.
    # Ahora repetimos el mismo filtro y recorremos las páginas hasta
    # volver a encontrar la coincidencia exacta. La validación sigue
    # siendo estricta: nombre + semana + unidad + programa + categoría.

    # La coincidencia YA quedo confirmada arriba. Aqui solo hay que volver
    # a dejar esa fila en pantalla, asi que vale la pena insistir: se
    # reintenta varias veces, con esperas cada vez mayores y probando
    # todos los terminos que produjeron resultados.
    terminos_recuperacion = [termino_encontrado]
    for termino in terminos_busqueda:
        if termino not in terminos_recuperacion:
            terminos_recuperacion.append(termino)

    espera_base = (
        4200
        if actividad["categoria_prizma"] == "CHALLENGE"
        else 2600
    )

    for intento in range(1, 4):

        espera_intento = espera_base + (intento - 1) * 2500

        for termino in terminos_recuperacion:

            resultado_recuperacion = _recuperar_fila_con_filtro(
                pagina,
                actividad,
                buscador,
                termino,
                espera_intento,
            )

            if resultado_recuperacion == "DUPLICADA":
                return (
                    None,
                    "ERROR_ACTIVIDAD_DUPLICADA",
                )

            if resultado_recuperacion is not None:
                return (
                    resultado_recuperacion,
                    None,
                )

        print(
            "Reintentando recuperacion de fila (intento",
            intento,
            "de 3)...",
        )

        # Una pausa antes del siguiente intento le da margen a PRIZMA
        # para terminar de responder.
        try:
            pagina.wait_for_timeout(700 * intento)
        except Exception:
            pass

    return (
        None,
        "ERROR_RECUPERANDO_FILA",
    )


# ============================================================
# ABRIR EDICIÓN
# ============================================================

def abrir_edicion(
    pagina,
    fila,
):

    botones = fila.locator(
        "button"
    )

    if botones.count() != 3:

        print(
            "❌ Cantidad inesperada de botones:",
            botones.count(),
        )

        return False

    try:

        botones.nth(1).click(
            timeout=10000
        )

    except PlaywrightTimeoutError:

        desbloquear_interfaz_post_guardado(
            pagina
        )

        try:

            botones.nth(1).click(
                timeout=10000
            )

        except Exception:
            return False

    except Exception:
        return False

    pagina.wait_for_timeout(
        1800
    )

    return True


# ============================================================
# CAMPO PRINCIPAL "RECURSO"
# ============================================================
#
# TODOS LOS ARCHIVOS SE CARGAN EN RECURSO.
#
# H5P -> RECURSO
# PDF -> RECURSO
#
# MATERIAL DESCARGABLE NO SE TOCA.
# ============================================================

def detectar_campo_recurso(
    pagina,
):

    archivos = pagina.locator(
        'input[type="file"]'
    )

    candidatos_recurso = []

    print()
    print(
        "Campos file detectados:",
        archivos.count(),
    )

    for i in range(
        archivos.count()
    ):

        archivo = archivos.nth(i)

        accept = (
            archivo.get_attribute(
                "accept"
            )
            or ""
        )

        accept_n = accept.lower()

        print()
        print(
            f"Campo #{i + 1}:"
        )

        print(
            "accept =",
            accept,
        )

        # El campo Recurso admite .h5p.
        if ".h5p" in accept_n:

            candidatos_recurso.append(
                archivo
            )

            print(
                "→ RECURSO PRINCIPAL"
            )

            continue

        # Material descargable.
        if accept_n.strip() == ".pdf":

            print(
                "→ MATERIAL DESCARGABLE - IGNORADO"
            )

            continue

        print(
            "→ CAMPO NO UTILIZADO"
        )

    if len(
        candidatos_recurso
    ) == 0:

        print(
            "❌ No se encontró RECURSO."
        )

        return None

    if len(
        candidatos_recurso
    ) > 1:

        print(
            "❌ Varios candidatos para RECURSO."
        )

        return None

    print()
    print(
        "✅ Campo RECURSO identificado."
    )

    return candidatos_recurso[0]


# ============================================================
# DESCRIPCIÓN
# ============================================================

def limpiar_descripcion(
    pagina,
):

    editores = pagina.locator(
        '[contenteditable="true"]'
    )

    objetivos = [
        normalizar_texto(
            texto
        )
        for texto in TEXTOS_DESCRIPCION_A_BORRAR
    ]

    exactos = []
    peligrosos = []

    for i in range(
        editores.count()
    ):

        try:

            editor = editores.nth(i)

            if not editor.is_visible():
                continue

            texto_original = (
                editor.inner_text().strip()
            )

            texto_n = normalizar_texto(
                texto_original
            )

            if texto_n in objetivos:

                exactos.append(
                    editor
                )

                continue

            for objetivo in objetivos:

                if (
                    objetivo
                    and objetivo in texto_n
                ):

                    peligrosos.append(
                        texto_original
                    )

        except Exception:
            pass

    if peligrosos:

        return (
            False,
            "ERROR_DESCRIPCION_CONTENIDO_ADICIONAL",
        )

    if len(exactos) == 0:

        return (
            True,
            None,
        )

    if len(exactos) > 1:

        return (
            False,
            "ERROR_DESCRIPCION_MULTIPLE",
        )

    editor = exactos[0]

    try:

        editor.click()

        editor.press(
            "Control+A"
        )

        editor.press(
            "Backspace"
        )

        pagina.wait_for_timeout(
            500
        )

        if normalizar_texto(
            editor.inner_text()
        ):

            return (
                False,
                "ERROR_LIMPIANDO_DESCRIPCION",
            )

        print(
            "✅ NO_DISPONIBLE eliminado."
        )

        return (
            True,
            None,
        )

    except Exception:

        return (
            False,
            "ERROR_LIMPIANDO_DESCRIPCION",
        )


# ============================================================
# RETOS EVALUATIVOS
# ============================================================

def es_programa_posgrado(
    programa,
):

    programa_n = normalizar_texto(
        programa
    )

    return (
        programa_n.startswith(
            "especializacion "
        )
        or programa_n == "especializacion"
        or programa_n.startswith(
            "esp "
        )
        or programa_n == "esp"
        or programa_n.startswith(
            "maestria "
        )
        or programa_n == "maestria"
    )


def obtener_numero_unidad(
    unidad,
):

    unidad_n = normalizar_texto(
        unidad
    )

    coincidencia = re.search(
        r"\b(\d+)\b",
        unidad_n,
    )

    if not coincidencia:
        return None

    return int(
        coincidencia.group(1)
    )


def obtener_configuracion_reto(
    actividad,
):

    numero_unidad = obtener_numero_unidad(
        actividad.get(
            "unidad",
            "",
        )
    )

    if numero_unidad is None:
        return (
            None,
            "ERROR_UNIDAD_RETO_NO_RECONOCIDA",
        )

    posgrado = es_programa_posgrado(
        actividad.get(
            "programa",
            "",
        )
    )

    if posgrado:

        reglas = {
            1: (
                "Corte 1",
                "Nivel intermedio",
            ),
            2: (
                "Corte 1",
                "Nivel intermedio",
            ),
            3: (
                "Corte 1",
                "Nivel avanzado",
            ),
            4: (
                "Corte 1",
                "Nivel avanzado",
            ),
        }

    else:

        reglas = {
            1: (
                "Corte 1",
                "Nivel básico",
            ),
            2: (
                "Corte 2",
                "Nivel intermedio",
            ),
            3: (
                "Corte 3",
                "Nivel avanzado",
            ),
        }

    configuracion = reglas.get(
        numero_unidad
    )

    if configuracion is None:

        if posgrado:
            return (
                None,
                "ERROR_UNIDAD_POSGRADO_NO_SOPORTADA",
            )

        return (
            None,
            "ERROR_UNIDAD_PREGRADO_NO_SOPORTADA",
        )

    return (
        {
            "corte": configuracion[0],
            "nivel": configuracion[1],
            "posgrado": posgrado,
            "numero_unidad": numero_unidad,
        },
        None,
    )


def seleccionar_autocomplete_reto(
    pagina,
    placeholder,
    valor,
):

    campo = pagina.locator(
        f'input[placeholder="{placeholder}"]'
    )

    if campo.count() != 1:
        return False

    campo = campo.first
    valor_n = normalizar_texto(
        valor
    )

    # Los autocompletes MUI de PRIZMA a veces tardan en montar sus
    # opciones. Hacemos hasta 3 intentos, pero solo aceptamos una
    # opción cuyo texto normalizado coincida EXACTAMENTE con el valor
    # esperado. Nunca elegimos la primera opción a ciegas.
    for intento in range(1, 4):

        try:
            campo.wait_for(
                state="visible",
                timeout=10000,
            )

            # Si el valor ya quedó seleccionado por un intento previo,
            # no lo tocamos nuevamente.
            try:
                valor_actual = normalizar_texto(
                    campo.input_value()
                )
                if valor_actual == valor_n:
                    return True
            except Exception:
                pass

            campo.click(
                timeout=10000
            )

            pagina.wait_for_timeout(
                500 + (intento * 300)
            )

            opciones = pagina.locator(
                '[role="option"]'
            )

            candidatas = []

            for indice in range(
                opciones.count()
            ):
                try:
                    opcion = opciones.nth(
                        indice
                    )

                    if not opcion.is_visible():
                        continue

                    texto_opcion = normalizar_texto(
                        opcion.inner_text()
                    )

                    if texto_opcion == valor_n:
                        candidatas.append(
                            opcion
                        )
                except Exception:
                    pass

            if len(candidatas) == 1:
                candidatas[0].click(
                    timeout=10000
                )

                pagina.wait_for_timeout(
                    700
                )

                valor_final = normalizar_texto(
                    campo.input_value()
                )

                if valor_final == valor_n:
                    return True

            # Algunos MUI filtran las opciones solamente cuando el
            # input recibe texto. Lo intentamos sin relajar la igualdad.
            try:
                campo.fill(
                    valor
                )

                pagina.wait_for_timeout(
                    800
                )

                opciones = pagina.locator(
                    '[role="option"]'
                )

                candidatas = []

                for indice in range(
                    opciones.count()
                ):
                    try:
                        opcion = opciones.nth(
                            indice
                        )

                        if not opcion.is_visible():
                            continue

                        if normalizar_texto(
                            opcion.inner_text()
                        ) == valor_n:
                            candidatas.append(
                                opcion
                            )
                    except Exception:
                        pass

                if len(candidatas) == 1:
                    candidatas[0].click(
                        timeout=10000
                    )

                    pagina.wait_for_timeout(
                        700
                    )

                    valor_final = normalizar_texto(
                        campo.input_value()
                    )

                    if valor_final == valor_n:
                        return True

            except Exception:
                pass

        except Exception:
            pass

        try:
            pagina.keyboard.press(
                "Escape"
            )
        except Exception:
            pass

        pagina.wait_for_timeout(
            350
        )

    return False


def detectar_campo_contenido_reto(
    pagina,
):

    archivos = pagina.locator(
        'input[type="file"]'
    )

    candidatos = []

    for i in range(
        archivos.count()
    ):

        try:

            archivo = archivos.nth(i)

            accept = (
                archivo.get_attribute(
                    "accept"
                )
                or ""
            ).lower()

            if ".pdf" not in accept:
                continue

            candidatos.append(
                archivo
            )

        except Exception:
            pass

    if len(candidatos) != 1:
        return None

    return candidatos[0]


# ============================================================
# BOTÓN GUARDAR
# ============================================================

def encontrar_boton_guardar(
    pagina,
):

    botones = pagina.get_by_role(
        "button",
        name="Editar",
        exact=True,
    )

    for i in range(
        botones.count()
    ):

        try:

            boton = botones.nth(i)

            if not boton.is_visible():
                continue

            if not boton.is_enabled():
                continue

            tipo = (
                boton.get_attribute(
                    "type"
                )
                or ""
            ).lower()

            if tipo == "submit":

                return boton

        except Exception:
            pass

    return None


# ============================================================
# ARCHIVO VISIBLE
# ============================================================

def archivo_visible(
    pagina,
    nombre_archivo,
):

    referencia = pagina.get_by_text(
        nombre_archivo,
        exact=False,
    )

    for i in range(
        referencia.count()
    ):

        try:

            if referencia.nth(i).is_visible():

                return True

        except Exception:
            pass

    return False


def _esperar_archivo_visible(
    pagina,
    nombre_archivo,
    timeout_ms=2200,
    intervalo_ms=100,
):
    """Termina en cuanto PRIZMA muestre el nombre del archivo cargado."""
    transcurrido = 0

    while True:
        if archivo_visible(pagina, nombre_archivo):
            return True

        if transcurrido >= timeout_ms:
            return False

        espera = min(intervalo_ms, timeout_ms - transcurrido)
        if espera <= 0:
            return False

        pagina.wait_for_timeout(espera)
        transcurrido += espera


# ============================================================
# PATCH
# ============================================================

def esperar_patch(
    pagina,
    respuestas,
    timeout_ms=15000,
):

    esperado = 0

    while esperado < timeout_ms:

        for respuesta in respuestas:

            if respuesta[
                "metodo"
            ] != "PATCH":
                continue

            if "/academic/activity/" not in respuesta[
                "url"
            ]:
                continue

            return respuesta

        pagina.wait_for_timeout(
            250
        )

        esperado += 250

    return None


# ============================================================
# LOGIN AUTOMÁTICO
# ============================================================


def validar_credenciales_prizma(
    usuario_prizma,
    contrasena_prizma,
    timeout_ms=30000,
):
    """Valida credenciales sin iniciar un cargue.

    Retorna un dict con ok/tipo/mensaje para distinguir credenciales
    rechazadas de una indisponibilidad técnica de PRIZMA.
    """
    usuario_prizma = str(usuario_prizma or "").strip()
    contrasena_prizma = str(contrasena_prizma or "")

    if not usuario_prizma or not contrasena_prizma:
        return {
            "ok": False,
            "tipo": "credenciales",
            "mensaje": "Debes ingresar usuario y contraseña de PRIZMA.",
        }

    try:
        with sync_playwright() as p:
            navegador = p.chromium.launch(
                headless=True,
                args=["--disable-dev-shm-usage"],
            )
            try:
                pagina = navegador.new_page(
                    viewport={"width": 1440, "height": 900}
                )
                pagina.goto(
                    URL_PRIZMA,
                    wait_until="domcontentloaded",
                    timeout=timeout_ms,
                )

                usuario = pagina.locator('input[name="identification_number"]')
                clave = pagina.locator('input[name="password"]')
                boton = pagina.get_by_role(
                    "button", name="Iniciar sesión", exact=True
                )

                usuario.wait_for(state="visible", timeout=timeout_ms)
                clave.wait_for(state="visible", timeout=timeout_ms)
                boton.wait_for(state="visible", timeout=timeout_ms)

                usuario.fill(usuario_prizma)
                clave.fill(contrasena_prizma)
                boton.click(timeout=10000)

                bienvenida = pagina.get_by_text(
                    "Bienvenido a Prizma admin", exact=False
                )

                # PRIZMA muestra el rechazo de credenciales como un toast breve.
                # Antes esperábamos hasta 15 s solo por la bienvenida; para cuando
                # revisábamos el body, el toast ya podía haber desaparecido.
                # Ahora observamos en paralelo éxito y rechazo desde el primer
                # instante posterior al clic.
                error_credenciales = pagina.get_by_text(
                    "Credenciales incorrectas", exact=False
                )

                limite = 15000
                transcurrido = 0
                intervalo = 250

                while transcurrido < limite:
                    try:
                        if bienvenida.count() > 0 and bienvenida.first.is_visible():
                            return {
                                "ok": True,
                                "tipo": "ok",
                                "mensaje": "Credenciales PRIZMA verificadas correctamente.",
                            }
                    except Exception:
                        pass

                    try:
                        if (
                            error_credenciales.count() > 0
                            and error_credenciales.first.is_visible()
                        ):
                            return {
                                "ok": False,
                                "tipo": "credenciales",
                                "mensaje": (
                                    "No se pudo iniciar sesión en PRIZMA. "
                                    "Verifica tu usuario y contraseña."
                                ),
                            }
                    except Exception:
                        pass

                    # Fallback por si PRIZMA cambia el contenedor del toast pero
                    # conserva el texto o una variante equivalente.
                    try:
                        texto = normalizar_texto(
                            pagina.locator("body").inner_text(timeout=1500)
                        )
                        senales_credenciales = [
                            "credenciales incorrectas",
                            "credenciales invalidas",
                            "usuario o contrasena",
                            "usuario y contrasena",
                            "contrasena incorrecta",
                            "usuario incorrecto",
                            "datos incorrectos",
                            "no autorizado",
                            "unauthorized",
                            "invalid credentials",
                        ]
                        if any(s in texto for s in senales_credenciales):
                            return {
                                "ok": False,
                                "tipo": "credenciales",
                                "mensaje": (
                                    "No se pudo iniciar sesión en PRIZMA. "
                                    "Verifica tu usuario y contraseña."
                                ),
                            }
                    except Exception:
                        pass

                    pagina.wait_for_timeout(intervalo)
                    transcurrido += intervalo

                # Si no vimos ni bienvenida ni el toast de rechazo, no culpamos
                # al usuario: puede ser lentitud, mantenimiento o fallo del backend.
                return {
                    "ok": False,
                    "tipo": "servicio",
                    "mensaje": (
                        "PRIZMA no respondió correctamente durante la validación "
                        "del acceso. Intenta nuevamente en unos minutos."
                    ),
                }
            finally:
                try:
                    navegador.close()
                except Exception:
                    pass

    except Exception as e:
        return {
            "ok": False,
            "tipo": "servicio",
            "mensaje": (
                "No fue posible conectar con PRIZMA para validar el acceso. "
                "Intenta nuevamente en unos minutos."
            ),
            "detalle": str(e),
        }


def iniciar_sesion_prizma(
    pagina,
    usuario_prizma,
    contrasena_prizma,
):

    print()
    print("--------------------------------------")
    print("LOGIN AUTOMÁTICO PRIZMA")
    print("--------------------------------------")

    pagina.goto(
        URL_PRIZMA,
        wait_until="domcontentloaded",
        timeout=60000,
    )

    campo_usuario = pagina.locator(
        'input[name="identification_number"]'
    )

    campo_contrasena = pagina.locator(
        'input[name="password"]'
    )

    boton_login = pagina.get_by_role(
        "button",
        name="Iniciar sesión",
        exact=True,
    )

    campo_usuario.wait_for(
        state="visible",
        timeout=30000,
    )

    campo_contrasena.wait_for(
        state="visible",
        timeout=30000,
    )

    boton_login.wait_for(
        state="visible",
        timeout=30000,
    )

    print(
        "✅ Formulario de login detectado."
    )

    campo_usuario.fill(
        usuario_prizma
    )

    campo_contrasena.fill(
        contrasena_prizma
    )

    print(
        "✅ Credenciales introducidas."
    )

    print(
        "Iniciando sesión..."
    )

    boton_login.click(
        timeout=10000
    )

    # --------------------------------------------------------
    # CONFIRMAR PANEL
    # --------------------------------------------------------

    bienvenida = pagina.get_by_text(
        "Bienvenido a Prizma admin",
        exact=False,
    )

    try:

        bienvenida.wait_for(
            state="visible",
            timeout=60000,
        )

    except Exception:

        try:

            texto_visible = pagina.locator(
                "body"
            ).inner_text()

            print()
            print(
                "No se pudo confirmar el panel."
            )

            print(
                texto_visible[:1200]
            )

        except Exception:
            pass

        raise RuntimeError(
            "ERROR_LOGIN_PRIZMA"
        )

    print(
        "✅ Login automático confirmado."
    )

    print(
        "URL después del login:",
        pagina.url,
    )

    # --------------------------------------------------------
    # ENTRAR A ACTIVIDADES
    # --------------------------------------------------------

    tarjeta_actividades = pagina.get_by_role(
        "button"
    ).filter(
        has_text="Actividades"
    ).first

    tarjeta_actividades.wait_for(
        state="visible",
        timeout=30000,
    )

    print(
        "✅ Tarjeta Actividades detectada."
    )

    tarjeta_actividades.click(
        timeout=10000
    )

    # --------------------------------------------------------
    # CONFIRMAR LISTADO
    # --------------------------------------------------------

    buscador = pagina.locator(
        'input[placeholder="Buscar..."]'
    )

    buscador.wait_for(
        state="visible",
        timeout=60000,
    )

    print(
        "✅ Módulo Actividades abierto."
    )

    return True


# ============================================================
# PROCESAR ACTIVIDAD
# ============================================================

def procesar_actividad(
    pagina,
    actividad,
    indice_recursos,
    carpeta_temp,
    respuestas,
    captura,
    actividades_curso,
    estado_navegacion=None,
):

    print()
    print()
    print("======================================")

    print(
        "ACTIVIDAD:",
        actividad["nombre"],
    )

    print(
        "FILA:",
        actividad["fila_excel"],
    )

    print(
        "CATEGORÍA:",
        actividad["categoria_prizma"],
    )

    print(
        "TIPO:",
        actividad["tipo_archivo"],
    )

    print("======================================")

    # 1. VALIDAR CARPETA DE RECURSO EN GOOGLE DRIVE (SI EXISTE)

    error_carpeta = validar_carpeta_drive_para_actividad(
        actividad
    )

    if error_carpeta:
        return {
            "ok": False,
            "error": error_carpeta,
            "recurso": "",
        }

    # 2. RECURSO DEL ZIP

    recurso, error = resolver_recurso(
        actividad,
        indice_recursos,
        carpeta_temp,
        actividades_curso,
    )

    if error:

        return {
            "ok": False,
            "error": error,
            "recurso": "",
        }

    ruta_recurso = recurso[
        "ruta"
    ]

    nombre_recurso = recurso[
        "nombre_original"
    ]

    # 2. LISTADO

    if not asegurar_listado(
        pagina
    ):

        return {
            "ok": False,
            "error":
                "ERROR_NO_SE_PUDO_VOLVER_AL_LISTADO",
            "recurso":
                nombre_recurso,
        }

    # 3. CATEGORÍA

    if not entrar_categoria(
        pagina,
        actividad["categoria_prizma"],
        estado_navegacion,
    ):

        return {
            "ok": False,
            "error":
                "ERROR_ENTRANDO_CATEGORIA",
            "recurso":
                nombre_recurso,
        }

    # 4. BUSCAR

    fila, error = buscar_actividad_correcta(
        pagina,
        actividad,
    )

    if error:

        return {
            "ok": False,
            "error": error,
            "recurso":
                nombre_recurso,
        }

    # 5. EDITAR

    if not abrir_edicion(
        pagina,
        fila,
    ):

        return {
            "ok": False,
            "error":
                "ERROR_ABRIENDO_EDICION",
            "recurso":
                nombre_recurso,
        }

    # 6. PREPARAR CAMPOS SEGÚN CATEGORÍA

    if actividad[
        "categoria_prizma"
    ] == "CHALLENGE":

        configuracion_reto, error_reto = (
            obtener_configuracion_reto(
                actividad
            )
        )

        if error_reto:

            cancelar_edicion_segura(
                pagina
            )

            return {
                "ok": False,
                "error": error_reto,
                "recurso": nombre_recurso,
            }

        if not seleccionar_autocomplete_reto(
            pagina,
            "Corte",
            configuracion_reto["corte"],
        ):

            cancelar_edicion_segura(
                pagina
            )

            return {
                "ok": False,
                "error":
                    "ERROR_SELECCIONANDO_CORTE",
                "recurso": nombre_recurso,
            }

        if not seleccionar_autocomplete_reto(
            pagina,
            "Nivel de dificultad",
            configuracion_reto["nivel"],
        ):

            cancelar_edicion_segura(
                pagina
            )

            return {
                "ok": False,
                "error":
                    "ERROR_SELECCIONANDO_NIVEL_DIFICULTAD",
                "recurso": nombre_recurso,
            }

        campo_recurso = detectar_campo_contenido_reto(
            pagina
        )

        if campo_recurso is None:

            cancelar_edicion_segura(
                pagina
            )

            return {
                "ok": False,
                "error":
                    "ERROR_CAMPO_CONTENIDO_RETO_NO_ENCONTRADO",
                "recurso": nombre_recurso,
            }

        print()
        print(
            "✅ Corte:",
            configuracion_reto["corte"],
        )
        print(
            "✅ Nivel:",
            configuracion_reto["nivel"],
        )
        print(
            "✅ El PDF será cargado en CONTENIDO."
        )

    else:

        campo_recurso = detectar_campo_recurso(
            pagina
        )

        if campo_recurso is None:

            cancelar_edicion_segura(
                pagina
            )

            return {
                "ok": False,
                "error":
                    "ERROR_CAMPO_RECURSO_NO_ENCONTRADO",
                "recurso":
                    nombre_recurso,
            }

        print()
        print(
            "✅ El archivo será cargado en RECURSO."
        )

        print(
            "Material descargable: IGNORADO"
        )

    # 7. DESCRIPCIÓN

    descripcion_ok, error_descripcion = limpiar_descripcion(
        pagina
    )

    if not descripcion_ok:

        cancelar_edicion_segura(
            pagina
        )

        return {
            "ok": False,
            "error":
                error_descripcion,
            "recurso":
                nombre_recurso,
        }

    # 8. CARGAR

    print()
    print(
        "Cargando:",
        nombre_recurso,
    )

    try:

        campo_recurso.set_input_files(
            ruta_recurso
        )

    except Exception:

        cancelar_edicion_segura(
            pagina
        )

        return {
            "ok": False,
            "error":
                "ERROR_SET_INPUT_FILES",
            "recurso":
                nombre_recurso,
        }

    # 9. VISIBLE - espera adaptativa; si PRIZMA responde rapido no pagamos
    # los 2.2 segundos fijos de la version anterior.
    if not _esperar_archivo_visible(
        pagina,
        nombre_recurso,
        timeout_ms=2200,
    ):

        cancelar_edicion_segura(
            pagina
        )

        return {
            "ok": False,
            "error":
                "ERROR_RECURSO_NO_VISIBLE",
            "recurso":
                nombre_recurso,
        }

    print(
        "✅ Recurso visible."
    )

    # 10. GUARDAR

    boton_guardar = encontrar_boton_guardar(
        pagina
    )

    if boton_guardar is None:

        cancelar_edicion_segura(
            pagina
        )

        return {
            "ok": False,
            "error":
                "ERROR_BOTON_GUARDAR",
            "recurso":
                nombre_recurso,
        }

    respuestas.clear()

    captura[
        "activa"
    ] = True

    print(
        "💾 Guardando..."
    )

    try:

        boton_guardar.click(
            timeout=10000
        )

    except Exception:

        captura[
            "activa"
        ] = False

        cancelar_edicion_segura(
            pagina
        )

        return {
            "ok": False,
            "error":
                "ERROR_CLIC_GUARDAR",
            "recurso":
                nombre_recurso,
        }

    # 11. PATCH

    respuesta_patch = esperar_patch(
        pagina,
        respuestas,
        timeout_ms=15000,
    )

    captura[
        "activa"
    ] = False

    if respuesta_patch is None:

        asegurar_listado(
            pagina
        )

        return {
            "ok": False,
            "error":
                "ERROR_PATCH_NO_CONFIRMADO",
            "recurso":
                nombre_recurso,
        }

    if not (
        200
        <= respuesta_patch[
            "status"
        ]
        < 300
    ):

        asegurar_listado(
            pagina
        )

        return {
            "ok": False,
            "error":
                "ERROR_PATCH_HTTP_"
                + str(
                    respuesta_patch["status"]
                ),
            "recurso":
                nombre_recurso,
        }

    print(
        "✅ PATCH confirmado:",
        respuesta_patch["status"],
    )

    # ========================================================
    # NO REABRIR.
    # DIRECTO A SIGUIENTE ACTIVIDAD.
    # ========================================================

    if not asegurar_listado(
        pagina
    ):

        return {
            "ok": False,
            "error":
                "ERROR_POST_GUARDADO_LISTADO",
            "recurso":
                nombre_recurso,
        }

    return {
        "ok": True,
        "error": "",
        "recurso":
            nombre_recurso,
    }


# ============================================================
# MOTOR COMPLETO
# ============================================================


def _es_error_estructural_prizma(error):
    texto = str(error or "")
    senales = [
        "Page crashed",
        "Target crashed",
        "Target page, context or browser has been closed",
        "Timeout 30000ms exceeded",
        "Timeout 60000ms exceeded",
        "ERROR_ENTRANDO_CATEGORIA",
        "ERROR_NO_SE_PUDO_VOLVER_AL_LISTADO",
        "ERROR_RECUPERANDO_TRAS_PAGE_CRASH",
        "ERROR_RECUPERANDO_PAGINA",
        "ERROR_RECUPERANDO_FILA",
        "ERROR_LOGIN_PRIZMA",
    ]
    return any(senal in texto for senal in senales)


def _cancelacion_solicitada(estado):
    return bool(estado.get("cancelar_solicitado"))


def ejecutar_cargue(
    ruta_excel,
    ruta_zip,
    carpeta_temp,
    ruta_reporte,
    procesar_ovi,
    procesar_ova,
    procesar_retos,
    usuario_prizma,
    contrasena_prizma,
    estado,
):
    navegador = None

    try:
        print()
        print("======================================")
        print("AUTO PRIZMA PRO")
        print("======================================")
        print("VERSION_SCRIPT =", VERSION_SCRIPT)
        print("BUILD_INTERNO =", BUILD_INTERNO)
        print("URL =", URL_PRIZMA)
        print("MODO = CURSO COMPLETO")
        print("NAVEGADOR = HEADLESS")
        print("LOGIN = AUTOMÁTICO")
        print()
        print("H5P -> RECURSO")
        print("PDF -> RECURSO")
        print("MATERIAL DESCARGABLE -> NO TOCAR")

        actualizar_estado(
            estado,
            etapa="preparando",
            mensaje="Leyendo Excel y preparando recursos...",
            terminado=False,
        )

        actividades = leer_actividades_excel(
            ruta_excel,
            procesar_ovi,
            procesar_ova,
            procesar_retos,
        )

        if not actividades:
            actualizar_estado(
                estado,
                etapa="error",
                mensaje="No se encontraron actividades OVI/OVA/Retos compatibles.",
                terminado=True,
            )
            return

        indice_recursos = crear_indice_recursos(ruta_zip)
        if not indice_recursos:
            actualizar_estado(
                estado,
                etapa="error",
                mensaje="No se encontraron archivos H5P/PDF dentro del ZIP.",
                terminado=True,
            )
            return

        actividades_a_procesar = actividades
        preparar_detalle_actividades(estado, actividades_a_procesar)

        actualizar_estado(
            estado,
            etapa="login",
            mensaje="Iniciando sesión automáticamente en PRIZMA...",
            total=len(actividades_a_procesar),
            procesadas=0,
            exitosas=0,
            errores=0,
        )

        with sync_playwright() as p:
            respuestas = []
            captura = {"activa": False}
            estado_navegacion = {"categoria_actual": None}

            def registrar_respuesta(response):
                if not captura["activa"]:
                    return
                try:
                    metodo = response.request.method.upper()
                    if metodo == "GET":
                        return
                    respuestas.append({
                        "metodo": metodo,
                        "status": response.status,
                        "url": response.url,
                    })
                except Exception:
                    pass

            def abrir_sesion_limpia(motivo=""):
                nonlocal navegador
                estado_navegacion["categoria_actual"] = None
                if motivo:
                    print("⚠️ Reiniciando sesión PRIZMA:", motivo)
                try:
                    if navegador is not None:
                        navegador.close()
                except Exception:
                    pass

                navegador = p.chromium.launch(
                    headless=True,
                    args=["--disable-dev-shm-usage"],
                )
                nueva_pagina = navegador.new_page(
                    viewport={"width": 1440, "height": 900}
                )
                nueva_pagina.on("response", registrar_respuesta)
                iniciar_sesion_prizma(
                    nueva_pagina,
                    usuario_prizma,
                    contrasena_prizma,
                )
                return nueva_pagina

            pagina = abrir_sesion_limpia("inicio del cargue")

            actualizar_estado(
                estado,
                etapa="procesando",
                mensaje="Login correcto. Módulo Actividades abierto.",
            )

            exitosas = 0
            errores = 0
            procesadas = 0
            errores_estructurales_consecutivos = 0

            for numero, actividad in enumerate(actividades_a_procesar, start=1):
                if _cancelacion_solicitada(estado):
                    actualizar_estado(
                        estado,
                        etapa="cancelado",
                        mensaje=(
                            "Proceso cancelado por el usuario. "
                            f"Se alcanzaron a procesar {procesadas} de "
                            f"{len(actividades_a_procesar)} actividades."
                        ),
                        terminado=True,
                    )
                    return

                print()
                print("######################################")
                print(f"ACTIVIDAD {numero} DE {len(actividades_a_procesar)}")
                print("######################################")

                actualizar_estado(
                    estado,
                    etapa="procesando",
                    mensaje=(
                        f"Procesando {numero} de {len(actividades_a_procesar)}: "
                        + actividad["nombre"]
                    ),
                )
                actualizar_detalle_actividad(estado, numero, "procesando")

                # Si la sesión quedó dañada en la actividad anterior, no se reutiliza.
                if errores_estructurales_consecutivos > 0:
                    try:
                        pagina = abrir_sesion_limpia(
                            "recuperación preventiva tras error estructural"
                        )
                        errores_estructurales_consecutivos = 0
                    except Exception as e:
                        print(traceback.format_exc())
                        resultado = {
                            "ok": False,
                            "error": "ERROR_RECUPERANDO_SESION: " + str(e),
                            "recurso": "",
                        }
                    else:
                        resultado = None
                else:
                    resultado = None

                if resultado is None:
                    # Preflight: si no estamos realmente en el listado, reiniciar antes de tocar la actividad.
                    try:
                        if not asegurar_listado(pagina):
                            pagina = abrir_sesion_limpia(
                                "no fue posible confirmar el listado antes de la actividad"
                            )
                    except Exception:
                        pagina = abrir_sesion_limpia(
                            "fallo comprobando el listado antes de la actividad"
                        )

                    resultado = None
                    ultimo_error = None

                    # Una actividad puede reintentarse una sola vez, siempre con navegador limpio.
                    for intento in (1, 2):
                        if _cancelacion_solicitada(estado):
                            actualizar_estado(
                                estado,
                                etapa="cancelado",
                                mensaje="Proceso cancelado por el usuario.",
                                terminado=True,
                            )
                            return

                        try:
                            resultado = procesar_actividad(
                                pagina,
                                actividad,
                                indice_recursos,
                                carpeta_temp,
                                respuestas,
                                captura,
                                actividades_a_procesar,
                                estado_navegacion,
                            )
                        except Exception as e:
                            print(traceback.format_exc())
                            ultimo_error = "ERROR_NO_CONTROLADO: " + str(e)
                            resultado = {
                                "ok": False,
                                "error": ultimo_error,
                                "recurso": "",
                            }

                        if resultado.get("ok"):
                            break

                        error_actual = resultado.get("error", "")
                        if intento == 1 and _es_error_estructural_prizma(error_actual):
                            actualizar_estado(
                                estado,
                                mensaje=(
                                    "PRIZMA o el navegador perdieron estabilidad. "
                                    "Reiniciando la sesión y reintentando la actividad actual..."
                                ),
                            )
                            try:
                                pagina = abrir_sesion_limpia(error_actual)
                                continue
                            except Exception as e_reinicio:
                                resultado = {
                                    "ok": False,
                                    "error": (
                                        "ERROR_RECUPERANDO_SESION: "
                                        + str(e_reinicio)
                                    ),
                                    "recurso": resultado.get("recurso", ""),
                                }
                        break

                procesadas += 1

                if resultado["ok"]:
                    exitosas += 1
                    errores_estructurales_consecutivos = 0
                    actualizar_detalle_actividad(estado, numero, "ok")
                    guardar_resultado(
                        ruta_reporte,
                        actividad,
                        "OK",
                        "Carga guardada - PATCH confirmado",
                        resultado["recurso"],
                    )
                else:
                    errores += 1
                    error_actual = resultado.get("error", "ERROR_DESCONOCIDO")
                    if _es_error_estructural_prizma(error_actual):
                        errores_estructurales_consecutivos += 1
                    else:
                        errores_estructurales_consecutivos = 0

                    actualizar_detalle_actividad(
                        estado, numero, "error", error_actual
                    )
                    guardar_resultado(
                        ruta_reporte,
                        actividad,
                        "ERROR",
                        error_actual,
                        resultado.get("recurso", ""),
                    )

                actualizar_estado(
                    estado,
                    procesadas=procesadas,
                    exitosas=exitosas,
                    errores=errores,
                    mensaje=(
                        f"Procesadas {procesadas} de {len(actividades_a_procesar)}"
                    ),
                )

                if _cancelacion_solicitada(estado):
                    actualizar_estado(
                        estado,
                        etapa="cancelado",
                        mensaje="Proceso cancelado por el usuario.",
                        terminado=True,
                    )
                    return

            try:
                if navegador is not None:
                    navegador.close()
            except Exception:
                pass

        actualizar_estado(
            estado,
            etapa="finalizado",
            mensaje=(
                "Cargue completo finalizado. "
                f"Exitosas: {exitosas}. Errores: {errores}."
            ),
            terminado=True,
        )

    except Exception as e:
        print()
        print(traceback.format_exc())
        mensaje_error = str(e)
        if "ERROR_LOGIN_PRIZMA" in mensaje_error:
            mensaje_error = (
                "No fue posible iniciar sesión en PRIZMA. "
                "Verifica el usuario y la contraseña."
            )
        actualizar_estado(
            estado,
            etapa="error",
            mensaje=mensaje_error,
            terminado=True,
        )
    finally:
        try:
            if navegador is not None:
                navegador.close()
        except Exception:
            pass

