from fastapi import (
    FastAPI,
    UploadFile,
    File,
    Form,
    BackgroundTasks,
    Request as FastAPIRequest,
)

from fastapi.responses import (
    Response,
    HTMLResponse,
    JSONResponse,
    FileResponse,
    RedirectResponse,
)

from io import BytesIO
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from openpyxl import load_workbook, Workbook

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleAuthRequest
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

import motor_prizma as motor_prizma_modulo

from motor_prizma import (
    ejecutar_cargue,
    validar_credenciales_prizma,
    determinar_tipo_archivo,
    normalizar_categoria,
    normalizar_texto,
    prevalidar_recursos_drive_zip,
)

import zipfile
import csv
import os
import uuid
import html
import json
import re
import unicodedata
import threading
import shutil
import hmac
import base64
import hashlib
import secrets
import contextvars
from collections import deque
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo


app = FastAPI(
    title="Auto Prizma Pro"
)


# ============================================================
# CARPETAS
# ============================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

UPLOADS_DIR = os.path.join(
    BASE_DIR,
    "uploads",
)

TEMP_DIR = os.path.join(
    BASE_DIR,
    "temp",
)

DATA_DIR = os.environ.get(
    "AUTO_PRIZMA_DATA_DIR",
    os.path.join(BASE_DIR, "data"),
)

RESULTADOS_DIR = os.path.join(
    DATA_DIR,
    "reportes",
)

HISTORIAL_PATH = os.path.join(
    DATA_DIR,
    "historial.json",
)

USUARIOS_PATH = os.path.join(
    DATA_DIR,
    "usuarios.json",
)

TOKENS_DIR = os.path.join(
    DATA_DIR,
    "tokens",
)

SECRETO_PATH = os.path.join(
    DATA_DIR,
    "secreto_sesion.txt",
)


for carpeta in [
    UPLOADS_DIR,
    TEMP_DIR,
    DATA_DIR,
    RESULTADOS_DIR,
    TOKENS_DIR,
]:

    os.makedirs(
        carpeta,
        exist_ok=True,
    )


# ============================================================
# SESIONES Y USUARIOS
# ============================================================
#
# - Cada persona entra con usuario y contrasena.
# - La contrasena se guarda como PBKDF2-SHA256 con sal propia.
# - La sesion viaja en una cookie firmada con HMAC; el servidor no
#   guarda estado de sesion, asi que reiniciar uvicorn no desloguea.
# - Cada usuario tiene SU PROPIO token de Google en data/tokens/.
#

SESION_COOKIE = "ap_sesion"
SESION_HORAS = 12
PBKDF2_ITERACIONES = 240000

USUARIOS_LOCK = threading.Lock()
INTENTOS_LOCK = threading.Lock()
INTENTOS_FALLIDOS = {}
MAX_INTENTOS = 5
BLOQUEO_SEGUNDOS = 120

# Se propaga a los endpoints sincronos (Starlette copia el contexto
# al threadpool) y se fija a mano en el hilo de cargue.
USUARIO_ACTUAL = contextvars.ContextVar("usuario_actual", default=None)


def _secreto_sesion():
    """Clave para firmar cookies. Prioriza la variable de entorno."""
    desde_entorno = str(
        os.environ.get("AUTO_PRIZMA_SECRET") or ""
    ).strip()

    if desde_entorno:
        return desde_entorno.encode("utf-8")

    if os.path.isfile(SECRETO_PATH):
        try:
            with open(SECRETO_PATH, "r", encoding="utf-8") as archivo:
                guardado = archivo.read().strip()
            if guardado:
                return guardado.encode("utf-8")
        except OSError:
            pass

    generado = secrets.token_urlsafe(48)

    try:
        with open(SECRETO_PATH, "w", encoding="utf-8") as archivo:
            archivo.write(generado)
    except OSError:
        pass

    return generado.encode("utf-8")


SECRETO_SESION = _secreto_sesion()


def _cookie_segura():
    """Marca la cookie como Secure cuando la app corre bajo https."""
    if str(os.environ.get("AUTO_PRIZMA_COOKIE_SECURE") or "").strip() == "1":
        return True

    redirect_uri = str(
        os.environ.get("GOOGLE_REDIRECT_URI") or ""
    ).strip().lower()

    return redirect_uri.startswith("https://")


def _normalizar_usuario(usuario):
    return str(usuario or "").strip().lower()


def _slug_usuario(usuario):
    """Nombre de archivo seguro para el token de Google de ese usuario."""
    base = re.sub(
        r"[^a-z0-9._-]+",
        "_",
        _normalizar_usuario(usuario),
    ).strip("._-")

    if not base:
        base = "usuario"

    firma = hashlib.sha256(
        _normalizar_usuario(usuario).encode("utf-8")
    ).hexdigest()[:10]

    return base[:40] + "-" + firma


def _cargar_usuarios():
    if not os.path.isfile(USUARIOS_PATH):
        return {}

    try:
        with open(USUARIOS_PATH, "r", encoding="utf-8") as archivo:
            datos = json.load(archivo)
        return datos if isinstance(datos, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def _guardar_usuarios(usuarios):
    temporal = USUARIOS_PATH + ".tmp"

    with open(temporal, "w", encoding="utf-8") as archivo:
        json.dump(usuarios, archivo, ensure_ascii=False, indent=2)

    os.replace(temporal, USUARIOS_PATH)


def _hash_contrasena(contrasena, sal=None):
    sal = sal or secrets.token_hex(16)

    derivada = hashlib.pbkdf2_hmac(
        "sha256",
        str(contrasena or "").encode("utf-8"),
        bytes.fromhex(sal),
        PBKDF2_ITERACIONES,
    )

    return sal, derivada.hex()


def _verificar_contrasena(contrasena, sal, esperado):
    try:
        _, calculado = _hash_contrasena(contrasena, sal)
    except (ValueError, TypeError):
        return False

    return hmac.compare_digest(calculado, str(esperado or ""))


def _crear_usuario(usuario, contrasena, nombre=""):
    """Alta o cambio de contrasena. Lo usa gestionar_usuarios.py."""
    usuario = _normalizar_usuario(usuario)

    if not usuario:
        raise ValueError("El usuario no puede estar vacio.")

    if len(str(contrasena or "")) < 8:
        raise ValueError("La contrasena debe tener al menos 8 caracteres.")

    sal, derivada = _hash_contrasena(contrasena)

    with USUARIOS_LOCK:
        usuarios = _cargar_usuarios()
        existente = usuarios.get(usuario) or {}

        usuarios[usuario] = {
            "nombre": str(nombre or existente.get("nombre") or usuario),
            "sal": sal,
            "hash": derivada,
            "activo": True,
            "creado": existente.get("creado") or datetime.now().isoformat(),
            "actualizado": datetime.now().isoformat(),
        }

        _guardar_usuarios(usuarios)

    return usuario


def _autenticar(usuario, contrasena):
    usuario = _normalizar_usuario(usuario)
    registro = _cargar_usuarios().get(usuario)

    if not registro or not registro.get("activo", True):
        # Se calcula igual para no revelar por tiempo si el usuario existe.
        _hash_contrasena(contrasena, secrets.token_hex(16))
        return None

    if not _verificar_contrasena(
        contrasena,
        registro.get("sal"),
        registro.get("hash"),
    ):
        return None

    return usuario


def _bloqueado(clave):
    with INTENTOS_LOCK:
        datos = INTENTOS_FALLIDOS.get(clave)

        if not datos:
            return 0

        intentos, hasta = datos

        if intentos < MAX_INTENTOS:
            return 0

        restante = hasta - datetime.now()

        if restante.total_seconds() <= 0:
            INTENTOS_FALLIDOS.pop(clave, None)
            return 0

        return int(restante.total_seconds()) + 1


def _registrar_fallo(clave):
    with INTENTOS_LOCK:
        intentos, _ = INTENTOS_FALLIDOS.get(clave, (0, datetime.now()))
        intentos += 1
        INTENTOS_FALLIDOS[clave] = (
            intentos,
            datetime.now() + timedelta(seconds=BLOQUEO_SEGUNDOS),
        )


def _limpiar_fallos(clave):
    with INTENTOS_LOCK:
        INTENTOS_FALLIDOS.pop(clave, None)


def _firmar_sesion(usuario):
    expira = int(
        (datetime.now() + timedelta(hours=SESION_HORAS)).timestamp()
    )

    cuerpo = (
        base64.urlsafe_b64encode(
            _normalizar_usuario(usuario).encode("utf-8")
        ).decode("ascii").rstrip("=")
        + "."
        + str(expira)
    )

    firma = hmac.new(
        SECRETO_SESION,
        cuerpo.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    return cuerpo + "." + firma


def _leer_sesion(valor):
    partes = str(valor or "").split(".")

    if len(partes) != 3:
        return None

    cuerpo = partes[0] + "." + partes[1]

    firma = hmac.new(
        SECRETO_SESION,
        cuerpo.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(firma, partes[2]):
        return None

    try:
        if int(partes[1]) < int(datetime.now().timestamp()):
            return None
    except ValueError:
        return None

    try:
        relleno = "=" * (-len(partes[0]) % 4)
        usuario = base64.urlsafe_b64decode(
            partes[0] + relleno
        ).decode("utf-8")
    except (ValueError, UnicodeDecodeError):
        return None

    usuario = _normalizar_usuario(usuario)
    registro = _cargar_usuarios().get(usuario)

    if not registro or not registro.get("activo", True):
        return None

    return usuario


def _usuario_de_request(request):
    return _leer_sesion(request.cookies.get(SESION_COOKIE))


def _usuario_actual():
    return USUARIO_ACTUAL.get()


def _nombre_visible(usuario):
    registro = _cargar_usuarios().get(_normalizar_usuario(usuario)) or {}
    return str(registro.get("nombre") or usuario or "")


def _es_de(trabajo, usuario):
    """True si ese trabajo pertenece al usuario indicado."""
    if not trabajo:
        return False
    return _normalizar_usuario(
        trabajo.get("usuario_app")
    ) == _normalizar_usuario(usuario)


# ============================================================
# GOOGLE OAUTH
# ============================================================

GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

# Token compartido de la version anterior. Ya NO se usa: cada persona
# tiene el suyo en data/tokens/. Se avisa para que se borre.
GOOGLE_TOKEN_COMPARTIDO_ANTIGUO = os.path.join(
    DATA_DIR,
    "google_token.json",
)

if os.path.isfile(GOOGLE_TOKEN_COMPARTIDO_ANTIGUO):
    print(
        "AVISO: existe data/google_token.json (token compartido de la version "
        "anterior). Ya no se usa. Revoca ese acceso en "
        "https://myaccount.google.com/permissions y borra el archivo."
    )


def _ruta_token_google(usuario):
    return os.path.join(
        TOKENS_DIR,
        _slug_usuario(usuario) + ".json",
    )


GOOGLE_OAUTH_STATES = {}
GOOGLE_OAUTH_LOCK = threading.Lock()


def _google_client_config():
    client_id = str(
        os.environ.get("GOOGLE_CLIENT_ID") or ""
    ).strip()

    client_secret = str(
        os.environ.get("GOOGLE_CLIENT_SECRET") or ""
    ).strip()

    redirect_uri = str(
        os.environ.get("GOOGLE_REDIRECT_URI") or ""
    ).strip()

    faltantes = []

    if not client_id:
        faltantes.append("GOOGLE_CLIENT_ID")

    if not client_secret:
        faltantes.append("GOOGLE_CLIENT_SECRET")

    if not redirect_uri:
        faltantes.append("GOOGLE_REDIRECT_URI")

    if faltantes:
        raise RuntimeError(
            "Faltan variables de entorno de Google: "
            + ", ".join(faltantes)
        )

    if (
        redirect_uri.startswith("http://127.0.0.1")
        or redirect_uri.startswith("http://localhost")
    ):
        os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"

    return {
        "web": {
            "client_id": client_id,
            "client_secret": client_secret,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [redirect_uri],
        }
    }, redirect_uri


def _guardar_credenciales_google(credenciales, usuario=None):
    usuario = usuario or _usuario_actual()

    if not usuario:
        raise RuntimeError(
            "No hay sesion activa para guardar el token de Google."
        )

    destino = _ruta_token_google(usuario)
    temporal = destino + ".tmp"

    with open(
        temporal,
        "w",
        encoding="utf-8",
    ) as archivo:
        archivo.write(
            credenciales.to_json()
        )

    os.replace(
        temporal,
        destino,
    )


def _cargar_credenciales_google(usuario=None):
    usuario = usuario or _usuario_actual()

    if not usuario:
        return None

    ruta = _ruta_token_google(usuario)

    if not os.path.isfile(ruta):
        return None

    try:
        credenciales = Credentials.from_authorized_user_file(
            ruta,
            GOOGLE_SCOPES,
        )

        if (
            credenciales.expired
            and credenciales.refresh_token
        ):
            credenciales.refresh(
                GoogleAuthRequest()
            )

            _guardar_credenciales_google(
                credenciales,
                usuario,
            )

        if not credenciales.valid:
            return None

        return credenciales

    except Exception:
        return None


def _google_conectado(usuario=None):
    return (
        _cargar_credenciales_google(usuario)
        is not None
    )


# ============================================================
# GOOGLE DRIVE PRIVADO - MODO ULTRARRAPIDO
# ============================================================
#
# Objetivo:
# - NO descargar H5P/PDF desde Google Drive.
# - NO abrir cada archivo.
# - Consultar SOLO metadatos: nombre + tamano exacto.
# - Consultar muchas carpetas en BATCH para evitar una peticion HTTP
#   secuencial por cada actividad.
# - Guardar el resultado en el cache del motor. Durante el cargue a
#   PRIZMA, el motor reutiliza ese cache y no vuelve a consultar Drive.
#

DRIVE_BATCH_MAX = 100

# Recursos resueltos por el preflight del trabajo que esta ejecutando
# cada worker. Esto evita volver a consultar Drive dentro de PRIZMA.
_DRIVE_THREAD_LOCAL = threading.local()
_CACHE_ARCHIVOS_DRIVE = {}
_CACHE_ARCHIVOS_DRIVE_LOCK = threading.Lock()
_CACHE_REFERENCIAS_DRIVE = {}
_CACHE_REFERENCIAS_DRIVE_LOCK = threading.Lock()


def _clave_actividad_drive(actividad):
    return (
        str(actividad.get("programa") or "").strip().casefold(),
        str(actividad.get("curso") or "").strip().casefold(),
        str(actividad.get("hoja") or "").strip().casefold(),
        int(actividad.get("fila_excel") or 0),
        str(actividad.get("nombre") or "").strip().casefold(),
        str(actividad.get("tipo_archivo") or "").strip().upper(),
    )


def _recurso_resuelto_thread(actividad):
    mapa = getattr(_DRIVE_THREAD_LOCAL, "recursos", None) or {}
    return mapa.get(_clave_actividad_drive(actividad))


def _extraer_referencia_drive(url):
    url = str(url or "").strip()
    if not url:
        return None, None

    patrones_carpeta = [
        r"drive\.google\.com/(?:drive/(?:u/\d+/)?folders|folders)/([A-Za-z0-9_-]+)",
    ]
    for patron in patrones_carpeta:
        m = re.search(patron, url, flags=re.IGNORECASE)
        if m:
            return "folder", m.group(1)

    patrones_archivo = [
        r"drive\.google\.com/file/d/([A-Za-z0-9_-]+)",
        r"drive\.google\.com/(?:open|uc)\?[^#]*\bid=([A-Za-z0-9_-]+)",
    ]
    for patron in patrones_archivo:
        m = re.search(patron, url, flags=re.IGNORECASE)
        if m:
            return "file", m.group(1)

    return None, None


def _extraer_gid(url_google_sheet):
    """Devuelve el gid (id de la pestaña) que trae el enlace, si lo trae."""
    m = re.search(
        r"[#?&]gid=(\d+)",
        str(url_google_sheet or ""),
    )
    return m.group(1) if m else ""


def _resolver_hoja_objetivo(url_google_sheet):
    """Traduce el gid del enlace al nombre real de la pestaña.

    Devuelve None si el enlace no trae gid: en ese caso se conserva el
    comportamiento anterior (todas las hojas).
    """
    gid = _extraer_gid(url_google_sheet)
    if not gid:
        return None

    spreadsheet_id = _extraer_spreadsheet_id(url_google_sheet)
    if not spreadsheet_id:
        return None

    credenciales = _cargar_credenciales_google()
    if credenciales is None:
        raise ValueError("Primero debes conectar tu cuenta de Google.")

    sheets = build(
        "sheets",
        "v4",
        credentials=credenciales,
        cache_discovery=False,
    )

    meta = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets.properties(sheetId,title)",
    ).execute()

    for item in meta.get("sheets", []):
        propiedades = item.get("properties", {}) or {}
        if str(propiedades.get("sheetId")) == gid:
            titulo = str(propiedades.get("title") or "").strip()
            if titulo:
                print("Hoja objetivo: gid=" + gid + " -> '" + titulo + "'")
                return titulo

    raise ValueError(
        "El enlace apunta a una pestaña (gid=" + gid + ") que ya no existe "
        "en esa matriz. Abre la pestaña correcta y copia el enlace de nuevo."
    )


def _misma_hoja(nombre_hoja, hoja_objetivo):
    if not hoja_objetivo:
        return True
    return normalizar_texto(nombre_hoja) == normalizar_texto(hoja_objetivo)


def _extraer_spreadsheet_id(url_google_sheet):
    m = re.search(
        r"docs\.google\.com/spreadsheets/d/([A-Za-z0-9_-]+)",
        str(url_google_sheet or ""),
        flags=re.IGNORECASE,
    )
    return m.group(1) if m else ""


def _url_desde_cell_data(celda):
    """Devuelve SOLO un enlace de Google Drive encontrado en la celda.

    La columna G puede contener Smart Chips, hipervinculos, formulas HYPERLINK
    o texto visible. No usamos enlaces de otras plataformas y, si la celda
    contiene varios links, priorizamos el primero que realmente sea Drive.
    """
    if not isinstance(celda, dict):
        return ""

    candidatos = []

    hyperlink = str(celda.get("hyperlink") or "").strip()
    if hyperlink:
        candidatos.append(hyperlink)

    for run in celda.get("chipRuns") or []:
        try:
            uri = str(
                run.get("chip", {})
                .get("richLinkProperties", {})
                .get("uri")
                or ""
            ).strip()
        except Exception:
            uri = ""
        if uri:
            candidatos.append(uri)

    valor = celda.get("userEnteredValue") or {}
    formula = str(valor.get("formulaValue") or "").strip()
    if formula:
        m = re.search(
            r'^\s*=\s*HYPERLINK\(\s*["\']([^"\']+)["\']',
            formula,
            flags=re.IGNORECASE,
        )
        if m:
            candidatos.append(m.group(1).strip())

    visible = str(celda.get("formattedValue") or "").strip()
    if visible.lower().startswith(("https://", "http://")):
        candidatos.append(visible)

    vistos = set()
    for url in candidatos:
        if not url or url in vistos:
            continue
        vistos.add(url)
        _, drive_id = _extraer_referencia_drive(url)
        if drive_id:
            return url

    return ""

def _leer_enlaces_columna_g_google_sheet(url_google_sheet, hoja_objetivo=None):
    """Lee SOLO enlaces Google Drive de la columna G (Enlaces).

    No usa valores de otras columnas para decidir que abrir en Drive. Tampoco
    descarga recursos: solamente recupera el link real de cada celda G,
    incluyendo Smart Chips.
    """
    spreadsheet_id = _extraer_spreadsheet_id(url_google_sheet)
    if not spreadsheet_id:
        raise ValueError("El enlace de Google Sheets no es valido.")

    credenciales = _cargar_credenciales_google()
    if credenciales is None:
        raise ValueError("Primero debes conectar tu cuenta de Google.")

    sheets = build(
        "sheets",
        "v4",
        credentials=credenciales,
        cache_discovery=False,
    )

    meta = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets.properties(title)",
    ).execute()

    titulos = [
        str(item.get("properties", {}).get("title") or "")
        for item in meta.get("sheets", [])
        if str(item.get("properties", {}).get("title") or "").strip()
    ]

    if hoja_objetivo:
        titulos = [
            titulo for titulo in titulos
            if _misma_hoja(titulo, hoja_objetivo)
        ]

    if not titulos:
        return {}

    rangos = [
        "'" + titulo.replace("'", "''") + "'!G:G"
        for titulo in titulos
    ]

    respuesta = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        ranges=rangos,
        includeGridData=True,
        fields=(
            "sheets(properties(title),"
            "data(startRow,rowData(values("
            "formattedValue,hyperlink,chipRuns,userEnteredValue"
            "))))"
        ),
    ).execute()

    enlaces_drive = {}

    for hoja in respuesta.get("sheets", []):
        titulo = str(hoja.get("properties", {}).get("title") or "")
        for bloque in hoja.get("data", []) or []:
            inicio = int(bloque.get("startRow") or 0)
            filas = bloque.get("rowData") or []
            for offset, fila in enumerate(filas):
                valores = fila.get("values") or []
                if not valores:
                    continue
                url_drive = _url_desde_cell_data(valores[0])
                if url_drive:
                    enlaces_drive[(titulo, inicio + offset + 1)] = url_drive

    print(
        "Sheets SOLO ENLACES:",
        len(titulos),
        "hoja(s) |",
        len(enlaces_drive),
        "enlace(s) Google Drive encontrados exclusivamente en columna G",
    )

    return enlaces_drive

def _clave_cache_drive(drive_id):
    """Las caches de Drive son por usuario.

    Cada persona tiene su propio token de Google y por tanto sus propios
    permisos: lo que una alcanza a ver no puede reutilizarse para otra.
    """
    return _normalizar_usuario(_usuario_actual()) + "|" + str(drive_id or "")


def _cache_drive_obtener(carpeta_id):
    try:
        with motor_prizma_modulo._CACHE_CARPETAS_DRIVE_LOCK:
            valor = motor_prizma_modulo._CACHE_CARPETAS_DRIVE.get(
                _clave_cache_drive(carpeta_id)
            )
    except Exception:
        return None

    # Las entradas guardadas por la version anterior no traen "subcarpetas"
    # y solo contienen hijos directos. Se descartan para volver a leerlas
    # con el recorrido nuevo.
    if isinstance(valor, dict) and "subcarpetas" not in valor:
        return None

    return valor


def _cache_drive_guardar(carpeta_id, resultado):
    try:
        with motor_prizma_modulo._CACHE_CARPETAS_DRIVE_LOCK:
            motor_prizma_modulo._CACHE_CARPETAS_DRIVE[
                _clave_cache_drive(carpeta_id)
            ] = resultado
    except Exception:
        pass


MIME_CARPETA_DRIVE = "application/vnd.google-apps.folder"
MIME_ATAJO_DRIVE = "application/vnd.google-apps.shortcut"

# Profundidad maxima al recorrer una carpeta de Drive.
# 0 = solo hijos directos (comportamiento anterior).
DRIVE_PROFUNDIDAD_MAX = 3


def _resultado_carpeta_vacio():
    return {
        "ok": True,
        "error": None,
        "archivos": [],
        "detalles": [],
        "subcarpetas": 0,
        "ignorados": [],
    }


def _agregar_detalle_carpeta(resultado, item, mime_type=None):
    """Agrega un archivo real (con tamano) al resultado de la carpeta."""
    nombre = str(item.get("name") or "").strip()
    if not nombre:
        return

    mime_type = str(mime_type or item.get("mimeType") or "").strip()

    tamano = item.get("size")
    try:
        tamano = int(tamano) if tamano is not None else None
    except (TypeError, ValueError):
        tamano = None

    identificador = str(item.get("id") or "").strip()

    for existente in resultado["detalles"]:
        if identificador and existente.get("id") == identificador:
            return

    resultado["archivos"].append(nombre)
    resultado["detalles"].append({
        "id": identificador,
        "nombre": nombre,
        "mimeType": mime_type,
        "tamano": tamano,
    })


def _normalizar_respuesta_carpeta_drive(respuesta):
    """Compatibilidad: normaliza una respuesta simple de files().list."""
    resultado = _resultado_carpeta_vacio()

    for item in (respuesta or {}).get("files", []):
        mime_type = str(item.get("mimeType") or "").strip()

        if mime_type == MIME_CARPETA_DRIVE:
            resultado["subcarpetas"] += 1
            continue

        if mime_type == MIME_ATAJO_DRIVE:
            continue

        _agregar_detalle_carpeta(resultado, item, mime_type)

    return resultado


def _listar_carpetas_drive_en_lote(drive, carpetas_raiz):
    """Lista por lotes el contenido de varias carpetas de Drive.

    A diferencia de la version anterior:
      - entra en subcarpetas (hasta DRIVE_PROFUNDIDAD_MAX niveles);
      - resuelve atajos (shortcuts) a archivos y a carpetas;
      - respeta nextPageToken (carpetas con muchos elementos).

    Sigue pidiendo unicamente id, name, mimeType y size:
    NO descarga ni abre ningun H5P/PDF.
    """
    carpetas_raiz = [c for c in dict.fromkeys(carpetas_raiz) if c]
    resultados = {raiz: _resultado_carpeta_vacio() for raiz in carpetas_raiz}

    if not carpetas_raiz:
        return resultados

    visitadas = {raiz: {raiz} for raiz in carpetas_raiz}
    atajos_a_archivo = {}

    # tarea = (raiz, carpeta_id, page_token, profundidad)
    pendientes = [(raiz, raiz, None, 0) for raiz in carpetas_raiz]

    while pendientes:
        siguientes = []

        for inicio in range(0, len(pendientes), DRIVE_BATCH_MAX):
            grupo = pendientes[inicio:inicio + DRIVE_BATCH_MAX]
            respuestas = {}

            def callback_lista(request_id, response, exception):
                if exception is not None:
                    respuestas[str(request_id)] = ("error", exception)
                else:
                    respuestas[str(request_id)] = ("ok", response)

            batch = drive.new_batch_http_request(callback=callback_lista)

            for indice, (raiz, carpeta_id, token, profundidad) in enumerate(grupo):
                batch.add(
                    drive.files().list(
                        q="'" + carpeta_id + "' in parents and trashed = false",
                        spaces="drive",
                        fields=(
                            "nextPageToken,files(id,name,mimeType,size,"
                            "shortcutDetails(targetId,targetMimeType))"
                        ),
                        pageSize=1000,
                        pageToken=token,
                        supportsAllDrives=True,
                        includeItemsFromAllDrives=True,
                    ),
                    request_id=str(indice),
                )

            batch.execute()

            for indice, tarea in enumerate(grupo):
                raiz, carpeta_id, token, profundidad = tarea
                estado, payload = respuestas.get(str(indice), ("error", None))

                if estado == "error":
                    if carpeta_id == raiz and token is None:
                        resultados[raiz]["ok"] = False
                        resultados[raiz]["error"] = "ERROR_CARPETA_RECURSO_NO_ACCESIBLE"
                    else:
                        resultados[raiz]["ignorados"].append({
                            "nombre": carpeta_id,
                            "mimeType": MIME_CARPETA_DRIVE,
                            "motivo": "subcarpeta no accesible",
                        })
                    continue

                payload = payload or {}

                for item in payload.get("files", []) or []:
                    nombre = str(item.get("name") or "").strip()
                    mime_type = str(item.get("mimeType") or "").strip()

                    if not nombre:
                        continue

                    if mime_type == MIME_CARPETA_DRIVE:
                        resultados[raiz]["subcarpetas"] += 1
                        hijo = str(item.get("id") or "").strip()
                        if (
                            hijo
                            and profundidad + 1 <= DRIVE_PROFUNDIDAD_MAX
                            and hijo not in visitadas[raiz]
                        ):
                            visitadas[raiz].add(hijo)
                            siguientes.append((raiz, hijo, None, profundidad + 1))
                        continue

                    if mime_type == MIME_ATAJO_DRIVE:
                        detalle = item.get("shortcutDetails") or {}
                        target_id = str(detalle.get("targetId") or "").strip()
                        target_mime = str(detalle.get("targetMimeType") or "").strip()

                        if not target_id:
                            continue

                        if target_mime == MIME_CARPETA_DRIVE:
                            if (
                                profundidad + 1 <= DRIVE_PROFUNDIDAD_MAX
                                and target_id not in visitadas[raiz]
                            ):
                                visitadas[raiz].add(target_id)
                                siguientes.append(
                                    (raiz, target_id, None, profundidad + 1)
                                )
                        else:
                            atajos_a_archivo.setdefault(target_id, set()).add(raiz)
                        continue

                    if item.get("size") is None:
                        resultados[raiz]["ignorados"].append({
                            "nombre": nombre,
                            "mimeType": mime_type,
                            "motivo": "archivo nativo de Google (sin tamano)",
                        })
                        continue

                    _agregar_detalle_carpeta(resultados[raiz], item, mime_type)

                token_siguiente = str(payload.get("nextPageToken") or "").strip()
                if token_siguiente:
                    siguientes.append((raiz, carpeta_id, token_siguiente, profundidad))

        pendientes = siguientes

    # Resolver los atajos que apuntan a archivos reales.
    objetivos = list(atajos_a_archivo.keys())

    for inicio in range(0, len(objetivos), DRIVE_BATCH_MAX):
        grupo = objetivos[inicio:inicio + DRIVE_BATCH_MAX]
        respuestas = {}

        def callback_atajo(request_id, response, exception):
            respuestas[str(request_id)] = None if exception is not None else response

        batch = drive.new_batch_http_request(callback=callback_atajo)

        for target_id in grupo:
            batch.add(
                drive.files().get(
                    fileId=target_id,
                    fields="id,name,mimeType,size",
                    supportsAllDrives=True,
                ),
                request_id=target_id,
            )

        batch.execute()

        for target_id in grupo:
            item = respuestas.get(target_id)
            if not item:
                continue
            for raiz in atajos_a_archivo.get(target_id, set()):
                if item.get("size") is None:
                    resultados[raiz]["ignorados"].append({
                        "nombre": str(item.get("name") or ""),
                        "mimeType": str(item.get("mimeType") or ""),
                        "motivo": "atajo a archivo sin tamano",
                    })
                    continue
                _agregar_detalle_carpeta(resultados[raiz], item)

    return resultados


def _listar_archivos_carpeta_drive_oauth(url_carpeta):
    """Fallback individual. Normalmente el preflight ya deja todo en cache."""

    carpeta_id = motor_prizma_modulo.extraer_id_carpeta_drive(url_carpeta)

    if not carpeta_id:
        return {
            "ok": False,
            "error": "ERROR_CARPETA_RECURSO_NO_ACCESIBLE",
            "archivos": [],
            "detalles": [],
        }

    cache = _cache_drive_obtener(carpeta_id)
    if cache is not None:
        return cache

    credenciales = _cargar_credenciales_google()
    if credenciales is None:
        return {
            "ok": False,
            "error": "ERROR_GOOGLE_NO_CONECTADO",
            "archivos": [],
            "detalles": [],
        }

    try:
        drive = build(
            "drive",
            "v3",
            credentials=credenciales,
            cache_discovery=False,
        )

        respuesta = drive.files().list(
            q=f"'{carpeta_id}' in parents and trashed = false",
            spaces="drive",
            fields="files(id,name,mimeType,size)",
            pageSize=1000,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()

        resultado = _normalizar_respuesta_carpeta_drive(respuesta)
        _cache_drive_guardar(carpeta_id, resultado)
        return resultado

    except Exception as error:
        print(
            "Drive API - error carpeta:",
            carpeta_id,
            type(error).__name__,
            str(error),
        )
        return {
            "ok": False,
            "error": "ERROR_CARPETA_RECURSO_NO_ACCESIBLE",
            "archivos": [],
            "detalles": [],
        }


def _precargar_referencias_drive_batch(actividades):
    """Precarga Drive rapido usando el tipo REAL del recurso, no solo la forma de la URL.

    Algunos Smart Chips usan enlaces tipo open?id=... incluso cuando apuntan a una
    carpeta. Primero consultamos metadatos de esas referencias ambiguas y solo despues
    decidimos si son carpeta o archivo. No se descarga ningun H5P/PDF.
    """

    carpetas_url = {}
    referencias_ambiguas = {}

    for actividad in actividades or []:
        url = str(actividad.get("_url_drive_sheet") or "").strip()
        tipo_ref, drive_id = _extraer_referencia_drive(url)
        if not drive_id:
            continue
        if tipo_ref == "folder":
            carpetas_url.setdefault(drive_id, url)
            with _CACHE_REFERENCIAS_DRIVE_LOCK:
                _CACHE_REFERENCIAS_DRIVE[_clave_cache_drive(drive_id)] = {
                    "tipo": "folder",
                    "target_id": drive_id,
                }
        else:
            referencias_ambiguas.setdefault(drive_id, url)

    credenciales = _cargar_credenciales_google()
    if credenciales is None:
        raise ValueError(
            "Primero debes conectar tu cuenta de Google antes de analizar la matriz."
        )

    drive = build(
        "drive",
        "v3",
        credentials=credenciales,
        cache_discovery=False,
    )

    # --------------------------------------------------------
    # 1) Resolver referencias ambiguas por SU MIMETYPE REAL.
    #    open?id=... puede ser archivo o carpeta.
    # --------------------------------------------------------
    ambiguas_pendientes = []
    with _CACHE_REFERENCIAS_DRIVE_LOCK:
        for drive_id in referencias_ambiguas:
            if _clave_cache_drive(drive_id) not in _CACHE_REFERENCIAS_DRIVE:
                ambiguas_pendientes.append(drive_id)

    metadatos_ambiguos = {}

    for inicio in range(0, len(ambiguas_pendientes), DRIVE_BATCH_MAX):
        grupo = ambiguas_pendientes[inicio:inicio + DRIVE_BATCH_MAX]
        respuestas = {}

        def callback_meta(request_id, response, exception):
            respuestas[str(request_id)] = None if exception is not None else response

        batch = drive.new_batch_http_request(callback=callback_meta)
        for drive_id in grupo:
            batch.add(
                drive.files().get(
                    fileId=drive_id,
                    fields="id,name,mimeType,size,shortcutDetails(targetId,targetMimeType)",
                    supportsAllDrives=True,
                ),
                request_id=drive_id,
            )
        batch.execute()

        for drive_id in grupo:
            metadatos_ambiguos[drive_id] = respuestas.get(drive_id)

    # Resolver shortcuts a archivos reales solo cuando existan.
    shortcut_archivos = {}
    for original_id, item in metadatos_ambiguos.items():
        if not item:
            continue
        mime_type = str(item.get("mimeType") or "")
        if mime_type == "application/vnd.google-apps.shortcut":
            detalle = item.get("shortcutDetails") or {}
            target_id = str(detalle.get("targetId") or "").strip()
            target_mime = str(detalle.get("targetMimeType") or "").strip()
            if target_id and target_mime != "application/vnd.google-apps.folder":
                shortcut_archivos[original_id] = target_id

    metadatos_targets = {}
    targets_unicos = list(dict.fromkeys(shortcut_archivos.values()))
    for inicio in range(0, len(targets_unicos), DRIVE_BATCH_MAX):
        grupo = targets_unicos[inicio:inicio + DRIVE_BATCH_MAX]
        respuestas = {}

        def callback_target(request_id, response, exception):
            respuestas[str(request_id)] = None if exception is not None else response

        batch = drive.new_batch_http_request(callback=callback_target)
        for target_id in grupo:
            batch.add(
                drive.files().get(
                    fileId=target_id,
                    fields="id,name,mimeType,size",
                    supportsAllDrives=True,
                ),
                request_id=target_id,
            )
        batch.execute()
        metadatos_targets.update(respuestas)

    carpetas_reales = dict(carpetas_url)
    archivos_reales = 0
    carpetas_detectadas_desde_enlace_archivo = 0
    shortcuts_detectados = 0

    with _CACHE_REFERENCIAS_DRIVE_LOCK, _CACHE_ARCHIVOS_DRIVE_LOCK:
        for drive_id in referencias_ambiguas:
            item = metadatos_ambiguos.get(drive_id)

            if not item:
                _CACHE_REFERENCIAS_DRIVE[_clave_cache_drive(drive_id)] = {
                    "tipo": "error",
                    "target_id": drive_id,
                }
                _CACHE_ARCHIVOS_DRIVE[_clave_cache_drive(drive_id)] = None
                continue

            mime_type = str(item.get("mimeType") or "").strip()
            target_id = drive_id

            if mime_type == "application/vnd.google-apps.shortcut":
                shortcuts_detectados += 1
                detalle = item.get("shortcutDetails") or {}
                target_id = str(detalle.get("targetId") or "").strip()
                target_mime = str(detalle.get("targetMimeType") or "").strip()

                if target_id and target_mime == "application/vnd.google-apps.folder":
                    carpetas_reales.setdefault(target_id, referencias_ambiguas[drive_id])
                    _CACHE_REFERENCIAS_DRIVE[_clave_cache_drive(drive_id)] = {
                        "tipo": "folder",
                        "target_id": target_id,
                    }
                    carpetas_detectadas_desde_enlace_archivo += 1
                    continue

                item = metadatos_targets.get(target_id)
                if not item:
                    _CACHE_REFERENCIAS_DRIVE[_clave_cache_drive(drive_id)] = {
                        "tipo": "error",
                        "target_id": target_id or drive_id,
                    }
                    _CACHE_ARCHIVOS_DRIVE[_clave_cache_drive(drive_id)] = None
                    continue
                mime_type = str(item.get("mimeType") or "").strip()

            if mime_type == "application/vnd.google-apps.folder":
                carpetas_reales.setdefault(target_id, referencias_ambiguas[drive_id])
                _CACHE_REFERENCIAS_DRIVE[_clave_cache_drive(drive_id)] = {
                    "tipo": "folder",
                    "target_id": target_id,
                }
                carpetas_detectadas_desde_enlace_archivo += 1
                continue

            tamano = item.get("size")
            try:
                tamano = int(tamano) if tamano is not None else None
            except (TypeError, ValueError):
                tamano = None

            archivo = {
                "id": str(item.get("id") or target_id or drive_id).strip(),
                "nombre": str(item.get("name") or "").strip(),
                "mimeType": mime_type,
                "tamano": tamano,
            }

            _CACHE_ARCHIVOS_DRIVE[_clave_cache_drive(drive_id)] = archivo
            _CACHE_REFERENCIAS_DRIVE[_clave_cache_drive(drive_id)] = {
                "tipo": "file",
                "target_id": archivo["id"],
                "item": archivo,
            }
            archivos_reales += 1

    # --------------------------------------------------------
    # 2) Listar hijos SOLO de las referencias que son carpetas
    #    reales. Ahora tambien se entra en subcarpetas y atajos,
    #    pero seguimos pidiendo unicamente nombre + tamano.
    # --------------------------------------------------------
    carpetas_pendientes = [
        drive_id for drive_id in carpetas_reales
        if _cache_drive_obtener(drive_id) is None
    ]

    carpetas_con_subcarpetas = 0
    carpetas_vacias = 0

    if carpetas_pendientes:
        resultados_carpetas = _listar_carpetas_drive_en_lote(
            drive,
            carpetas_pendientes,
        )

        for carpeta_id in carpetas_pendientes:
            resultado = resultados_carpetas.get(carpeta_id)

            if not resultado:
                resultado = {
                    "ok": False,
                    "error": "ERROR_CARPETA_RECURSO_NO_ACCESIBLE",
                    "archivos": [],
                    "detalles": [],
                    "subcarpetas": 0,
                    "ignorados": [],
                }

            if resultado.get("subcarpetas"):
                carpetas_con_subcarpetas += 1

            if resultado.get("ok") and not resultado.get("detalles"):
                carpetas_vacias += 1

            _cache_drive_guardar(carpeta_id, resultado)

    print(
        "Drive rapido tipo-real:",
        len(carpetas_url),
        "carpeta(s) por URL +",
        len(referencias_ambiguas),
        "referencia(s) ambigua(s) |",
        carpetas_detectadas_desde_enlace_archivo,
        "eran carpeta(s) |",
        archivos_reales,
        "eran archivo(s) |",
        shortcuts_detectados,
        "shortcut(s) | solo nombre + tamano",
    )

    print(
        "Drive estructura:",
        len(carpetas_pendientes),
        "carpeta(s) leida(s) |",
        carpetas_con_subcarpetas,
        "con subcarpetas (se recorrieron hasta",
        DRIVE_PROFUNDIDAD_MAX,
        "nivel(es)) |",
        carpetas_vacias,
        "sin ningun archivo con tamano",
    )

    return {
        "total_carpetas": len(carpetas_reales),
        "total_archivos_directos": archivos_reales,
        "consultadas_google": len(ambiguas_pendientes) + len(carpetas_pendientes) + len(targets_unicos),
    }

_EQUIVALENCIAS_UNICODE = {
    0x2010: "-",
    0x2011: "-",
    0x2012: "-",
    0x2013: "-",
    0x2014: "-",
    0x2015: "-",
    0x2212: "-",
    0x2018: "'",
    0x2019: "'",
    0x201C: '"',
    0x201D: '"',
    0x00A0: " ",
    0x2007: " ",
    0x202F: " ",
}


def _nombre_base_archivo(nombre):
    """Ultimo componente del nombre, sin importar si separa con / o con \\."""
    texto = str(nombre or "").replace("\\", "/")
    return texto.rsplit("/", 1)[-1].strip()


def _clave_nombre_drive_zip(nombre):
    """Nombre exacto de archivo, normalizando solo Unicode NFC."""
    return unicodedata.normalize("NFC", _nombre_base_archivo(nombre))


def _clave_nombre_flexible(nombre):
    """Mismo nombre escrito de forma equivalente.

    Tolera mayusculas/minusculas, guiones tipograficos, comillas curvas,
    espacios duros y espacios repetidos. NO es busqueda difusa: sigue
    exigiendo el mismo nombre, no un nombre parecido.
    """
    texto = unicodedata.normalize("NFC", _nombre_base_archivo(nombre))
    texto = texto.translate(_EQUIVALENCIAS_UNICODE)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto.casefold()


def _indice_zip_nombre_tamano(ruta_zip):
    """Indexa el ZIP por nombre exacto y por nombre equivalente."""
    indice = {
        "exacto": {},
        "flexible": {},
        "total": 0,
    }

    with zipfile.ZipFile(ruta_zip, "r") as archivo_zip:
        for info in archivo_zip.infolist():
            if info.is_dir():
                continue

            nombre = _nombre_base_archivo(info.filename)
            extension = os.path.splitext(nombre)[1].lower()

            if extension not in (".h5p", ".pdf"):
                continue

            entrada = {
                "nombre": nombre,
                "tamano": int(info.file_size),
                "miembro": info.filename,
            }

            indice["exacto"].setdefault(
                _clave_nombre_drive_zip(nombre),
                [],
            ).append(entrada)

            indice["flexible"].setdefault(
                _clave_nombre_flexible(nombre),
                [],
            ).append(entrada)

            indice["total"] += 1

    return indice


def _buscar_en_zip(indice_zip, nombre_drive, tamano_drive):
    """Empareja UN archivo de Drive contra el ZIP.

    Devuelve (estado, entradas):
      "ok"              -> entradas = [entrada unica del ZIP]
      "duplicado"       -> el ZIP tiene varias copias distintas iguales en nombre+tamano
      "tamano_distinto" -> el nombre existe en el ZIP pero con otro tamano
      "sin_nombre"      -> ese nombre no existe en el ZIP
    """
    try:
        tamano_drive = int(tamano_drive)
    except (TypeError, ValueError):
        return "sin_nombre", []

    candidatos = list(
        indice_zip["exacto"].get(_clave_nombre_drive_zip(nombre_drive), [])
    )

    if not candidatos:
        candidatos = list(
            indice_zip["flexible"].get(_clave_nombre_flexible(nombre_drive), [])
        )

    if not candidatos:
        return "sin_nombre", []

    iguales = [
        entrada for entrada in candidatos
        if int(entrada.get("tamano") or -1) == tamano_drive
    ]

    if len(iguales) == 1:
        return "ok", iguales

    if len(iguales) > 1:
        miembros = {str(entrada.get("miembro") or "") for entrada in iguales}
        if len(miembros) == 1:
            return "ok", iguales[:1]
        return "duplicado", iguales

    return "tamano_distinto", candidatos


def _desempatar_por_tipo_declarado(actividad, coincidencias):
    """Si la matriz declara H5P o PDF y solo un candidato tiene esa extension,
    se toma ese. No se usa el nombre de la actividad: solo el tipo declarado.
    """
    tipo = str(actividad.get("tipo_archivo") or "").strip().upper()
    extension = {"H5P": ".h5p", "PDF": ".pdf"}.get(tipo)

    if not extension:
        return coincidencias

    filtradas = [
        par for par in coincidencias
        if os.path.splitext(str(par[0].get("nombre") or ""))[1].lower() == extension
    ]

    return filtradas if len(filtradas) == 1 else coincidencias


def _explicar_fallas_drive_zip(fallas):
    """Convierte los fallos por archivo en un mensaje que dice la causa real."""
    if not fallas:
        return "la carpeta de Drive no tiene archivos .h5p o .pdf comparables con el ZIP."

    distintos = [f for f in fallas if f[1] == "tamano_distinto"]
    duplicados = [f for f in fallas if f[1] == "duplicado"]
    faltantes = [f for f in fallas if f[1] == "sin_nombre"]

    partes = []

    for item, _, entradas in distintos[:3]:
        tamanos = ", ".join(
            str(int(entrada.get("tamano") or 0)) for entrada in entradas[:3]
        )
        partes.append(
            '"' + str(item.get("nombre") or "") + '" esta en Drive con '
            + str(int(item.get("tamano") or 0)) + " bytes y en el ZIP con "
            + tamanos + " bytes"
        )

    for item, _, _ in duplicados[:2]:
        partes.append(
            '"' + str(item.get("nombre") or "")
            + '" aparece duplicado en el ZIP con el mismo tamano'
        )

    if faltantes:
        nombres = ", ".join(
            '"' + str(item.get("nombre") or "") + '"'
            for item, _, _ in faltantes[:3]
        )
        extra = "" if len(faltantes) <= 3 else " y " + str(len(faltantes) - 3) + " mas"
        partes.append("el ZIP no contiene " + nombres + extra)

    return "; ".join(partes) + "."


def _registrar_diagnostico_drive(
    diagnostico,
    descripcion,
    url_drive,
    drive_id,
    motivo,
    resultado_drive=None,
    fallas=None,
):
    """Guarda hasta 10 filas fallidas con TODO lo necesario para clasificarlas."""
    if len(diagnostico) >= 10:
        return

    resultado_drive = resultado_drive or {}
    detalles = resultado_drive.get("detalles") or []

    diagnostico.append({
        "fila": descripcion,
        "valor_en_G": url_drive,
        "drive_id": drive_id,
        "motivo": motivo,
        "subcarpetas_vistas": resultado_drive.get("subcarpetas", 0),
        "archivos_en_drive": [
            str(item.get("nombre") or "")
            + " | " + str(item.get("mimeType") or "")
            + " | " + str(item.get("tamano"))
            for item in detalles[:8]
        ],
        "ignorados_en_drive": [
            str(item.get("nombre") or "") + " | " + str(item.get("motivo") or "")
            for item in (resultado_drive.get("ignorados") or [])[:5]
        ],
        "resultado_vs_zip": [
            str(item.get("nombre") or "") + " -> " + str(estado)
            for item, estado, _ in (fallas or [])[:8]
        ],
    })


def _prevalidar_recursos_drive_zip_ultrarapido(
    url_google_sheet,
    ruta_excel,
    ruta_zip,
    procesar_ovi=True,
    procesar_ova=True,
    procesar_retos=True,
    hoja_objetivo=None,
):
    """Valida Drive SOLO desde la columna G y SOLO por nombre + tamano.

    Reglas:
    - Drive se consulta unicamente si esa misma fila tiene un enlace Drive real
      en la columna G obtenido directamente con la API de Sheets.
    - No se usa el nombre de la actividad para adivinar el recurso.
    - No se usa OVI/OVA/RETO para decidir que archivo tomar de Drive.
    - No se abre ni se descarga ningun H5P/PDF de Drive.
    - Carpeta: se listan solo metadatos de sus hijos y se busca UN archivo que
      exista en el ZIP con el mismo nombre exacto y el mismo tamano exacto.
    - Archivo directo: solo participa si es .h5p/.pdf y tiene tamano; otros
      enlaces Drive se ignoran y el motor conserva su resolucion historica ZIP.
    - Filas sin enlace Drive en G NO son error.
    """

    _DRIVE_THREAD_LOCAL.hoja_objetivo = hoja_objetivo

    actividades = motor_prizma_modulo.leer_actividades_excel(
        ruta_excel,
        procesar_ovi,
        procesar_ova,
        procesar_retos,
    )

    if not actividades:
        return {
            "ok": False,
            "total_actividades": 0,
            "total_enlaces_drive": 0,
            "validadas_drive": 0,
            "errores": ["No se encontraron actividades compatibles en la matriz."],
            "recursos": [],
            "_mapa_resueltos": {},
        }

    indice_zip = _indice_zip_nombre_tamano(ruta_zip)
    if not indice_zip["total"]:
        return {
            "ok": False,
            "total_actividades": len(actividades),
            "total_enlaces_drive": 0,
            "validadas_drive": 0,
            "errores": ["El ZIP no contiene archivos H5P o PDF."],
            "recursos": [],
            "_mapa_resueltos": {},
        }

    # Fuente unica de Drive: columna G leida directamente por Sheets API.
    enlaces_drive = _leer_enlaces_columna_g_google_sheet(
        url_google_sheet,
        hoja_objetivo,
    )

    actividades_con_drive = []
    for actividad in actividades:
        clave_fila = (
            str(actividad.get("hoja") or ""),
            int(actividad.get("fila_excel") or 0),
        )
        url_drive = str(enlaces_drive.get(clave_fila) or "").strip()
        actividad["_url_drive_sheet"] = url_drive
        if url_drive:
            actividades_con_drive.append(actividad)

    # Precarga por lotes SOLO las filas realmente enlazadas a Drive.
    if actividades_con_drive:
        _precargar_referencias_drive_batch(actividades_con_drive)

    errores = []
    recursos = []
    mapa_resueltos = {}
    ignorados_archivo_no_recurso = 0
    diagnostico = []

    for actividad in actividades_con_drive:
        descripcion = (
            f'Hoja {actividad.get("hoja", "")} · '
            f'fila {actividad.get("fila_excel", "")} · '
            f'{actividad.get("nombre", "")}'
        )

        url_drive = str(actividad.get("_url_drive_sheet") or "").strip()
        _, drive_id = _extraer_referencia_drive(url_drive)

        if not drive_id:
            # En teoria no ocurre porque el lector ya filtra solo Drive.
            continue

        with _CACHE_REFERENCIAS_DRIVE_LOCK:
            referencia = dict(
                _CACHE_REFERENCIAS_DRIVE.get(_clave_cache_drive(drive_id))
                or {}
            )

        tipo_ref = str(referencia.get("tipo") or "")
        target_id = str(referencia.get("target_id") or drive_id)
        elegido = None
        zip_elegido = None

        if tipo_ref == "folder":
            resultado_drive = _cache_drive_obtener(target_id)

            if not resultado_drive or not resultado_drive.get("ok"):
                errores.append(
                    descripcion
                    + ": no fue posible leer los metadatos de la carpeta Drive "
                    + "(revisa permisos de lectura sobre esa carpeta)."
                )
                _registrar_diagnostico_drive(
                    diagnostico,
                    descripcion,
                    url_drive,
                    target_id,
                    "carpeta no legible",
                    resultado_drive,
                )
                continue

            detalles = resultado_drive.get("detalles") or []

            candidatos_drive = []
            for item in detalles:
                nombre_item = str(item.get("nombre") or "").strip()
                extension = os.path.splitext(nombre_item)[1].lower()

                if extension not in (".h5p", ".pdf"):
                    continue
                if item.get("tamano") is None:
                    continue

                candidatos_drive.append(item)

            if not candidatos_drive:
                errores.append(
                    descripcion
                    + ": la carpeta de Drive no contiene ningun .h5p ni .pdf "
                    + "(elementos con tamano leidos: " + str(len(detalles))
                    + ", subcarpetas recorridas: "
                    + str(resultado_drive.get("subcarpetas", 0)) + ")."
                )
                _registrar_diagnostico_drive(
                    diagnostico,
                    descripcion,
                    url_drive,
                    target_id,
                    "carpeta sin h5p/pdf",
                    resultado_drive,
                )
                continue

            coincidencias = []
            fallas = []

            for item in candidatos_drive:
                estado, entradas = _buscar_en_zip(
                    indice_zip,
                    str(item.get("nombre") or ""),
                    item.get("tamano"),
                )

                if estado == "ok":
                    coincidencias.append((item, entradas[0]))
                else:
                    fallas.append((item, estado, entradas))

            if len(coincidencias) > 1:
                coincidencias = _desempatar_por_tipo_declarado(
                    actividad,
                    coincidencias,
                )

            if len(coincidencias) == 0:
                errores.append(
                    descripcion + ": " + _explicar_fallas_drive_zip(fallas)
                )
                _registrar_diagnostico_drive(
                    diagnostico,
                    descripcion,
                    url_drive,
                    target_id,
                    "sin coincidencia nombre+tamano",
                    resultado_drive,
                    fallas,
                )
                continue

            if len(coincidencias) > 1:
                nombres = ", ".join(
                    str(item.get("nombre") or "")
                    for item, _ in coincidencias[:6]
                )
                errores.append(
                    descripcion
                    + ": la carpeta tiene mas de un recurso que coincide con el ZIP "
                    + "por nombre y tamano (" + nombres + "). "
                    + "No se elige ninguno para no adivinar."
                )
                _registrar_diagnostico_drive(
                    diagnostico,
                    descripcion,
                    url_drive,
                    target_id,
                    "ambiguo: varios coinciden",
                    resultado_drive,
                    fallas,
                )
                continue

            elegido, zip_elegido = coincidencias[0]

        elif tipo_ref == "file":
            elegido = referencia.get("item")
            if not elegido:
                with _CACHE_ARCHIVOS_DRIVE_LOCK:
                    elegido = _CACHE_ARCHIVOS_DRIVE.get(
                        _clave_cache_drive(drive_id)
                    )

            if not elegido:
                errores.append(
                    descripcion + ": no fue posible leer los metadatos del archivo Drive."
                )
                _registrar_diagnostico_drive(
                    diagnostico,
                    descripcion,
                    url_drive,
                    target_id,
                    "archivo no legible",
                )
                continue

            nombre_drive = str(elegido.get("nombre") or "").strip()
            tamano_drive = elegido.get("tamano")
            extension = os.path.splitext(nombre_drive)[1].lower()

            # Un Drive directo que no sea un recurso cargable no interviene.
            # Esto evita falsos conflictos con Docs, Sheets, enlaces auxiliares, etc.
            if extension not in (".h5p", ".pdf") or tamano_drive is None:
                ignorados_archivo_no_recurso += 1
                continue

            estado, entradas = _buscar_en_zip(
                indice_zip,
                nombre_drive,
                tamano_drive,
            )

            if estado != "ok":
                errores.append(
                    descripcion + ": "
                    + _explicar_fallas_drive_zip([(elegido, estado, entradas)])
                )
                _registrar_diagnostico_drive(
                    diagnostico,
                    descripcion,
                    url_drive,
                    target_id,
                    "archivo directo sin coincidencia",
                    {"detalles": [elegido]},
                    [(elegido, estado, entradas)],
                )
                continue

            zip_elegido = entradas[0]

        else:
            # Si Google devuelve un link que no es recurso utilizable, no hacemos
            # una segunda busqueda ni usamos otra columna: se deja al flujo ZIP.
            continue

        nombre_drive = str(elegido.get("nombre") or "").strip()
        tamano_drive = int(elegido.get("tamano") or 0)

        if not nombre_drive or not zip_elegido:
            continue

        actividad["nombre_recurso_drive"] = nombre_drive

        resuelto = {
            "nombre": nombre_drive,
            "tamano": tamano_drive,
            "tipo_ref": tipo_ref,
            "drive_id": target_id,
        }
        mapa_resueltos[_clave_actividad_drive(actividad)] = resuelto

        recursos.append({
            "hoja": actividad.get("hoja", ""),
            "fila": actividad.get("fila_excel", 0),
            "actividad": actividad.get("nombre", ""),
            "archivo": nombre_drive,
            "tamano": tamano_drive,
        })

    if diagnostico:
        print("")
        print("=" * 78)
        print("DIAGNOSTICO DRIVE vs ZIP - primeras", len(diagnostico), "fila(s) con problema")
        print("=" * 78)
        for numero, caso in enumerate(diagnostico, start=1):
            print("")
            print(str(numero) + ") " + str(caso.get("fila")))
            print("   motivo            :", caso.get("motivo"))
            print("   valor en G        :", caso.get("valor_en_G"))
            print("   drive id          :", caso.get("drive_id"))
            print("   subcarpetas vistas:", caso.get("subcarpetas_vistas"))
            print("   archivos en Drive (nombre | mime | bytes):")
            for linea in caso.get("archivos_en_drive") or ["   (ninguno)"]:
                print("      -", linea)
            if caso.get("ignorados_en_drive"):
                print("   ignorados en Drive:")
                for linea in caso["ignorados_en_drive"]:
                    print("      -", linea)
            if caso.get("resultado_vs_zip"):
                print("   comparacion contra el ZIP:")
                for linea in caso["resultado_vs_zip"]:
                    print("      -", linea)
        print("")
        print("=" * 78)
        print("")

    print(
        "ZIP indexado:",
        indice_zip["total"],
        "archivo(s) .h5p/.pdf",
    )

    print(
        "Drive SOLO ENLACES:",
        len(actividades_con_drive),
        "fila(s) con Drive en G |",
        len(recursos),
        "recurso(s) confirmados por nombre+tamano |",
        ignorados_archivo_no_recurso,
        "archivo(s) Drive no-H5P/PDF ignorado(s) |",
        len(actividades) - len(actividades_con_drive),
        "actividad(es) sin Drive en G: flujo ZIP normal",
    )

    return {
        "ok": len(errores) == 0,
        "total_actividades": len(actividades),
        "total_enlaces_drive": len(actividades_con_drive),
        "validadas_drive": len(recursos),
        "ignoradas_drive_no_recurso": ignorados_archivo_no_recurso,
        "errores": errores,
        "recursos": recursos,
        "_mapa_resueltos": mapa_resueltos,
    }

def _validar_drive_desde_preflight(actividad):
    """Drive ya fue resuelto antes de abrir PRIZMA.

    Si la fila tenia un enlace Drive de recurso, reutilizamos el nombre exacto.
    Si no lo tenia, NO volvemos a Drive y dejamos que resolver_recurso use el
    flujo historico del ZIP.
    """
    resuelto = _recurso_resuelto_thread(actividad)
    if resuelto:
        actividad["nombre_recurso_drive"] = resuelto["nombre"]
    return None

_LEER_ACTIVIDADES_EXCEL_ORIGINAL = motor_prizma_modulo.leer_actividades_excel


def _leer_actividades_excel_hoja_objetivo(
    ruta_excel,
    procesar_ovi=True,
    procesar_ova=True,
    procesar_retos=True,
):
    """Igual que el lector original, pero limitado a la pestaña del enlace.

    El motor llama a leer_actividades_excel como global del modulo, asi que
    reemplazarlo aqui cubre tanto el preflight como el cargue a PRIZMA.
    """
    actividades = _LEER_ACTIVIDADES_EXCEL_ORIGINAL(
        ruta_excel,
        procesar_ovi,
        procesar_ova,
        procesar_retos,
    )

    objetivo = getattr(_DRIVE_THREAD_LOCAL, "hoja_objetivo", None)

    if not objetivo:
        return actividades

    return [
        actividad for actividad in actividades
        if _misma_hoja(actividad.get("hoja"), objetivo)
    ]


motor_prizma_modulo.leer_actividades_excel = _leer_actividades_excel_hoja_objetivo


# FALTABA: esta referencia no estaba definida en ninguna parte del proyecto.
# Si _resolver_recurso_desde_preflight llegaba a ejecutarse para una fila SIN
# Drive en la columna G, reventaba con NameError.
_RESOLVER_RECURSO_ORIGINAL = motor_prizma_modulo.resolver_recurso


def _resolver_recurso_desde_preflight(
    actividad,
    indice_recursos,
    carpeta_temp,
    actividades_curso,
):
    resuelto = _recurso_resuelto_thread(actividad)
    if not resuelto:
        return _RESOLVER_RECURSO_ORIGINAL(
            actividad,
            indice_recursos,
            carpeta_temp,
            actividades_curso,
        )

    nombre_esperado = str(resuelto.get("nombre") or "").strip()
    nombre_esperado_clave = _clave_nombre_drive_zip(nombre_esperado)

    exactos = [
        recurso
        for recurso in indice_recursos
        if _clave_nombre_drive_zip(recurso.get("nombre")) == nombre_esperado_clave
    ]

    if len(exactos) == 0:
        return None, "ERROR_RECURSO_DRIVE_NO_ESTA_EN_ZIP"
    if len(exactos) != 1:
        return None, "ERROR_RECURSO_DUPLICADO"

    elegido = exactos[0]
    carpeta_fila = os.path.join(
        carpeta_temp,
        "fila_" + str(actividad.get("fila_excel") or 0),
    )
    os.makedirs(carpeta_fila, exist_ok=True)
    ruta_destino = os.path.join(carpeta_fila, elegido["nombre"])

    try:
        with zipfile.ZipFile(elegido["zip"], "r") as zip_ref:
            info = zip_ref.getinfo(elegido["miembro"])
            tamano_esperado = int(resuelto.get("tamano") or -1)
            if tamano_esperado >= 0 and int(info.file_size) != tamano_esperado:
                return None, "ERROR_RECURSO_DRIVE_TAMANO_CAMBIO"

            with zip_ref.open(elegido["miembro"]) as origen:
                with open(ruta_destino, "wb") as destino:
                    shutil.copyfileobj(origen, destino)
    except KeyError:
        return None, "ERROR_RECURSO_DRIVE_NO_ESTA_EN_ZIP"
    except Exception:
        return None, "ERROR_EXTRAYENDO_RECURSO"

    if not os.path.isfile(ruta_destino):
        return None, "ERROR_EXTRAYENDO_RECURSO"

    print(
        "Recurso exacto prevalidado Drive -> ZIP:",
        elegido["nombre"],
        "|",
        int(resuelto.get("tamano") or 0),
        "bytes",
    )
    return {
        "ruta": ruta_destino,
        "nombre_original": elegido["nombre"],
        "puntuacion": 10000,
    }, None

# ============================================================
# TRABAJOS
# ============================================================

TRABAJOS = {}
HISTORIAL_LOCK = threading.Lock()
TRABAJOS_LOCK = threading.RLock()
COLA_CARGUES = deque()
COLA_CONDICION = threading.Condition(TRABAJOS_LOCK)

# Cuantos cargues pueden correr al mismo tiempo.
# 0 = sin limite: cada cargue arranca en su propio hilo apenas se pide,
# sin importar cuantos haya corriendo ya.
# Cada cargue abre su propio Chromium, asi que si la maquina se queda
# sin RAM se puede poner un tope con la variable de entorno
# AUTO_PRIZMA_MAX_SIMULTANEOS.
try:
    MAX_CARGUES_SIMULTANEOS = int(
        os.environ.get("AUTO_PRIZMA_MAX_SIMULTANEOS") or 0
    )
except ValueError:
    MAX_CARGUES_SIMULTANEOS = 0

CARGUES_ACTIVOS = 0
ZONA_HORARIA_COLOMBIA = ZoneInfo("America/Bogota")


def _ahora_colombia():
    return datetime.now(ZONA_HORARIA_COLOMBIA)


def _ahora_colombia_iso():
    return _ahora_colombia().isoformat(timespec="seconds")


def _sanitizar_parte_nombre(valor, limite=70):
    texto = str(valor or "").strip()
    texto = re.sub(r'[\\/:*?"<>|]+', "-", texto)
    texto = re.sub(r"\s+", " ", texto).strip(" .-")
    if not texto:
        texto = "Sin nombre"
    return texto[:limite].rstrip()


def _cursos_desde_hojas(hojas):
    cursos = []
    vistos = set()

    for hoja in hojas or []:
        curso = str(hoja.get("curso") or "").strip() or "Curso sin nombre"
        programa = str(hoja.get("programa") or "").strip() or "Programa sin nombre"
        clave = (curso.casefold(), programa.casefold())
        if clave in vistos:
            continue
        vistos.add(clave)
        cursos.append({"curso": curso, "programa": programa})

    return cursos


def _nombre_reporte(cursos, trabajo_id):
    fecha = _ahora_colombia().strftime("%Y-%m-%d_%H-%M-%S")

    if cursos:
        curso = _sanitizar_parte_nombre(cursos[0]["curso"], 58)
        programa = _sanitizar_parte_nombre(cursos[0]["programa"], 58)
        extra = ""
        if len(cursos) > 1:
            extra = f"_y_{len(cursos) - 1}_curso_mas"
        return f"{fecha} - {curso} - {programa}{extra} - {trabajo_id[:6]}.csv"

    return f"{fecha} - resultado_prizma - {trabajo_id[:6]}.csv"


def _cargar_historial():
    if not os.path.isfile(HISTORIAL_PATH):
        return []

    try:
        with open(HISTORIAL_PATH, "r", encoding="utf-8") as archivo:
            datos = json.load(archivo)
        return datos if isinstance(datos, list) else []
    except Exception:
        return []


def _guardar_historial(registros):
    temporal = HISTORIAL_PATH + ".tmp"
    with open(temporal, "w", encoding="utf-8") as archivo:
        json.dump(registros, archivo, ensure_ascii=False, indent=2)
    os.replace(temporal, HISTORIAL_PATH)


def _registrar_reporte_final(trabajo):
    ruta_reporte = trabajo.get("ruta_reporte")
    if not ruta_reporte or not os.path.isfile(ruta_reporte):
        return

    with HISTORIAL_LOCK:
        registros = _cargar_historial()
        trabajo_id = trabajo.get("id")

        if any(r.get("id") == trabajo_id for r in registros):
            trabajo["historial_registrado"] = True
            return

        fecha_iso = _ahora_colombia_iso()
        registros.append({
            "id": trabajo_id,
            "fecha_iso": fecha_iso,
            "archivo_reporte": os.path.basename(ruta_reporte),
            "cursos": trabajo.get("cursos", []),
            "estado_final": trabajo.get("etapa", "finalizado"),
            "usuario_app": trabajo.get("usuario_app"),
        })
        _guardar_historial(registros)
        trabajo["historial_registrado"] = True


def ejecutar_cargue_con_historial(
    ruta_excel,
    ruta_zip,
    carpeta_temp,
    ruta_reporte,
    procesar_ovi,
    procesar_ova,
    procesar_retos,
    usuario_prizma,
    contrasena_prizma,
    trabajo,
):
    _DRIVE_THREAD_LOCAL.recursos = trabajo.get(
        "recursos_drive_resueltos",
        {},
    )
    _DRIVE_THREAD_LOCAL.hoja_objetivo = trabajo.get(
        "hoja_objetivo",
    )

    # El hilo del worker no hereda el contexto de la peticion.
    testigo_usuario = USUARIO_ACTUAL.set(trabajo.get("usuario_app"))

    try:
        ejecutar_cargue(
            ruta_excel,
            ruta_zip,
            carpeta_temp,
            ruta_reporte,
            procesar_ovi,
            procesar_ova,
            procesar_retos,
            usuario_prizma,
            contrasena_prizma,
            trabajo,
        )
    finally:
        try:
            _DRIVE_THREAD_LOCAL.recursos = {}
            _DRIVE_THREAD_LOCAL.hoja_objetivo = None
            USUARIO_ACTUAL.reset(testigo_usuario)
        except Exception:
            pass
        _registrar_reporte_final(trabajo)


# ============================================================
# COLA GLOBAL DE CARGUES - HASTA 2 PROCESOS A LA VEZ
# ============================================================

def _posicion_en_cola(trabajo_id):
    with TRABAJOS_LOCK:
        pendientes = [
            tid for tid in COLA_CARGUES
            if tid in TRABAJOS
            and TRABAJOS[tid].get("etapa") == "en_cola"
            and not TRABAJOS[tid].get("terminado")
        ]
        try:
            return pendientes.index(trabajo_id) + 1
        except ValueError:
            return None


def _encolar_trabajo(trabajo_id, usuario_prizma, contrasena_prizma):
    with COLA_CONDICION:
        trabajo = TRABAJOS[trabajo_id]
        trabajo["usuario_prizma_temporal"] = usuario_prizma
        trabajo["contrasena_prizma_temporal"] = contrasena_prizma
        trabajo["etapa"] = "en_cola"
        trabajo["mensaje"] = "Cargue agregado a la cola de procesamiento."
        trabajo["encolado_en"] = _ahora_colombia_iso()
        trabajo["cancelar_solicitado"] = False
        trabajo["terminado"] = False
        if trabajo_id not in COLA_CARGUES:
            COLA_CARGUES.append(trabajo_id)

        # Con MAX_CARGUES_SIMULTANEOS = 0 esto arranca de inmediato:
        # nadie espera a que otro usuario termine.
        _despachar_cargues_pendientes()


def _hay_capacidad():
    """Debe llamarse con COLA_CONDICION tomado."""
    if MAX_CARGUES_SIMULTANEOS <= 0:
        return True
    return CARGUES_ACTIVOS < MAX_CARGUES_SIMULTANEOS


def _despachar_cargues_pendientes():
    """Lanza un hilo propio por cada cargue que quepa.

    Debe llamarse con COLA_CONDICION tomado.
    """
    global CARGUES_ACTIVOS

    while COLA_CARGUES and _hay_capacidad():
        trabajo_id = COLA_CARGUES.popleft()
        trabajo = TRABAJOS.get(trabajo_id)

        if not trabajo:
            continue

        if trabajo.get("terminado") or trabajo.get("cancelar_solicitado"):
            continue

        CARGUES_ACTIVOS += 1

        threading.Thread(
            target=_ejecutar_cargue_en_hilo,
            args=(trabajo_id,),
            name="auto-prizma-cargue-" + str(trabajo_id)[:8],
            daemon=True,
        ).start()

    COLA_CONDICION.notify_all()


def _ejecutar_cargue_en_hilo(trabajo_id):
    """Corre UN cargue completo. Cada cargue tiene su propio hilo y su
    propio navegador, asi que varios usuarios pueden cargar a la vez."""
    global CARGUES_ACTIVOS

    usuario = ""
    contrasena = ""

    with COLA_CONDICION:
        trabajo = TRABAJOS.get(trabajo_id)

        if not trabajo:
            CARGUES_ACTIVOS -= 1
            COLA_CONDICION.notify_all()
            return

        usuario = trabajo.pop("usuario_prizma_temporal", "")
        contrasena = trabajo.pop("contrasena_prizma_temporal", "")
        trabajo["etapa"] = "iniciando"
        trabajo["mensaje"] = "Iniciando navegador y autenticación PRIZMA..."
        trabajo["iniciado_en"] = _ahora_colombia_iso()
        trabajo["terminado"] = False

    try:
        ejecutar_cargue_con_historial(
            trabajo["ruta_excel"],
            trabajo["ruta_zip"],
            trabajo["carpeta_temp"],
            trabajo["ruta_reporte"],
            trabajo["procesar_ovi"],
            trabajo["procesar_ova"],
            trabajo["procesar_retos"],
            usuario,
            contrasena,
            trabajo,
        )
    except Exception as exc:
        trabajo["etapa"] = "error"
        trabajo["mensaje"] = str(exc)
        trabajo["terminado"] = True
    finally:
        usuario = ""
        contrasena = ""
        trabajo.pop("usuario_prizma_temporal", None)
        trabajo.pop("contrasena_prizma_temporal", None)
        trabajo["finalizado_en"] = _ahora_colombia_iso()

        with COLA_CONDICION:
            CARGUES_ACTIVOS -= 1
            _despachar_cargues_pendientes()


print(
    "Cargues simultaneos:",
    "sin limite" if MAX_CARGUES_SIMULTANEOS <= 0
    else MAX_CARGUES_SIMULTANEOS,
)


# ============================================================
# CONVERTIR CSV A XLSX INTERNO
# ============================================================

def convertir_csv_a_xlsx(
    contenido_csv,
):

    texto = None

    for codificacion in [
        "utf-8-sig",
        "utf-8",
        "cp1252",
        "latin-1",
    ]:

        try:

            texto = contenido_csv.decode(
                codificacion
            )

            break

        except UnicodeDecodeError:
            continue

    if texto is None:
        raise ValueError(
            "No fue posible leer el archivo CSV."
        )

    muestra = texto[:10000]

    try:

        dialecto = csv.Sniffer().sniff(
            muestra,
            delimiters=",;\t|",
        )

        delimitador = dialecto.delimiter

    except csv.Error:

        delimitador = ","

    filas = list(
        csv.reader(
            texto.splitlines(),
            delimiter=delimitador,
        )
    )

    if not filas:
        raise ValueError(
            "El archivo CSV está vacío."
        )

    libro = Workbook()

    hoja = libro.active
    hoja.title = "Matriz"

    for numero_fila, fila in enumerate(
        filas,
        start=1,
    ):

        for numero_columna, valor in enumerate(
            fila,
            start=1,
        ):

            hoja.cell(
                row=numero_fila,
                column=numero_columna,
                value=valor,
            )

    salida = BytesIO()
    libro.save(salida)

    return salida.getvalue()


# ============================================================
# DESCARGAR MATRIZ DESDE GOOGLE SHEETS
# ============================================================

GOOGLE_SHEETS_MAX_BYTES = 50 * 1024 * 1024


def descargar_google_sheet_xlsx(url_google_sheet):
    """Exporta un Google Sheet privado a XLSX usando la cuenta OAuth conectada."""

    url_google_sheet = str(url_google_sheet or "").strip()

    coincidencia = re.search(
        r"docs\.google\.com/spreadsheets/d/([A-Za-z0-9_-]+)",
        url_google_sheet,
        flags=re.IGNORECASE,
    )

    if not coincidencia:
        raise ValueError(
            "El enlace de Google Sheets no es válido. Copia el enlace completo de la matriz."
        )

    credenciales = _cargar_credenciales_google()

    if credenciales is None:
        raise ValueError(
            "Primero debes conectar tu cuenta de Google antes de analizar una matriz privada."
        )

    spreadsheet_id = coincidencia.group(1)

    try:
        drive = build(
            "drive",
            "v3",
            credentials=credenciales,
            cache_discovery=False,
        )

        metadata = drive.files().get(
            fileId=spreadsheet_id,
            fields="id,name,mimeType",
            supportsAllDrives=True,
        ).execute()

        mime_type = str(metadata.get("mimeType") or "")

        if mime_type != "application/vnd.google-apps.spreadsheet":
            raise ValueError(
                "El enlace no corresponde a una hoja nativa de Google Sheets."
            )

        salida = BytesIO()

        solicitud = drive.files().export_media(
            fileId=spreadsheet_id,
            mimeType=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        descargador = MediaIoBaseDownload(
            salida,
            solicitud,
            chunksize=1024 * 1024,
        )

        terminado = False

        while not terminado:
            _, terminado = descargador.next_chunk()

            if salida.tell() > GOOGLE_SHEETS_MAX_BYTES:
                raise ValueError(
                    "La matriz de Google Sheets supera el tamaño máximo permitido."
                )

        contenido = salida.getvalue()

    except ValueError:
        raise

    except HttpError as error:
        estado = getattr(getattr(error, "resp", None), "status", None)

        if estado in (401, 403):
            mensaje = (
                "La cuenta de Google conectada no tiene permiso para leer esta matriz. "
                "Verifica que puedas abrirla con esa misma cuenta."
            )
        elif estado == 404:
            mensaje = (
                "Google Drive no encontró la matriz o la cuenta conectada no tiene acceso."
            )
        else:
            mensaje = (
                "Google Drive no pudo exportar la matriz. "
                f"Código HTTP: {estado or 'desconocido'}."
            )

        raise ValueError(mensaje) from error

    except Exception as error:
        raise ValueError(
            "No fue posible descargar la matriz mediante la cuenta de Google conectada."
        ) from error

    if not contenido:
        raise ValueError(
            "Google Drive devolvió una matriz vacía."
        )

    if not contenido.startswith(b"PK"):
        raise ValueError(
            "Google Drive no devolvió un archivo Excel válido."
        )

    try:
        libro_prueba = load_workbook(
            BytesIO(contenido),
            read_only=True,
            data_only=False,
        )
        libro_prueba.close()
    except Exception as error:
        raise ValueError(
            "La matriz descargada desde Google Sheets no pudo leerse como XLSX."
        ) from error

    return contenido


# ============================================================
# ANALIZAR MATRIZ
# ============================================================

def analizar_excel(
    contenido_excel,
    procesar_ovi=True,
    procesar_ova=True,
    procesar_retos=True,
    hoja_objetivo=None,
):

    libro = load_workbook(
        BytesIO(contenido_excel),
        data_only=True,
    )

    hojas_validas = []

    for nombre_hoja in libro.sheetnames:

        if not _misma_hoja(nombre_hoja, hoja_objetivo):
            continue

        hoja = libro[
            nombre_hoja
        ]

        fila_cabecera = None

        for fila in range(
            1,
            min(
                hoja.max_row,
                30,
            ) + 1,
        ):

            valor = hoja.cell(
                row=fila,
                column=1,
            ).value

            if normalizar_texto(
                valor
            ) == normalizar_texto(
                "Semana correspondiente"
            ):

                fila_cabecera = fila
                break

        if fila_cabecera is None:
            continue

        programa = hoja.cell(
            row=1,
            column=1,
        ).value

        curso = hoja.cell(
            row=2,
            column=1,
        ).value

        actividades = []

        cantidad_ovi = 0
        cantidad_ova = 0
        cantidad_retos = 0
        cantidad_h5p = 0
        cantidad_pdf = 0

        for fila in range(
            fila_cabecera + 1,
            hoja.max_row + 1,
        ):

            semana = hoja.cell(
                row=fila,
                column=1,
            ).value

            unidad = hoja.cell(
                row=fila,
                column=2,
            ).value

            nombre = hoja.cell(
                row=fila,
                column=4,
            ).value

            categoria = hoja.cell(
                row=fila,
                column=5,
            ).value

            tipo_recurso = hoja.cell(
                row=fila,
                column=6,
            ).value

            enlace = hoja.cell(
                row=fila,
                column=7,
            ).value

            if not nombre:
                continue

            if not categoria:
                continue

            categoria_prizma = normalizar_categoria(
                categoria
            )

            if categoria_prizma not in [
                "OVI",
                "OVA",
                "CHALLENGE",
            ]:
                continue

            if (
                categoria_prizma == "OVI"
                and not procesar_ovi
            ):
                continue

            if (
                categoria_prizma == "OVA"
                and not procesar_ova
            ):
                continue

            if (
                categoria_prizma == "CHALLENGE"
                and not procesar_retos
            ):
                continue

            tipo_archivo = determinar_tipo_archivo(
                categoria,
                tipo_recurso,
                enlace,
            )

            if tipo_archivo is None:
                continue

            if categoria_prizma == "OVI":
                cantidad_ovi += 1

            if categoria_prizma == "OVA":
                cantidad_ova += 1

            if categoria_prizma == "CHALLENGE":
                cantidad_retos += 1

            if tipo_archivo == "H5P":
                cantidad_h5p += 1

            if tipo_archivo == "PDF":
                cantidad_pdf += 1

            actividades.append(
                {
                    "fila": fila,

                    "semana": (
                        str(semana).strip()
                        if semana
                        else ""
                    ),

                    "unidad": (
                        str(unidad).strip()
                        if unidad
                        else ""
                    ),

                    "nombre": str(
                        nombre
                    ).strip(),

                    "categoria":
                        categoria_prizma,

                    "tipo":
                        tipo_archivo,
                }
            )

        hojas_validas.append(
            {
                "hoja":
                    nombre_hoja,

                "programa": (
                    str(programa).strip()
                    if programa
                    else ""
                ),

                "curso": (
                    str(curso).strip()
                    if curso
                    else nombre_hoja
                ),

                "actividades":
                    actividades,

                "ovi":
                    cantidad_ovi,

                "ova":
                    cantidad_ova,

                "retos":
                    cantidad_retos,

                "h5p":
                    cantidad_h5p,

                "pdf":
                    cantidad_pdf,
            }
        )

    return hojas_validas


# ============================================================
# ANALIZAR ZIP
# ============================================================

def analizar_zip(
    contenido_zip,
):

    cantidad_h5p = 0
    cantidad_pdf = 0

    with zipfile.ZipFile(
        BytesIO(contenido_zip),
        "r",
    ) as archivo_zip:

        for miembro in archivo_zip.namelist():

            if miembro.endswith("/"):
                continue

            nombre = os.path.basename(
                miembro
            )

            extension = os.path.splitext(
                nombre
            )[1].lower()

            if extension == ".h5p":

                cantidad_h5p += 1

            elif extension == ".pdf":

                cantidad_pdf += 1

    return {
        "h5p":
            cantidad_h5p,

        "pdf":
            cantidad_pdf,

        "total":
            cantidad_h5p
            + cantidad_pdf,
    }


# ============================================================
# HTML PRINCIPAL
# ============================================================

def generar_html(
    resultado=None,
    error=None,
    trabajo_id=None,
):

    def e(valor):
        return html.escape(
            str(valor or "")
        )

    contenido = ""

    # ========================================================
    # PANTALLA INICIAL
    # ========================================================

    if resultado is None:

        bloque_error = ""

        if error:
            bloque_error = f"""
            <div class="alerta-error">
                <div class="alerta-icono">!</div>
                <div>
                    <strong>No se pudo continuar</strong>
                    <p>{e(error)}</p>
                </div>
            </div>
            """

        google_conectado = _google_conectado()

        if google_conectado:
            bloque_google = """
            <section class="google-conexion google-ok">
                <div class="google-icono">G</div>
                <div class="google-texto">
                    <strong>Google conectado</strong>
                    <span>Esta sesión usa tu propia cuenta de Google, con permiso de solo lectura para Sheets y Drive.</span>
                </div>
                <form action="/google/desconectar" method="post">
                    <button class="google-boton secundario" type="submit">Desconectar</button>
                </form>
            </section>
            """
        else:
            bloque_google = """
            <section class="google-conexion">
                <div class="google-icono">G</div>
                <div class="google-texto">
                    <strong>Conecta tu cuenta de Google</strong>
                    <span>Cada persona conecta la suya. Necesario para leer matrices y carpetas privadas.</span>
                </div>
                <a class="google-boton" href="/google/conectar">Conectar con Google</a>
            </section>
            """

        contenido = f"""
        <section class="encabezado-pagina">
            <div>
                <h1>¡Hola, {e(_nombre_visible(_usuario_actual()))}!</h1>
                <p>Automatiza el cargue de actividades en PRIZMA de forma rápida y segura.</p>
            </div>
        </section>

        {bloque_error}

        {bloque_google}

        <form
            id="cargue-principal"
            action="/analizar"
            method="post"
            enctype="multipart/form-data"
            class="panel panel-principal"
        >
            <div class="titulo-seccion">
                <span class="numero-seccion">1</span>
                <div>
                    <h2>Archivos</h2>
                    <p>Pega el enlace de Google Sheets y selecciona el paquete de recursos.</p>
                </div>
            </div>

            <div class="grid-archivos">
                <div class="tarjeta-archivo">
                    <div class="archivo-cabecera">
                        <div class="archivo-icono verde" aria-hidden="true">
                            <svg viewBox="0 0 24 24"><path d="M6 2h8l4 4v16H6z"/><path d="M14 2v5h5"/><path d="M9 11l6 6M15 11l-6 6"/></svg>
                        </div>
                        <div>
                            <strong>Google Sheets de la matriz</strong>
                            <span>Pega el enlace de la pestaña que vas a cargar. Solo se procesa esa pestaña.</span>
                        </div>
                    </div>
                    <div class="selector-archivo">
                        <input
                            id="google-sheet-url"
                            type="url"
                            name="google_sheet_url"
                            placeholder="https://docs.google.com/spreadsheets/d/..."
                            autocomplete="off"
                            required
                            style="display:block;width:100%;border:0;outline:0;background:transparent;color:#101828;font:inherit;font-size:12px;"
                        >
                    </div>
                </div>

                <label class="tarjeta-archivo zona-drop" for="archivo-zip" data-input="archivo-zip">
                    <div class="archivo-cabecera">
                        <div class="archivo-icono morado" aria-hidden="true">
                            <svg viewBox="0 0 24 24"><path d="M6 2h8l4 4v16H6z"/><path d="M14 2v5h5"/><path d="M11 5h2M11 8h2M11 11h2M11 14h2"/></svg>
                        </div>
                        <div>
                            <strong>ZIP de recursos</strong>
                            <span>Archivos H5P y PDF · también puedes arrastrarlo aquí</span>
                        </div>
                    </div>
                    <div class="selector-archivo">
                        <span class="boton-selector">Seleccionar archivo</span>
                        <span id="nombre-zip" class="nombre-archivo">Ningún archivo seleccionado</span>
                    </div>
                    <input
                        id="archivo-zip"
                        type="file"
                        name="recursos"
                        accept=".zip"
                        required
                    >
                </label>
            </div>

            <div class="separador"></div>

            <div class="titulo-seccion">
                <span class="numero-seccion">2</span>
                <div>
                    <h2>Tipos de actividades</h2>
                    <p>Selecciona qué actividades deseas procesar.</p>
                </div>
            </div>

            <div class="grid-tipos">
                <label class="tarjeta-tipo tipo-verde">
                    <input type="checkbox" name="ovi" value="1" checked>
                    <span class="check-personalizado">✓</span>
                    <span class="tipo-icono icono-ovi" aria-hidden="true">
                        <svg viewBox="0 0 24 24"><path d="M4 5.5A2.5 2.5 0 0 1 6.5 3H11v16H6.5A2.5 2.5 0 0 0 4 21.5z"/><path d="M20 5.5A2.5 2.5 0 0 0 17.5 3H13v16h4.5a2.5 2.5 0 0 1 2.5 2.5z"/></svg>
                    </span>
                    <strong>OVI</strong>
                    <small>Objetos Virtuales de Información</small>
                </label>

                <label class="tarjeta-tipo tipo-azul">
                    <input type="checkbox" name="ova" value="1" checked>
                    <span class="check-personalizado">✓</span>
                    <span class="tipo-icono icono-ova" aria-hidden="true">
                        <svg viewBox="0 0 24 24"><rect x="3" y="4" width="18" height="13" rx="2"/><path d="M8 21h8M12 17v4"/><path d="m9 9 2 2 4-4"/></svg>
                    </span>
                    <strong>OVA</strong>
                    <small>Objetos Virtuales de Aprendizaje</small>
                </label>

                <label class="tarjeta-tipo tipo-morado">
                    <input type="checkbox" name="retos" value="1" checked>
                    <span class="check-personalizado">✓</span>
                    <span class="tipo-icono icono-reto" aria-hidden="true">
                        <svg viewBox="0 0 24 24"><path d="M8 3h8v4a4 4 0 0 1-8 0z"/><path d="M8 5H4v2a4 4 0 0 0 4 4M16 5h4v2a4 4 0 0 1-4 4"/><path d="M12 11v5M8 21h8M10 16h4v5h-4z"/></svg>
                    </span>
                    <strong>Retos Evaluativos</strong>
                    <small>Retos evaluativos en PDF</small>
                </label>
            </div>

            <button class="boton-principal" type="submit">
                <span>↗</span>
                Analizar matriz y recursos
            </button>

            <p class="nota-seguridad">
                Tus archivos se utilizan únicamente durante el proceso de cargue.
            </p>
        </form>

        <aside id="ayuda" class="panel panel-ayuda">
            <h3>¿Cómo funciona?</h3>
            <div class="paso"><b>1</b><span>Pega el enlace de Google Sheets de la matriz.</span></div>
            <div class="paso"><b>2</b><span>Sube el ZIP con los recursos correspondientes.</span></div>
            <div class="paso"><b>3</b><span>Selecciona los tipos de actividades.</span></div>
            <div class="paso"><b>4</b><span>Analiza, revisa y luego inicia el cargue.</span></div>

            <div class="ayuda-separador"></div>

            <h3 class="titulo-consejos">Consejos</h3>
            <p class="consejo">✓ La matriz debe conservar los enlaces de Google Drive de cada recurso.</p>
            <p class="consejo">✓ El ZIP debe contener los H5P y PDF correspondientes.</p>
            <p class="consejo">✓ Revisa el análisis antes de iniciar el cargue.</p>
        </aside>
        """

    # ========================================================
    # PANTALLA DE REVISIÓN
    # ========================================================

    else:

        total_actividades = 0
        total_ovi = 0
        total_ova = 0
        total_retos = 0
        total_h5p = 0
        total_pdf = 0
        filas_html = ""

        drive_info = resultado.get("drive") or {}
        drive_validados = int(drive_info.get("validadas_drive") or 0)
        recursos_drive = {
            (str(item.get("hoja") or ""), int(item.get("fila") or 0)):
                str(item.get("archivo") or "")
            for item in drive_info.get("recursos", [])
        }

        for hoja in resultado["hojas"]:

            total_actividades += len(
                hoja["actividades"]
            )
            total_ovi += hoja.get("ovi", 0)
            total_ova += hoja.get("ova", 0)
            total_retos += hoja.get("retos", 0)
            total_h5p += hoja.get("h5p", 0)
            total_pdf += hoja.get("pdf", 0)

            for actividad in hoja["actividades"]:

                categoria = actividad["categoria"]
                clase_categoria = {
                    "OVI": "badge-verde",
                    "OVA": "badge-azul",
                    "CHALLENGE": "badge-morado",
                }.get(categoria, "badge-gris")

                etiqueta_categoria = (
                    "RETO"
                    if categoria == "CHALLENGE"
                    else categoria
                )

                clase_tipo = (
                    "badge-pdf"
                    if actividad["tipo"] == "PDF"
                    else "badge-h5p"
                )

                archivo_drive = recursos_drive.get(
                    (str(hoja.get("hoja") or ""), int(actividad.get("fila") or 0)),
                    "",
                )

                filas_html += f"""
                <tr>
                    <td>{e(actividad["semana"])}</td>
                    <td>{e(actividad["unidad"])}</td>
                    <td class="celda-actividad">{e(actividad["nombre"])}</td>
                    <td><span class="badge {clase_categoria}">{e(etiqueta_categoria)}</span></td>
                    <td><span class="badge {clase_tipo}">{e(actividad["tipo"])}</span></td>
                    <td class="celda-actividad" title="{e(archivo_drive)}">{e(archivo_drive)}</td>
                </tr>
                """

        bloque_error_revision = ""
        if error:
            bloque_error_revision = f"""
            <div class="alerta-error" style="grid-column:1/-1;margin-bottom:18px;">
                <div class="alerta-icono">!</div>
                <div><strong>No se pudo iniciar el cargue</strong><p>{e(error)}</p></div>
            </div>
            """

        contenido = f"""
        <section id="cargue-principal" class="encabezado-exito panel">
            <div class="check-grande">✓</div>
            <div>
                <h1>Análisis completado</h1>
                <p>{total_actividades} actividades listas para cargar · {drive_validados} recursos verificados con Google Drive</p>
            </div>
        </section>

        <section class="resumen-grid">
            <div class="resumen-card"><span>Actividades</span><strong>{total_actividades}</strong></div>
            <div class="resumen-card verde"><span>OVI</span><strong>{total_ovi}</strong></div>
            <div class="resumen-card azul"><span>OVA</span><strong>{total_ova}</strong></div>
            <div class="resumen-card morado"><span>Retos</span><strong>{total_retos}</strong></div>
            <div class="resumen-card violeta"><span>H5P</span><strong>{total_h5p}</strong></div>
            <div class="resumen-card rojo"><span>PDF</span><strong>{total_pdf}</strong></div>
        </section>

        {bloque_error_revision}

        <section class="revision-grid">
            <div class="panel tabla-panel">
                <div class="titulo-bloque">
                    <h2>Actividades detectadas</h2>
                    <p>Revisa que las actividades y categorías sean correctas.</p>
                </div>

                <div class="tabla-scroll">
                    <table>
                        <thead>
                            <tr>
                                <th>Semana</th>
                                <th>Unidad</th>
                                <th>Actividad</th>
                                <th>Categoría</th>
                                <th>Tipo</th>
                                <th>Archivo verificado</th>
                            </tr>
                        </thead>
                        <tbody>
                            {filas_html}
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="panel credenciales-panel">
                <div class="titulo-bloque">
                    <h2>Acceso a PRIZMA</h2>
                    <p>Ingresa tus credenciales para autorizar este cargue.</p>
                </div>

                <form action="/iniciar/{trabajo_id}" method="post" autocomplete="off">
                    <label class="campo-label" for="usuario_prizma">Usuario PRIZMA</label>
                    <div class="campo-moderno">
                        <span>♙</span>
                        <input
                            id="usuario_prizma"
                            type="text"
                            name="usuario_prizma"
                            placeholder="Ingresa tu usuario"
                            autocomplete="off"
                            required
                        >
                    </div>

                    <label class="campo-label" for="contrasena_prizma">Contraseña PRIZMA</label>
                    <div class="campo-moderno">
                        <span>⌾</span>
                        <input
                            id="contrasena_prizma"
                            type="password"
                            name="contrasena_prizma"
                            placeholder="Ingresa tu contraseña"
                            autocomplete="new-password"
                            required
                        >
                    </div>

                    <div class="caja-seguridad">
                        <strong>Credenciales de uso temporal</strong>
                        <span>La aplicación no las escribe en archivos, reportes ni base de datos.</span>
                    </div>

                    <button class="boton-principal" type="submit">
                        <span>↗</span>
                        Iniciar cargue en PRIZMA
                    </button>
                </form>
            </div>
        </section>

        <a class="boton-secundario" href="/">← Cambiar matriz o ZIP</a>
        """

    return """
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Auto Prizma Pro</title>
        <style>
            :root {
                --fondo: #f7f8fc;
                --panel: #ffffff;
                --texto: #101828;
                --muted: #667085;
                --borde: #e5e7ef;
                --morado: #5b48e8;
                --morado-2: #4338ca;
                --verde: #0ea968;
                --azul: #1976d2;
                --rojo: #ef4444;
                --sombra: 0 10px 30px rgba(29, 41, 57, .06);
            }

            * { box-sizing: border-box; }

            body {
                margin: 0;
                font-family: Inter, ui-sans-serif, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                background: var(--fondo);
                color: var(--texto);
            }

            .app { min-height: 100vh; display: grid; grid-template-columns: 235px 1fr; }

            .sidebar {
                position: sticky;
                top: 0;
                height: 100vh;
                background: #fff;
                border-right: 1px solid var(--borde);
                padding: 28px 20px;
                display: flex;
                flex-direction: column;
            }

            .marca { display: flex; align-items: center; gap: 12px; margin-bottom: 34px; }
            .logo {
                width: 44px; height: 44px; border-radius: 13px;
                display: grid; place-items: center;
                background: linear-gradient(145deg, #6d5dfc, #4338ca);
                box-shadow: 0 8px 20px rgba(79, 70, 229, .25);
                flex: 0 0 44px;
            }
            .logo svg { width: 29px; height: 29px; overflow: visible; }
            .logo svg path:first-child { fill: #fff; }
            .logo svg path:last-child { fill: #c7d2fe; }
            .marca strong { display: block; font-size: 18px; }
            .marca span { display: block; color: var(--muted); font-size: 12px; margin-top: 3px; }

            .nav { display: grid; gap: 8px; }
            .nav-item {
                padding: 12px 14px; border-radius: 11px; color: #475467; font-size: 14px;
                display: flex; gap: 11px; align-items: center; text-decoration: none; transition: .18s ease;
            }
            a.nav-item:hover { background: #f7f5ff; color: #4f46e5; }
            .nav-item.activo { background: #f1efff; color: #4f46e5; font-weight: 700; }
            .nav-item.proximamente { opacity: .48; cursor: default; }
            .nav-item.proximamente small { margin-left: auto; font-size: 9px; }

            .estado-servicio {
                margin-top: auto; border: 1px solid var(--borde); border-radius: 14px;
                padding: 15px; background: #fff;
            }
            .servicio-linea { display: flex; align-items: center; gap: 8px; color: #07894f; font-size: 13px; font-weight: 700; }
            .punto { width: 8px; height: 8px; border-radius: 50%; background: #15b76a; }
            .servicio-mini { margin-top: 13px; display: flex; justify-content: space-between; font-size: 12px; color: var(--muted); }
            .chip { background: #eef4ff; color: #3538cd; padding: 4px 8px; border-radius: 999px; }

            .contenido { padding: 28px 34px 45px; max-width: 1500px; width: 100%; margin: 0 auto; }
            .layout-inicio { display: grid; grid-template-columns: minmax(0, 1fr) 290px; gap: 24px; align-items: start; }

            .panel {
                background: var(--panel); border: 1px solid var(--borde); border-radius: 16px;
                box-shadow: var(--sombra);
            }

            .encabezado-pagina {
                grid-column: 1 / -1; background: #fff; border: 1px solid var(--borde);
                border-radius: 16px; padding: 24px 28px; box-shadow: var(--sombra);
            }
            .encabezado-pagina h1, .encabezado-exito h1 { margin: 0; font-size: 25px; }
            .encabezado-pagina p, .encabezado-exito p { margin: 6px 0 0; color: var(--muted); }

            .google-conexion {
                grid-column: 1 / -1;
                display: flex;
                align-items: center;
                gap: 14px;
                padding: 16px 18px;
                background: #fff;
                border: 1px solid var(--borde);
                border-radius: 14px;
                box-shadow: var(--sombra);
            }
            .google-conexion.google-ok {
                border-color: #a7e7c7;
                background: #fbfffd;
            }
            .google-icono {
                width: 42px;
                height: 42px;
                border-radius: 11px;
                display: grid;
                place-items: center;
                background: #f5f6f8;
                color: #344054;
                font-size: 18px;
                font-weight: 900;
                flex: 0 0 42px;
            }
            .google-texto { min-width: 0; flex: 1 1 auto; }
            .google-texto strong { display: block; font-size: 14px; }
            .google-texto span {
                display: block;
                margin-top: 4px;
                color: var(--muted);
                font-size: 12px;
                line-height: 1.45;
            }
            .google-boton {
                border: 0;
                border-radius: 9px;
                padding: 10px 14px;
                background: #4f46e5;
                color: #fff;
                text-decoration: none;
                font-size: 12px;
                font-weight: 800;
                cursor: pointer;
                white-space: nowrap;
            }
            .google-boton.secundario {
                background: #fff;
                color: #475467;
                border: 1px solid #d0d5dd;
            }

            .panel-principal { padding: 26px; }
            .titulo-seccion { display: flex; gap: 12px; align-items: flex-start; margin-bottom: 18px; }
            .titulo-seccion h2, .titulo-bloque h2 { margin: 0; font-size: 18px; }
            .titulo-seccion p, .titulo-bloque p { margin: 5px 0 0; color: var(--muted); font-size: 13px; }
            .numero-seccion {
                width: 28px; height: 28px; border-radius: 9px; display: grid; place-items: center;
                background: #f0edff; color: #5145cd; font-weight: 800; font-size: 13px;
            }

            .grid-archivos { display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }
            .tarjeta-archivo {
                border: 1px dashed #cfd4e1; border-radius: 14px; padding: 20px;
                cursor: pointer; transition: .18s ease; background: #fff;
                min-width: 0; overflow: hidden;
            }
            .tarjeta-archivo:hover { border-color: #8176f2; background: #fbfaff; transform: translateY(-1px); }
            .tarjeta-archivo.arrastrando { border-color: #5b48e8; background: #f5f3ff; box-shadow: inset 0 0 0 2px rgba(91,72,232,.10); }
            .tarjeta-archivo input { display: none; }
            .archivo-cabecera { display: flex; align-items: center; gap: 12px; min-width: 0; }
            .archivo-cabecera > div:last-child { min-width: 0; }
            .archivo-cabecera strong { display: block; font-size: 14px; }
            .archivo-cabecera span { display: block; font-size: 12px; color: var(--muted); margin-top: 4px; }
            .archivo-icono {
                width: 40px; height: 40px; border-radius: 10px; display: grid; place-items: center;
                flex: 0 0 40px;
            }
            .archivo-icono svg { width: 23px; height: 23px; fill: none; stroke: currentColor; stroke-width: 1.8; stroke-linecap: round; stroke-linejoin: round; }
            .archivo-icono.verde { background: #e8f8f0; color: #0a9b5b; }
            .archivo-icono.morado { background: #f0edff; color: #5b48e8; }
            .selector-archivo {
                margin-top: 18px; min-height: 46px; border: 1px solid var(--borde); border-radius: 10px;
                padding: 9px 12px; display: flex; align-items: center; gap: 12px;
                min-width: 0; overflow: hidden;
            }
            .boton-selector {
                background: #5b48e8; color: #fff; padding: 8px 12px; border-radius: 8px;
                font-size: 12px; font-weight: 700; white-space: nowrap;
            }
            .nombre-archivo { color: var(--muted); font-size: 12px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; min-width: 0; flex: 1 1 auto; display: block; }

            .separador { height: 1px; background: var(--borde); margin: 26px 0; }
            .grid-tipos { display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; }
            .tarjeta-tipo {
                position: relative; border: 1px solid var(--borde); border-radius: 14px; padding: 18px;
                min-height: 125px; cursor: pointer; transition: .18s ease; display: flex; flex-direction: column;
            }
            .tarjeta-tipo:hover { transform: translateY(-1px); box-shadow: 0 8px 22px rgba(16,24,40,.06); }
            .tarjeta-tipo input { position: absolute; opacity: 0; pointer-events: none; }
            .check-personalizado {
                position: absolute; top: 15px; right: 15px; width: 22px; height: 22px;
                border-radius: 6px; display: grid; place-items: center; color: white; font-size: 13px;
                background: #d0d5dd;
            }
            .tarjeta-tipo:has(input:checked).tipo-verde { border-color: #65d6a1; background: #fbfffd; }
            .tarjeta-tipo:has(input:checked).tipo-verde .check-personalizado { background: #0ea968; }
            .tarjeta-tipo:has(input:checked).tipo-azul { border-color: #7ab6ef; background: #fbfdff; }
            .tarjeta-tipo:has(input:checked).tipo-azul .check-personalizado { background: #1976d2; }
            .tarjeta-tipo:has(input:checked).tipo-morado { border-color: #b9a8f5; background: #fdfcff; }
            .tarjeta-tipo:has(input:checked).tipo-morado .check-personalizado { background: #8b5cf6; }
            .tipo-icono { width: 42px; height: 42px; border-radius: 12px; display: grid; place-items: center; margin-bottom: 15px; }
            .tipo-icono svg { width: 25px; height: 25px; fill: none; stroke: currentColor; stroke-width: 1.8; stroke-linecap: round; stroke-linejoin: round; }
            .icono-ovi { color:#079455; background:#e9f9f1; }
            .icono-ova { color:#1976d2; background:#eaf4ff; }
            .icono-reto { color:#7f56d9; background:#f1ebff; }
            .tarjeta-tipo strong { font-size: 16px; }
            .tarjeta-tipo small { color: var(--muted); margin-top: 6px; line-height: 1.4; }

            .boton-principal {
                width: 100%; border: 0; border-radius: 11px; padding: 15px 20px; margin-top: 24px;
                color: #fff; font-size: 15px; font-weight: 800; cursor: pointer;
                background: linear-gradient(90deg, #5145e5, #4f46e5 55%, #5f45df);
                box-shadow: 0 8px 20px rgba(79,70,229,.20);
            }
            .boton-principal:hover { filter: brightness(.98); transform: translateY(-1px); }
            .boton-principal span { margin-right: 8px; }
            .nota-seguridad { text-align: center; color: var(--muted); font-size: 12px; margin: 14px 0 0; }

            .panel-ayuda { padding: 24px; }
            .panel-ayuda h3 { margin: 0 0 20px; font-size: 17px; }
            .paso { display: grid; grid-template-columns: 30px 1fr; gap: 12px; align-items: start; margin-bottom: 20px; color: #475467; font-size: 13px; line-height: 1.5; }
            .paso b { width: 28px; height: 28px; border-radius: 50%; background: #f1efff; color: #4f46e5; display: grid; place-items: center; }
            .ayuda-separador { height: 1px; background: var(--borde); margin: 24px 0; }
            .titulo-consejos { color: #4f46e5; }
            .consejo { color: #475467; font-size: 13px; line-height: 1.55; }

            .alerta-error {
                grid-column: 1 / -1; display: flex; gap: 12px; padding: 16px 18px; border-radius: 12px;
                background: #fff1f2; border: 1px solid #fecdd3; color: #9f1239;
            }
            .alerta-error p { margin: 4px 0 0; }
            .alerta-icono { width: 28px; height: 28px; border-radius: 50%; background: #ef4444; color: #fff; display: grid; place-items: center; font-weight: 900; }

            .encabezado-exito { padding: 22px 24px; display: flex; gap: 15px; align-items: center; margin-bottom: 0; grid-column: 1 / -1; }
            .check-grande { width: 48px; height: 48px; border-radius: 14px; background: #e9f9f1; color: #08a45e; display: grid; place-items: center; font-size: 25px; font-weight: 900; }
            .resumen-grid { display: grid; grid-template-columns: repeat(6, minmax(110px, 1fr)); gap: 14px; margin: 18px 0 20px; grid-column: 1 / -1; min-width: 0; }
            .resumen-card { background: #fff; border: 1px solid var(--borde); border-radius: 14px; padding: 17px; box-shadow: var(--sombra); text-align: center; }
            .resumen-card span { display: block; color: var(--muted); font-size: 12px; }
            .resumen-card strong { display: block; font-size: 27px; margin-top: 6px; }
            .resumen-card.verde strong { color: #0a9b5b; }
            .resumen-card.azul strong { color: #1976d2; }
            .resumen-card.morado strong, .resumen-card.violeta strong { color: #6d4ce8; }
            .resumen-card.rojo strong { color: #e5484d; }

            .revision-grid { display: grid; grid-template-columns: minmax(0, 1.55fr) minmax(310px, .75fr); gap: 20px; align-items: start; grid-column: 1 / -1; min-width: 0; }
            .tabla-panel, .credenciales-panel { padding: 24px; }
            .titulo-bloque { margin-bottom: 18px; }
            .tabla-scroll { max-height: 570px; overflow: auto; border: 1px solid var(--borde); border-radius: 12px; }
            table { width: 100%; border-collapse: collapse; font-size: 13px; }
            thead { position: sticky; top: 0; background: #f8f9fc; z-index: 1; }
            th { text-align: left; color: #475467; font-size: 12px; padding: 12px 13px; border-bottom: 1px solid var(--borde); }
            td { padding: 12px 13px; border-bottom: 1px solid #eef0f4; vertical-align: middle; }
            tbody tr:last-child td { border-bottom: 0; }
            .celda-actividad { min-width: 260px; }
            .badge { display: inline-block; padding: 4px 8px; border-radius: 999px; font-size: 10px; font-weight: 800; }
            .badge-verde { background: #e9f9f1; color: #07894f; }
            .badge-azul { background: #eaf4ff; color: #1565c0; }
            .badge-morado { background: #f1ebff; color: #7048d7; }
            .badge-gris { background: #f2f4f7; color: #475467; }
            .badge-h5p { background: #f1ebff; color: #6941c6; }
            .badge-pdf { background: #fff0f1; color: #d92d20; }

            .campo-label { display: block; font-size: 13px; font-weight: 700; margin: 18px 0 8px; }
            .campo-moderno { display: flex; align-items: center; gap: 10px; border: 1px solid #d7dbe5; border-radius: 10px; padding: 0 12px; background: #fff; }
            .campo-moderno span { color: #667085; }
            .campo-moderno input { width: 100%; border: 0; outline: 0; padding: 13px 0; font-size: 14px; background: transparent; }
            .caja-seguridad { margin-top: 20px; background: #f8f7ff; border: 1px solid #d9d4ff; border-radius: 11px; padding: 14px; }
            .caja-seguridad strong, .caja-seguridad span { display: block; }
            .caja-seguridad strong { font-size: 12px; color: #4f46e5; }
            .caja-seguridad span { font-size: 12px; color: var(--muted); margin-top: 5px; line-height: 1.45; }
            .boton-secundario { grid-column: 1 / -1; display: inline-block; margin-top: 20px; color: #4f46e5; text-decoration: none; font-weight: 700; font-size: 14px; padding: 11px 15px; border: 1px solid #d8d6f8; border-radius: 10px; background: #fff; }

            @media (max-width: 1050px) {
                .app { grid-template-columns: 1fr; }
                .sidebar { display: none; }
                .contenido { padding: 20px; }
                .layout-inicio { grid-template-columns: 1fr; }
                .panel-ayuda { order: 2; }
                .resumen-grid { grid-template-columns: repeat(3, 1fr); }
                .revision-grid { grid-template-columns: 1fr; }
            }

            @media (max-width: 700px) {
                .google-conexion { align-items: flex-start; flex-wrap: wrap; }
                .google-conexion .google-boton { width: 100%; text-align: center; }
                .grid-archivos, .grid-tipos { grid-template-columns: 1fr; }
                .resumen-grid { grid-template-columns: repeat(2, 1fr); }
                .contenido { padding: 14px; }
                .panel-principal, .tabla-panel, .credenciales-panel { padding: 18px; }
            }
        </style><link rel="stylesheet" href="/estilos-responsive.css">
    </head>
    <body>
        <div class="app">
            <aside class="sidebar">
                <div class="marca">
                    <div class="logo" aria-label="Auto Prizma Pro">
                        <svg viewBox="0 0 48 48" aria-hidden="true">
                            <path d="M9 35.5 20.5 8.5c.8-1.9 3.4-1.9 4.2 0l4.1 9.6-5.4 12.7-3.1-7.4-5.2 12.1z"/>
                            <path d="M26.4 14.5 39 35.5h-8.2l-8.5-14.2z"/>
                        </svg>
                    </div>
                    <div>
                        <strong>Auto Prizma Pro</strong>
                        <span>Automatización PRIZMA</span>
                    </div>
                </div>

                <nav class="nav">
                    <a class="nav-item activo" href="/">⌂ <span>Inicio</span></a>
                    <a class="nav-item" href="/cargue-actual">⇧ <span>Cargue actual</span></a>
                    <a class="nav-item" href="/historial">◷ <span>Historial</span></a>
                    <a class="nav-item" href="/reportes">▥ <span>Reportes</span></a><a class="nav-item" href="/cambiar-clave">✎ <span>Mi contraseña</span></a><a class="nav-item" href="/salir">⏻ <span>Salir</span></a>
                </nav>

                <div class="estado-servicio">
                    <div class="servicio-linea"><span class="punto"></span> Servicio activo</div>
                    <div class="servicio-mini"><span>Navegador</span><span class="chip">Chromium</span></div>
                    <div class="servicio-mini"><span>Conexión</span><span class="chip">Estable</span></div>
                </div>
            </aside>

            <main class="contenido">
                <div class="layout-inicio">
                    """ + contenido + """
                </div>
            </main>
        </div>

        <script>
            const matriz = document.getElementById('archivo-matriz');
            const zip = document.getElementById('archivo-zip');

            function actualizarNombre(input, destinoId) {
                const destino = document.getElementById(destinoId);
                if (!destino || !input) return;
                destino.textContent = input.files.length ? input.files[0].name : 'Ningún archivo seleccionado';
                destino.title = input.files.length ? input.files[0].name : '';
            }

            if (matriz) matriz.addEventListener('change', () => actualizarNombre(matriz, 'nombre-matriz'));
            if (zip) zip.addEventListener('change', () => actualizarNombre(zip, 'nombre-zip'));

            document.querySelectorAll('.zona-drop').forEach((zona) => {
                const input = document.getElementById(zona.dataset.input);
                if (!input) return;

                ['dragenter', 'dragover'].forEach((evento) => {
                    zona.addEventListener(evento, (e) => {
                        e.preventDefault();
                        e.stopPropagation();
                        zona.classList.add('arrastrando');
                    });
                });

                ['dragleave', 'drop'].forEach((evento) => {
                    zona.addEventListener(evento, (e) => {
                        e.preventDefault();
                        e.stopPropagation();
                        zona.classList.remove('arrastrando');
                    });
                });

                zona.addEventListener('drop', (e) => {
                    const archivos = e.dataTransfer.files;
                    if (!archivos || !archivos.length) return;
                    const transferencia = new DataTransfer();
                    transferencia.items.add(archivos[0]);
                    input.files = transferencia.files;
                    input.dispatchEvent(new Event('change', { bubbles: true }));
                });
            });
        </script>
    </body>
    </html>
    """


# ============================================================
# GOOGLE OAUTH - CONECTAR / CALLBACK / DESCONECTAR
# ============================================================

# ============================================================
# LOGIN
# ============================================================

RUTAS_PUBLICAS = {
    "/login",
    "/salir",
    "/favicon.ico",
    "/estilos-responsive.css",
}


# Se carga DESPUES del <style> de cada pagina, por eso puede corregir
# lo que traigan los estilos propios sin tener que tocarlos uno por uno.
CSS_RESPONSIVE = r"""
/* ---------- base ---------- */
html { -webkit-text-size-adjust: 100%; }
img, svg, video { max-width: 100%; height: auto; }
* { min-width: 0; }

/* ---------- pantallas medianas y telefonos ---------- */
@media (max-width: 1050px) {

  .app { grid-template-columns: 1fr !important; }

  /* El menu lateral pasa a ser una barra superior.
     Antes simplemente se ocultaba y en el telefono no habia
     forma de ir a Historial, Reportes ni Salir. */
  .sidebar {
    display: flex !important;
    position: sticky !important;
    top: 0;
    z-index: 40;
    height: auto !important;
    min-height: 0 !important;
    flex-direction: column !important;
    gap: 10px;
    padding: 12px 14px !important;
    border-right: 0 !important;
    border-bottom: 1px solid #e5e7ef;
    background: #fff;
  }

  .sidebar .marca { margin-bottom: 0 !important; }
  .sidebar .marca span { display: none !important; }
  .sidebar .marca strong { font-size: 16px !important; }
  .sidebar .logo {
    width: 36px !important;
    height: 36px !important;
    border-radius: 10px !important;
    flex: 0 0 36px;
  }
  .sidebar .logo svg { width: 23px !important; height: 23px !important; }

  .sidebar .nav {
    display: flex !important;
    flex-direction: row !important;
    gap: 6px !important;
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    scrollbar-width: none;
    padding-bottom: 2px;
  }
  .sidebar .nav::-webkit-scrollbar { display: none; }

  .sidebar .nav-item {
    padding: 10px 13px !important;
    font-size: 13px !important;
    white-space: nowrap;
    flex: 0 0 auto;
  }

  .sidebar .estado-servicio { display: none !important; }

  .contenido { padding: 18px 16px 32px !important; }

  /* Cualquier rejilla de dos o mas columnas pasa a una sola. */
  .layout-inicio,
  .revision-grid,
  .carga-card,
  .zona,
  .herramientas,
  .grid-archivos,
  .grid-tipos { grid-template-columns: 1fr !important; }

  .accion-bloque { border-left: 0 !important; padding-left: 0 !important; }
}

/* ---------- telefonos ---------- */
@media (max-width: 700px) {

  .contenido { padding: 14px 12px 28px !important; }

  .panel-principal,
  .tabla-panel,
  .credenciales-panel,
  .panel,
  .vacio,
  .carga-card { padding: 16px !important; border-radius: 14px !important; }

  .vacio { padding: 30px 18px !important; }

  h1 { font-size: 21px !important; line-height: 1.25; }
  h2 { font-size: 17px !important; }

  .cabecera,
  .cabecera-top,
  .cabecera-linea {
    flex-direction: column !important;
    align-items: flex-start !important;
    gap: 12px;
  }

  .resumen-grid,
  .numeros,
  .metricas { grid-template-columns: repeat(2, 1fr) !important; }

  /* Tarjeta de conexion con Google: el boton baja y ocupa el ancho. */
  .google-conexion {
    flex-direction: column !important;
    align-items: flex-start !important;
    gap: 12px;
  }
  .google-conexion .google-boton,
  .google-boton {
    width: 100% !important;
    text-align: center !important;
    justify-content: center !important;
  }

  /* Botones y enlaces de accion a lo ancho, mas facil de tocar. */
  .boton,
  .boton-secundario,
  button[type="submit"],
  .accion-bloque a {
    width: 100% !important;
    justify-content: center !important;
    text-align: center !important;
    padding: 14px 16px !important;
  }

  /* 16px evita que iOS haga zoom al enfocar un campo. */
  input, select, textarea { font-size: 16px !important; }
  .campo-moderno input { padding: 12px 0 !important; }

  /* Las tablas anchas se deslizan en vez de romper la pantalla. */
  table {
    display: block !important;
    width: 100% !important;
    overflow-x: auto !important;
    -webkit-overflow-scrolling: touch;
    white-space: nowrap;
  }

  /* Nombres largos de curso o de archivo no desbordan. */
  .curso-texto h2,
  .curso-texto p { white-space: normal !important; }

  td, th { word-break: break-word; }
}

/* ---------- pantallas muy angostas ---------- */
@media (max-width: 420px) {
  .resumen-grid,
  .numeros,
  .metricas { grid-template-columns: 1fr !important; }

  .contenido { padding: 12px 10px 24px !important; }
  .sidebar .nav-item { padding: 9px 11px !important; font-size: 12px !important; }
}

/* ---------- impresion ---------- */
@media print {
  .sidebar, .nav, .boton, .boton-secundario { display: none !important; }
  .app { grid-template-columns: 1fr !important; }
}
"""


@app.get("/estilos-responsive.css")
def estilos_responsive():
    return Response(
        content=CSS_RESPONSIVE,
        media_type="text/css",
        headers={"Cache-Control": "public, max-age=300"},
    )


def _pagina_login(error="", usuario_previo=""):
    plantilla = r"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Ingresar - Auto Prizma Pro</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:#f7f8fc;
color:#101828;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px}
.caja{background:#fff;border:1px solid #e4e7ec;border-radius:18px;padding:36px 32px;
width:100%;max-width:400px;box-shadow:0 12px 40px rgba(16,24,40,.07)}
.marca{display:flex;align-items:center;gap:12px;margin-bottom:26px}
.logo{width:42px;height:42px;border-radius:12px;background:#1570ef;display:flex;
align-items:center;justify-content:center}
.logo svg{width:26px;height:26px;fill:#fff}
.marca strong{display:block;font-size:16px}
.marca span{font-size:12px;color:#667085}
h1{font-size:20px;margin-bottom:6px}
.sub{font-size:13px;color:#667085;margin-bottom:22px}
label{display:block;font-size:13px;font-weight:600;margin-bottom:6px}
input{width:100%;padding:11px 13px;border:1px solid #d0d5dd;border-radius:10px;
font-size:14px;margin-bottom:16px;font-family:inherit}
input:focus{outline:none;border-color:#1570ef;box-shadow:0 0 0 3px rgba(21,112,239,.12)}
button{width:100%;padding:12px;background:#1570ef;color:#fff;border:none;
border-radius:10px;font-size:15px;font-weight:600;cursor:pointer;font-family:inherit}
button:hover{background:#1059c4}
.error{background:#fef3f2;border:1px solid #fecdca;color:#b42318;padding:11px 13px;
border-radius:10px;font-size:13px;margin-bottom:18px}
.pie{margin-top:20px;font-size:12px;color:#98a2b3;text-align:center;line-height:1.6}
</style><link rel="stylesheet" href="/estilos-responsive.css"></head><body>
<form class="caja" method="post" action="/login">
  <div class="marca">
    <div class="logo"><svg viewBox="0 0 48 48"><path d="M9 35.5 20.5 8.5c.8-1.9 3.4-1.9 4.2 0l4.1 9.6-5.4 12.7-3.1-7.4-5.2 12.1z"/><path d="M26.4 14.5 39 35.5h-8.2l-8.5-14.2z"/></svg></div>
    <div><strong>Auto Prizma Pro</strong><span>Automatizacion PRIZMA</span></div>
  </div>
  <h1>Ingresar</h1>
  <p class="sub">Usa el usuario que te entregaron.</p>
  __ERROR__
  <label for="usuario">Usuario</label>
  <input id="usuario" name="usuario" autocomplete="username" required
         value="__USUARIO__" autofocus>
  <label for="contrasena">Contrasena</label>
  <input id="contrasena" name="contrasena" type="password"
         autocomplete="current-password" required>
  <button type="submit">Entrar</button>
  <p class="pie">Cada persona ve unicamente sus propios cargues,<br>su historial y sus reportes.</p>
</form></body></html>"""

    bloque_error = ""
    if error:
        bloque_error = '<div class="error">' + html.escape(str(error)) + "</div>"

    return plantilla.replace(
        "__ERROR__",
        bloque_error,
    ).replace(
        "__USUARIO__",
        html.escape(str(usuario_previo or ""), quote=True),
    )


@app.middleware("http")
async def exigir_sesion(request: FastAPIRequest, call_next):
    ruta = request.url.path

    if ruta in RUTAS_PUBLICAS:
        return await call_next(request)

    usuario = _usuario_de_request(request)

    if not usuario:
        acepta = str(request.headers.get("accept") or "")

        if ruta.startswith("/estado/") or "application/json" in acepta:
            return JSONResponse(
                {"error": "Sesion expirada", "login": "/login"},
                status_code=401,
            )

        return RedirectResponse("/login", status_code=303)

    testigo = USUARIO_ACTUAL.set(usuario)

    try:
        return await call_next(request)
    finally:
        USUARIO_ACTUAL.reset(testigo)


@app.get("/login", response_class=HTMLResponse)
def login_formulario(request: FastAPIRequest):
    if _usuario_de_request(request):
        return RedirectResponse("/", status_code=303)

    return HTMLResponse(_pagina_login())


@app.post("/login", response_class=HTMLResponse)
def login_enviar(
    request: FastAPIRequest,
    usuario: str = Form(...),
    contrasena: str = Form(...),
):
    usuario = _normalizar_usuario(usuario)

    origen = str(
        request.client.host if request.client else "desconocido"
    )
    clave_intentos = usuario + "|" + origen

    espera = _bloqueado(clave_intentos)

    if espera:
        return HTMLResponse(
            _pagina_login(
                "Demasiados intentos fallidos. Espera "
                + str(espera)
                + " segundos.",
                usuario,
            ),
            status_code=429,
        )

    if not _cargar_usuarios():
        return HTMLResponse(
            _pagina_login(
                "Todavia no hay usuarios creados. Ejecuta: "
                "python gestionar_usuarios.py crear",
                usuario,
            ),
            status_code=503,
        )

    autenticado = _autenticar(usuario, contrasena)

    if not autenticado:
        _registrar_fallo(clave_intentos)
        return HTMLResponse(
            _pagina_login("Usuario o contrasena incorrectos.", usuario),
            status_code=401,
        )

    _limpiar_fallos(clave_intentos)

    respuesta = RedirectResponse("/", status_code=303)
    respuesta.set_cookie(
        SESION_COOKIE,
        _firmar_sesion(autenticado),
        max_age=SESION_HORAS * 3600,
        httponly=True,
        samesite="lax",
        secure=_cookie_segura(),
        path="/",
    )

    print("Login OK:", autenticado, "desde", origen)

    return respuesta


@app.get("/salir")
@app.post("/salir")
def cerrar_sesion():
    respuesta = RedirectResponse("/login", status_code=303)
    respuesta.delete_cookie(SESION_COOKIE, path="/")
    return respuesta


@app.get("/cambiar-clave", response_class=HTMLResponse)
def cambiar_clave_formulario(mensaje: str = "", error: str = ""):
    plantilla = r"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Cambiar contrasena - Auto Prizma Pro</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#f7f8fc;color:#101828;
min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px}
.caja{background:#fff;border:1px solid #e4e7ec;border-radius:18px;padding:34px 32px;
width:100%;max-width:400px;box-shadow:0 12px 40px rgba(16,24,40,.07)}
h1{font-size:19px;margin-bottom:6px}
.sub{font-size:13px;color:#667085;margin-bottom:20px}
label{display:block;font-size:13px;font-weight:600;margin-bottom:6px}
input{width:100%;padding:11px 13px;border:1px solid #d0d5dd;border-radius:10px;
font-size:14px;margin-bottom:16px;font-family:inherit}
button{width:100%;padding:12px;background:#1570ef;color:#fff;border:none;
border-radius:10px;font-size:15px;font-weight:600;cursor:pointer;font-family:inherit}
.error{background:#fef3f2;border:1px solid #fecdca;color:#b42318;padding:11px;
border-radius:10px;font-size:13px;margin-bottom:16px}
.ok{background:#ecfdf3;border:1px solid #abefc6;color:#067647;padding:11px;
border-radius:10px;font-size:13px;margin-bottom:16px}
.volver{display:block;text-align:center;margin-top:16px;font-size:13px;color:#667085}
</style><link rel="stylesheet" href="/estilos-responsive.css"></head><body>
<form class="caja" method="post" action="/cambiar-clave">
  <h1>Cambiar contrasena</h1>
  <p class="sub">Sesion de <strong>__USUARIO__</strong></p>
  __AVISO__
  <label>Contrasena actual</label>
  <input name="actual" type="password" autocomplete="current-password" required>
  <label>Nueva contrasena</label>
  <input name="nueva" type="password" autocomplete="new-password" required
         minlength="8">
  <label>Repite la nueva</label>
  <input name="repetir" type="password" autocomplete="new-password" required
         minlength="8">
  <button type="submit">Guardar</button>
  <a class="volver" href="/">Volver al inicio</a>
</form></body></html>"""

    aviso = ""
    if error:
        aviso = '<div class="error">' + html.escape(str(error)) + "</div>"
    elif mensaje:
        aviso = '<div class="ok">' + html.escape(str(mensaje)) + "</div>"

    return HTMLResponse(
        plantilla.replace(
            "__USUARIO__",
            html.escape(_nombre_visible(_usuario_actual())),
        ).replace("__AVISO__", aviso)
    )


@app.post("/cambiar-clave", response_class=HTMLResponse)
def cambiar_clave_enviar(
    actual: str = Form(...),
    nueva: str = Form(...),
    repetir: str = Form(...),
):
    usuario = _usuario_actual()

    if not _autenticar(usuario, actual):
        return cambiar_clave_formulario(
            error="La contrasena actual no es correcta."
        )

    if nueva != repetir:
        return cambiar_clave_formulario(
            error="La nueva contrasena y su repeticion no coinciden."
        )

    try:
        _crear_usuario(usuario, nueva)
    except ValueError as problema:
        return cambiar_clave_formulario(error=str(problema))

    return cambiar_clave_formulario(
        mensaje="Contrasena actualizada."
    )


# ============================================================
# GOOGLE OAUTH (por usuario)
# ============================================================

@app.get("/google/conectar")
def google_conectar():
    try:
        client_config, redirect_uri = _google_client_config()

        flow = Flow.from_client_config(
            client_config,
            scopes=GOOGLE_SCOPES,
            autogenerate_code_verifier=True,
        )

        flow.redirect_uri = redirect_uri

        authorization_response = (
            redirect_uri
            + ("?" + request.url.query if request.url.query else "")
        )

        flow.fetch_token(
            authorization_response=authorization_response
        )

        _guardar_credenciales_google(
            flow.credentials,
            usuario_del_flujo or _usuario_actual(),
        )

        return RedirectResponse(
            "/",
            status_code=303,
        )

    except Exception as error:
        return HTMLResponse(
            generar_html(
                error=(
                    "Google respondió, pero no fue posible guardar "
                    "la autorización. "
                    + str(error)
                )
            ),
            status_code=500,
        )


@app.post("/google/desconectar")
def google_desconectar():
    try:
        ruta = _ruta_token_google(_usuario_actual())

        if os.path.isfile(ruta):
            os.remove(ruta)
    except OSError:
        pass

    return RedirectResponse(
        "/",
        status_code=303,
    )


# ============================================================
# INICIO
# ============================================================

@app.get(
    "/",
    response_class=HTMLResponse,
)
def inicio():

    return generar_html()


# ============================================================
# ANALIZAR
# ============================================================

@app.post(
    "/analizar",
    response_class=HTMLResponse,
)
async def analizar(
    google_sheet_url: str = Form(...),
    recursos: UploadFile = File(...),
    ovi: str | None = Form(default=None),
    ova: str | None = Form(default=None),
    retos: str | None = Form(default=None),
):

    try:

        if not recursos.filename.lower().endswith(
            ".zip"
        ):

            return generar_html(
                error=(
                    "Los recursos deben estar "
                    "en un archivo .zip"
                )
            )

        procesar_ovi = (
            ovi is not None
        )

        procesar_ova = (
            ova is not None
        )

        procesar_retos = (
            retos is not None
        )

        if (
            not procesar_ovi
            and not procesar_ova
            and not procesar_retos
        ):

            return generar_html(
                error=(
                    "Selecciona OVI, OVA y/o Retos Evaluativos."
                )
            )

        contenido_excel = descargar_google_sheet_xlsx(
            google_sheet_url
        )

        # Solo se procesa la pestaña que viene en el enlace (#gid=...).
        hoja_objetivo = _resolver_hoja_objetivo(google_sheet_url)

        contenido_zip = await recursos.read()

        hojas = analizar_excel(
            contenido_excel,
            procesar_ovi,
            procesar_ova,
            procesar_retos,
            hoja_objetivo,
        )

        if not hojas:

            if hoja_objetivo:
                return generar_html(
                    error=(
                        "La pestaña \"" + hoja_objetivo + "\" no tiene una "
                        "matriz PRIZMA válida (no encontré la fila de "
                        "encabezado \"Semana correspondiente\"). "
                        "Abre la pestaña correcta y copia el enlace otra vez."
                    )
                )

            return generar_html(
                error=(
                    "No encontré una matriz "
                    "PRIZMA válida."
                )
            )

        try:

            zip_info = analizar_zip(
                contenido_zip
            )

        except zipfile.BadZipFile:

            return generar_html(
                error=(
                    "El ZIP no es válido."
                )
            )

        # ====================================================
        # CREAR TRABAJO
        # ====================================================

        trabajo_id = uuid.uuid4().hex

        carpeta_trabajo = os.path.join(
            UPLOADS_DIR,
            trabajo_id,
        )

        os.makedirs(
            carpeta_trabajo,
            exist_ok=True,
        )

        ruta_excel = os.path.join(
            carpeta_trabajo,
            "matriz.xlsx",
        )

        ruta_zip = os.path.join(
            carpeta_trabajo,
            "recursos.zip",
        )

        with open(
            ruta_excel,
            "wb",
        ) as archivo:

            archivo.write(
                contenido_excel
            )

        with open(
            ruta_zip,
            "wb",
        ) as archivo:

            archivo.write(
                contenido_zip
            )

        carpeta_temp = os.path.join(
            TEMP_DIR,
            trabajo_id,
        )

        os.makedirs(
            carpeta_temp,
            exist_ok=True,
        )

        preflight_drive = _prevalidar_recursos_drive_zip_ultrarapido(
            google_sheet_url,
            ruta_excel,
            ruta_zip,
            procesar_ovi,
            procesar_ova,
            procesar_retos,
            hoja_objetivo,
        )

        recursos_drive_resueltos = preflight_drive.pop(
            "_mapa_resueltos",
            {},
        )

        if not preflight_drive.get("ok"):
            errores_drive = preflight_drive.get("errores") or []
            resumen_errores = " | ".join(errores_drive[:8])

            if len(errores_drive) > 8:
                resumen_errores += (
                    " | ... y "
                    + str(len(errores_drive) - 8)
                    + " error(es) más."
                )

            shutil.rmtree(carpeta_trabajo, ignore_errors=True)
            shutil.rmtree(carpeta_temp, ignore_errors=True)

            return generar_html(
                error=(
                    "La validación Google Drive vs ZIP encontró problemas. "
                    + resumen_errores
                )
            )

        cursos_trabajo = _cursos_desde_hojas(hojas)

        ruta_reporte = os.path.join(
            RESULTADOS_DIR,
            _nombre_reporte(cursos_trabajo, trabajo_id),
        )

        TRABAJOS[
            trabajo_id
        ] = {
            "id":
                trabajo_id,

            "ruta_excel":
                ruta_excel,

            "ruta_zip":
                ruta_zip,

            "carpeta_temp":
                carpeta_temp,

            "ruta_reporte":
                ruta_reporte,

            "procesar_ovi":
                procesar_ovi,

            "procesar_ova":
                procesar_ova,

            "procesar_retos":
                procesar_retos,

            "etapa":
                "analizado",

            "mensaje":
                "Archivos analizados correctamente.",

            "total":
                0,

            "procesadas":
                0,

            "exitosas":
                0,

            "errores":
                0,

            "terminado":
                False,

            "detalle_actividades":
                [],

            "recursos_drive_resueltos":
                recursos_drive_resueltos,

            "hoja_objetivo":
                hoja_objetivo,

            "usuario_app":
                _usuario_actual(),

            "cursos":
                cursos_trabajo,

            "historial_registrado":
                False,

            "creado_en":
                _ahora_colombia_iso(),

            "cancelar_solicitado":
                False,
        }

        resultado = {
            "hojas":
                hojas,

            "zip":
                zip_info,

            "drive":
                preflight_drive,
        }

        TRABAJOS[trabajo_id]["resultado_analisis"] = resultado

        return generar_html(
            resultado=resultado,
            trabajo_id=trabajo_id,
        )

    except Exception as e:

        return generar_html(
            error=str(e)
        )


# ============================================================
# INICIAR
# ============================================================

@app.post(
    "/iniciar/{trabajo_id}",
    response_class=HTMLResponse,
)
def iniciar_trabajo(
    trabajo_id: str,
    background_tasks: BackgroundTasks,
    usuario_prizma: str = Form(...),
    contrasena_prizma: str = Form(...),
):
    trabajo = TRABAJOS.get(trabajo_id)
    if not trabajo or not _es_de(trabajo, _usuario_actual()):
        return HTMLResponse("<h2>Trabajo no encontrado.</h2><a href='/'>Volver</a>", status_code=404)

    usuario_prizma = usuario_prizma.strip()
    if not usuario_prizma or not contrasena_prizma:
        return generar_html(
            resultado=trabajo.get("resultado_analisis"),
            trabajo_id=trabajo_id,
            error="Debes ingresar el usuario y la contraseña de PRIZMA.",
        )

    if trabajo.get("etapa") not in ["analizado", "error"]:
        return HTMLResponse("<h2>Este trabajo ya fue iniciado.</h2>")

    # Validar el acceso antes de encolar el curso.
    trabajo["etapa"] = "validando_login"
    trabajo["mensaje"] = "Validando usuario y contraseña en PRIZMA..."
    validacion = validar_credenciales_prizma(usuario_prizma, contrasena_prizma)

    if not validacion.get("ok"):
        trabajo["etapa"] = "analizado"
        trabajo["mensaje"] = validacion.get("mensaje", "No fue posible validar el acceso a PRIZMA.")
        return generar_html(
            resultado=trabajo.get("resultado_analisis"),
            trabajo_id=trabajo_id,
            error=trabajo["mensaje"],
        )

    _encolar_trabajo(trabajo_id, usuario_prizma, contrasena_prizma)

    html_progreso = r"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Cargue en progreso - Auto Prizma Pro</title>
        <style>
            :root {
                --fondo:#f7f8fc; --panel:#fff; --texto:#101828; --muted:#667085;
                --borde:#e5e7ef; --morado:#5548e8; --verde:#0ea968; --rojo:#ef4444;
                --sombra:0 10px 30px rgba(29,41,57,.06);
            }
            * { box-sizing:border-box; }
            body { margin:0; font-family:Inter,ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; background:var(--fondo); color:var(--texto); }
            .app { min-height:100vh; display:grid; grid-template-columns:235px 1fr; }
            .sidebar { position:sticky; top:0; height:100vh; background:#fff; border-right:1px solid var(--borde); padding:28px 20px; display:flex; flex-direction:column; }
            .marca { display:flex; align-items:center; gap:12px; margin-bottom:34px; }
            .logo { width:44px; height:44px; border-radius:13px; display:grid; place-items:center; background:linear-gradient(145deg,#6d5dfc,#4338ca); box-shadow:0 8px 20px rgba(79,70,229,.25); flex:0 0 44px; }
            .logo svg{width:29px;height:29px;overflow:visible}.logo svg path:first-child{fill:#fff}.logo svg path:last-child{fill:#c7d2fe}
            .marca strong { display:block; font-size:18px; }
            .marca span { display:block; color:var(--muted); font-size:12px; margin-top:3px; }
            .nav { display:grid; gap:8px; }
            .nav-item { padding:12px 14px; border-radius:11px; color:#475467; font-size:14px; display:flex; gap:11px; align-items:center; text-decoration:none; transition:.18s ease; }
            a.nav-item:hover{background:#f7f5ff;color:#4f46e5}.nav-item.activo{background:#f1efff;color:#4f46e5;font-weight:700}.nav-item.proximamente{opacity:.48;cursor:default}.nav-item.proximamente small{margin-left:auto;font-size:9px}
            .estado-servicio { margin-top:auto; border:1px solid var(--borde); border-radius:14px; padding:15px; background:#fff; }
            .servicio-linea { display:flex; align-items:center; gap:8px; color:#07894f; font-size:13px; font-weight:700; }
            .punto { width:8px; height:8px; border-radius:50%; background:#15b76a; }
            .servicio-mini { margin-top:13px; display:flex; justify-content:space-between; font-size:12px; color:var(--muted); }
            .chip { background:#eef4ff; color:#3538cd; padding:4px 8px; border-radius:999px; }
            .contenido { padding:28px 34px 45px; max-width:1450px; width:100%; margin:0 auto; }
            .panel { background:#fff; border:1px solid var(--borde); border-radius:16px; box-shadow:var(--sombra); }
            .cabecera-progreso { padding:25px 28px; }
            .cabecera-linea { display:flex; justify-content:space-between; align-items:center; gap:20px; }
            .cabecera-titulo { display:flex; gap:13px; align-items:center; }
            .icono-subida { width:43px; height:43px; border-radius:12px; background:#f1efff; color:#4f46e5; display:grid; place-items:center; font-size:21px; }
            .cabecera-titulo h1 { margin:0; font-size:22px; }
            .porcentaje { font-size:14px; color:var(--muted); }
            .porcentaje strong { font-size:18px; color:#4f46e5; }
            .barra { height:11px; background:#eef0f5; border-radius:999px; overflow:hidden; margin:22px 0 14px; }
            .progreso { height:100%; width:0%; border-radius:999px; background:linear-gradient(90deg,#6759f5,#4f46e5); transition:width .35s ease; }
            .estado { color:#4f46e5; font-size:14px; font-weight:600; line-height:1.45; }
            .numeros { display:grid; grid-template-columns:repeat(3,1fr); gap:18px; margin:20px 0; }
            .numero { padding:20px; display:flex; align-items:center; gap:15px; }
            .numero-icono { width:48px; height:48px; border-radius:50%; display:grid; place-items:center; font-weight:900; font-size:20px; }
            .numero.procesadas .numero-icono { background:#eef4ff; color:#4f46e5; }
            .numero.exitosas .numero-icono { background:#e9f9f1; color:#0a9b5b; }
            .numero.errores .numero-icono { background:#fff0f1; color:#e5484d; }
            .numero span { display:block; color:var(--muted); font-size:13px; }
            .numero strong { display:block; font-size:28px; margin-top:2px; }
            .zona { display:grid; grid-template-columns:minmax(0,1.55fr) minmax(280px,.75fr); gap:20px; align-items:start; }
            .lista-panel, .estado-panel { padding:23px; }
            .titulo-panel { margin:0 0 16px; font-size:17px; }
            .lista-actividades { max-height:570px; overflow:auto; border:1px solid var(--borde); border-radius:12px; }
            .actividad-progreso { display:grid; grid-template-columns:27px 1fr; gap:10px; align-items:start; padding:11px 13px; border-bottom:1px solid #eef0f4; background:#fff; }
            .actividad-progreso:last-child { border-bottom:0; }
            .actividad-icono { width:23px; height:23px; display:grid; place-items:center; font-size:15px; }
            .actividad-nombre { font-size:13px; line-height:1.4; }
            .actividad-error { color:#b42318; font-size:11px; margin-top:4px; word-break:break-word; }
            .actividad-ok { background:#fbfffd; }
            .actividad-error-fila { background:#fffafa; }
            .actividad-procesando { background:#f7f6ff; box-shadow:inset 3px 0 0 #5b48e8; }
            .estado-panel .estado-caja { background:#f8f9fc; border:1px solid var(--borde); border-radius:12px; padding:17px; margin-bottom:14px; }
            .estado-caja span { display:block; color:var(--muted); font-size:12px; }
            .estado-caja strong { display:block; margin-top:6px; font-size:14px; line-height:1.45; }
            .leyenda { display:grid; gap:12px; margin-top:22px; }
            .leyenda div { display:flex; align-items:center; gap:9px; color:#475467; font-size:13px; }
            .final { margin-top:18px; padding:18px; border-radius:12px; background:#ecfdf5; border:1px solid #a7f3d0; }
            .final h3 { margin:0 0 6px; }
            .final p { margin:0; color:#475467; font-size:13px; }
            .descarga { display:inline-flex; margin-top:14px; padding:11px 14px; border-radius:9px; background:#4f46e5; color:#fff; text-decoration:none; font-size:13px; font-weight:800; }
            .pie { margin-top:18px; color:var(--muted); font-size:12px; text-align:center; }
            @media(max-width:1050px) { .app{grid-template-columns:1fr}.sidebar{display:none}.contenido{padding:20px}.zona{grid-template-columns:1fr} }
            @media(max-width:680px) { .numeros{grid-template-columns:1fr}.contenido{padding:14px}.cabecera-linea{align-items:flex-start;flex-direction:column} }
        </style><link rel="stylesheet" href="/estilos-responsive.css">
    </head>
    <body>
        <div class="app">
            <aside class="sidebar">
                <div class="marca">
                    <div class="logo" aria-label="Auto Prizma Pro">
                        <svg viewBox="0 0 48 48" aria-hidden="true">
                            <path d="M9 35.5 20.5 8.5c.8-1.9 3.4-1.9 4.2 0l4.1 9.6-5.4 12.7-3.1-7.4-5.2 12.1z"/>
                            <path d="M26.4 14.5 39 35.5h-8.2l-8.5-14.2z"/>
                        </svg>
                    </div>
                    <div><strong>Auto Prizma Pro</strong><span>Cargue automático</span></div>
                </div>
                <nav class="nav">
                    <a class="nav-item" href="/">⌂ <span>Inicio</span></a>
                    <a class="nav-item activo" href="/cargue-actual">⇧ <span>Cargue actual</span></a>
                    <a class="nav-item" href="/historial">◷ <span>Historial</span></a>
                    <a class="nav-item" href="/reportes">▥ <span>Reportes</span></a><a class="nav-item" href="/cambiar-clave">✎ <span>Mi contraseña</span></a><a class="nav-item" href="/salir">⏻ <span>Salir</span></a>
                </nav>
                <div class="estado-servicio">
                    <div class="servicio-linea"><span class="punto"></span> Servicio activo</div>
                    <div class="servicio-mini"><span>Navegador</span><span class="chip">Chromium</span></div>
                    <div class="servicio-mini"><span>Conexión</span><span class="chip">Estable</span></div>
                </div>
            </aside>

            <main id="cargue-progreso" class="contenido">
                <section class="panel cabecera-progreso">
                    <div class="cabecera-linea">
                        <div class="cabecera-titulo">
                            <div class="icono-subida">⇧</div>
                            <h1>Cargue en progreso</h1>
                        </div>
                        <div class="porcentaje"><strong id="porcentaje">0%</strong> completado</div>
                    </div>
                    <div class="barra"><div id="progreso" class="progreso"></div></div>
                    <div id="estado" class="estado">Iniciando...</div>
                </section>

                <section class="numeros">
                    <div class="panel numero procesadas">
                        <div class="numero-icono">↻</div>
                        <div><span>Procesadas</span><strong id="procesadas">0</strong></div>
                    </div>
                    <div class="panel numero exitosas">
                        <div class="numero-icono">✓</div>
                        <div><span>Exitosas</span><strong id="exitosas">0</strong></div>
                    </div>
                    <div class="panel numero errores">
                        <div class="numero-icono">×</div>
                        <div><span>Errores</span><strong id="errores">0</strong></div>
                    </div>
                </section>

                <section class="zona">
                    <div class="panel lista-panel">
                        <h2 class="titulo-panel">Actividades del cargue</h2>
                        <div id="lista-actividades" class="lista-actividades">
                            <div class="actividad-progreso">Preparando lista de actividades...</div>
                        </div>
                        <div class="pie">El progreso se actualiza automáticamente.</div>
                    </div>

                    <div class="panel estado-panel">
                        <h2 class="titulo-panel">Estado del proceso</h2>
                        <div class="estado-caja">
                            <span>Actividad actual</span>
                            <strong id="actividad-actual">Preparando cargue...</strong>
                        </div>
                        <div class="estado-caja">
                            <span>Total de actividades</span>
                            <strong id="total-actividades">—</strong>
                        </div>
                        <div class="leyenda">
                            <div>⏳ Pendiente</div>
                            <div>🔄 Procesando</div>
                            <div>✅ Completada</div>
                            <div>❌ Con error</div>
                        </div>
                        <form action="/cancelar/__TRABAJO_ID__" method="post" style="margin-top:18px;">
                            <button type="submit" style="width:100%;border:1px solid #fecaca;background:#fff;color:#b42318;border-radius:9px;padding:11px 14px;cursor:pointer;font-weight:800;">Cancelar proceso</button>
                        </form>
                        <div id="final"></div>
                    </div>
                </section>
            </main>
        </div>

        <script>
            const trabajoId = "__TRABAJO_ID__";

            function escaparHtml(texto) {
                return String(texto ?? "")
                    .replaceAll("&", "&amp;")
                    .replaceAll("<", "&lt;")
                    .replaceAll(">", "&gt;")
                    .replaceAll('"', "&quot;")
                    .replaceAll("'", "&#039;");
            }

            function renderizarActividades(actividades) {
                const contenedor = document.getElementById("lista-actividades");

                if (!Array.isArray(actividades) || actividades.length === 0) {
                    contenedor.innerHTML = '<div class="actividad-progreso">Preparando lista de actividades...</div>';
                    return;
                }

                let html = "";
                let actual = "Preparando cargue...";

                for (const actividad of actividades) {
                    let icono = "⏳";
                    let clase = "";

                    if (actividad.estado === "procesando") {
                        icono = "🔄";
                        clase = "actividad-procesando";
                        actual = actividad.numero + ". " + actividad.nombre;
                    } else if (actividad.estado === "ok") {
                        icono = "✅";
                        clase = "actividad-ok";
                    } else if (actividad.estado === "error") {
                        icono = "❌";
                        clase = "actividad-error-fila";
                    }

                    html += '<div class="actividad-progreso ' + clase + '">' +
                        '<div class="actividad-icono">' + icono + '</div>' +
                        '<div><div class="actividad-nombre">' +
                        actividad.numero + '. ' + escaparHtml(actividad.nombre) + '</div>';

                    if (actividad.estado === "error" && actividad.error) {
                        html += '<div class="actividad-error">' + escaparHtml(actividad.error) + '</div>';
                    }

                    html += '</div></div>';
                }

                contenedor.innerHTML = html;
                document.getElementById("actividad-actual").innerText = actual;
            }

            async function revisarEstado() {
                try {
                    const respuesta = await fetch("/estado/" + trabajoId);
                    const datos = await respuesta.json();

                    document.getElementById("estado").innerText = datos.mensaje + (datos.posicion_cola ? ' · Posición en cola: ' + datos.posicion_cola : '');
                    document.getElementById("procesadas").innerText = datos.procesadas;
                    document.getElementById("exitosas").innerText = datos.exitosas;
                    document.getElementById("errores").innerText = datos.errores;
                    document.getElementById("total-actividades").innerText = datos.total || "—";

                    renderizarActividades(datos.detalle_actividades);

                    let porcentaje = 0;
                    if (datos.total > 0) {
                        porcentaje = (datos.procesadas / datos.total) * 100;
                    }
                    porcentaje = Math.min(100, porcentaje);

                    document.getElementById("progreso").style.width = porcentaje + "%";
                    document.getElementById("porcentaje").innerText = Math.round(porcentaje) + "%";

                    if (datos.terminado) {
                        document.getElementById("actividad-actual").innerText = "Proceso finalizado";

                        let html = '<div class="final"><h3>Proceso terminado</h3><p>' +
                            escaparHtml(datos.mensaje) + '</p>';

                        if (datos.reporte_disponible) {
                            html += '<a class="descarga" href="/reporte/' + trabajoId + '">Descargar reporte CSV</a>';
                        }

                        html += '</div>';
                        document.getElementById("final").innerHTML = html;
                        return;
                    }

                    setTimeout(revisarEstado, 1500);
                } catch (error) {
                    document.getElementById("estado").innerText = "Error consultando estado.";
                    setTimeout(revisarEstado, 3000);
                }
            }

            revisarEstado();
        </script>
    </body>
    </html>
    """

    pagina_progreso = html_progreso.replace(
        "__TRABAJO_ID__",
        trabajo_id,
    )
    trabajo["pagina_progreso"] = pagina_progreso

    return HTMLResponse(pagina_progreso)


def _pagina_sin_cargue_actual():
    return HTMLResponse(r'''<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Cargue actual - Auto Prizma Pro</title>
<style>
:root{--fondo:#f7f8fc;--texto:#101828;--muted:#667085;--borde:#e5e7ef;--morado:#5548e8;--sombra:0 12px 34px rgba(29,41,57,.06)}*{box-sizing:border-box}body{margin:0;font-family:Inter,ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--fondo);color:var(--texto)}.app{min-height:100vh;display:grid;grid-template-columns:235px 1fr}.sidebar{height:100vh;background:#fff;border-right:1px solid var(--borde);padding:28px 20px;display:flex;flex-direction:column}.marca{display:flex;align-items:center;gap:12px;margin-bottom:34px}.logo{width:44px;height:44px;border-radius:13px;display:grid;place-items:center;background:linear-gradient(145deg,#6d5dfc,#4338ca);box-shadow:0 8px 20px rgba(79,70,229,.25)}.logo svg{width:29px;height:29px}.logo svg path:first-child{fill:#fff}.logo svg path:last-child{fill:#c7d2fe}.marca strong{display:block;font-size:18px}.marca span{display:block;color:var(--muted);font-size:12px;margin-top:3px}.nav{display:grid;gap:8px}.nav-item{padding:12px 14px;border-radius:11px;color:#475467;font-size:14px;display:flex;gap:11px;align-items:center;text-decoration:none}.nav-item:hover{background:#f7f5ff;color:#4f46e5}.nav-item.activo{background:#f1efff;color:#4f46e5;font-weight:700}.estado-servicio{margin-top:auto;border:1px solid var(--borde);border-radius:14px;padding:15px}.servicio-linea{font-size:12px;font-weight:800;color:#07894f;margin-bottom:14px}.punto{width:8px;height:8px;background:#12b76a;border-radius:50%;display:inline-block;margin-right:7px}.servicio-mini{display:flex;justify-content:space-between;align-items:center;font-size:11px;color:var(--muted);margin-top:10px}.chip{background:#eef2ff;color:#4f46e5;border-radius:999px;padding:4px 8px}.contenido{padding:30px 34px;display:grid;place-items:center}.vacio{width:min(680px,100%);background:#fff;border:1px solid var(--borde);border-radius:18px;box-shadow:var(--sombra);padding:54px 34px;text-align:center}.icono{width:68px;height:68px;border-radius:20px;background:#f1efff;color:#5548e8;display:grid;place-items:center;margin:0 auto 20px;font-size:28px;font-weight:800}.vacio h1{margin:0 0 10px;font-size:27px}.vacio p{margin:0 auto 24px;color:var(--muted);font-size:14px;line-height:1.6;max-width:480px}.boton{display:inline-flex;padding:12px 18px;border-radius:10px;background:linear-gradient(90deg,#5548e8,#6546e8);color:#fff;text-decoration:none;font-size:13px;font-weight:800}@media(max-width:900px){.app{grid-template-columns:1fr}.sidebar{display:none}.contenido{padding:18px}}
</style><link rel="stylesheet" href="/estilos-responsive.css"></head><body><div class="app"><aside class="sidebar"><div class="marca"><div class="logo"><svg viewBox="0 0 48 48"><path d="M9 35.5 20.5 8.5c.8-1.9 3.4-1.9 4.2 0l4.1 9.6-5.4 12.7-3.1-7.4-5.2 12.1z"/><path d="M26.4 14.5 39 35.5h-8.2l-8.5-14.2z"/></svg></div><div><strong>Auto Prizma Pro</strong><span>Automatización PRIZMA</span></div></div><nav class="nav"><a class="nav-item" href="/">⌂ <span>Inicio</span></a><a class="nav-item activo" href="/cargue-actual">⇧ <span>Cargue actual</span></a><a class="nav-item" href="/historial">◷ <span>Historial</span></a><a class="nav-item" href="/reportes">▥ <span>Reportes</span></a><a class="nav-item" href="/cambiar-clave">✎ <span>Mi contraseña</span></a><a class="nav-item" href="/salir">⏻ <span>Salir</span></a></nav><div class="estado-servicio"><div class="servicio-linea"><span class="punto"></span> Servicio activo</div><div class="servicio-mini"><span>Navegador</span><span class="chip">Chromium</span></div><div class="servicio-mini"><span>Conexión</span><span class="chip">Estable</span></div></div></aside><main class="contenido"><section class="vacio"><div class="icono">⇧</div><h1>No hay cargues activos</h1><p>Cuando alguien inicie un proceso desde Inicio, aparecerá aquí para que el equipo pueda consultar su progreso.</p><a class="boton" href="/">Iniciar un nuevo cargue</a></section></main></div></body></html>''')


def _formatear_inicio_cargue(fecha_iso):
    if not fecha_iso:
        return "Iniciando..."
    try:
        fecha = datetime.fromisoformat(fecha_iso)
        if fecha.tzinfo is None:
            fecha = fecha.replace(tzinfo=ZONA_HORARIA_COLOMBIA)
        else:
            fecha = fecha.astimezone(ZONA_HORARIA_COLOMBIA)
        return fecha.strftime("%d/%m/%Y · %I:%M %p").replace("AM", "a. m.").replace("PM", "p. m.")
    except Exception:
        return "En progreso"


def _pagina_cargues_activos(trabajos):
    tarjetas = []
    for trabajo in trabajos:
        cursos = trabajo.get("cursos") or []
        if cursos:
            curso = cursos[0].get("curso") or "Curso sin nombre"
            programa = cursos[0].get("programa") or "Programa sin nombre"
            if len(cursos) > 1:
                curso = f"{curso} y {len(cursos) - 1} curso(s) más"
        else:
            curso, programa = "Curso en proceso", "Programa sin nombre"
        total = int(trabajo.get("total") or 0)
        procesadas = int(trabajo.get("procesadas") or 0)
        exitosas = int(trabajo.get("exitosas") or 0)
        errores = int(trabajo.get("errores") or 0)
        porcentaje = max(0, min(100, round((procesadas / total) * 100) if total > 0 else 0))
        trabajo_id = html.escape(str(trabajo.get("id") or ""))
        en_cola = trabajo.get("etapa") == "en_cola"
        posicion = _posicion_en_cola(str(trabajo.get("id") or "")) if en_cola else None
        estado_texto = f"En cola · #{posicion}" if posicion else "En progreso"
        inicio_label = "En cola desde" if en_cola else "Iniciado"
        fecha_mostrar = trabajo.get("encolado_en") if en_cola else trabajo.get("iniciado_en")
        tarjetas.append(f'''<article class="carga-card"><div class="curso-bloque"><div class="curso-icono">▣</div><div class="curso-texto"><h2>{html.escape(curso)}</h2><p>Programa: {html.escape(programa)}</p><span class="estado-chip"><span></span> {html.escape(estado_texto)}</span></div></div><div class="inicio-bloque"><small>◷ {inicio_label}</small><strong>{html.escape(_formatear_inicio_cargue(fecha_mostrar))}</strong></div><div class="avance-bloque"><div class="avance-cab"><span>Progreso general</span><strong>{porcentaje}%</strong></div><div class="barra"><div class="barra-interna" style="width:{porcentaje}%"></div></div><div class="metricas"><div><i class="verde"></i><small>Procesadas</small><strong>{procesadas}</strong></div><div><i class="verde"></i><small>Exitosas</small><strong>{exitosas}</strong></div><div><i class="rojo"></i><small>Errores</small><strong>{errores}</strong></div></div></div><div class="accion-bloque"><a href="/proceso/{trabajo_id}">Entrar al proceso <b>›</b></a><form action="/cancelar/{trabajo_id}" method="post" style="margin-top:8px"><button type="submit" style="border:1px solid #fecaca;background:#fff;color:#b42318;border-radius:8px;padding:8px 10px;cursor:pointer;font-weight:700">Cancelar</button></form></div></article>''')
    cuerpo = "".join(tarjetas)
    cantidad = len(trabajos)
    plural = "s" if cantidad != 1 else ""
    return HTMLResponse(f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Cargue actual - Auto Prizma Pro</title><style>
:root{{--fondo:#f7f8fc;--texto:#101828;--muted:#667085;--borde:#e5e7ef;--morado:#5548e8;--sombra:0 10px 30px rgba(29,41,57,.05)}}*{{box-sizing:border-box}}body{{margin:0;font-family:Inter,ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--fondo);color:var(--texto)}}.app{{min-height:100vh;display:grid;grid-template-columns:235px 1fr}}.sidebar{{position:sticky;top:0;height:100vh;background:#fff;border-right:1px solid var(--borde);padding:28px 20px;display:flex;flex-direction:column}}.marca{{display:flex;align-items:center;gap:12px;margin-bottom:34px}}.logo{{width:44px;height:44px;border-radius:13px;display:grid;place-items:center;background:linear-gradient(145deg,#6d5dfc,#4338ca);box-shadow:0 8px 20px rgba(79,70,229,.25)}}.logo svg{{width:29px;height:29px}}.logo svg path:first-child{{fill:#fff}}.logo svg path:last-child{{fill:#c7d2fe}}.marca strong{{display:block;font-size:18px}}.marca span{{display:block;color:var(--muted);font-size:12px;margin-top:3px}}.nav{{display:grid;gap:8px}}.nav-item{{padding:12px 14px;border-radius:11px;color:#475467;font-size:14px;display:flex;gap:11px;align-items:center;text-decoration:none}}.nav-item:hover{{background:#f7f5ff;color:#4f46e5}}.nav-item.activo{{background:#f1efff;color:#4f46e5;font-weight:700}}.estado-servicio{{margin-top:auto;border:1px solid var(--borde);border-radius:14px;padding:15px}}.servicio-linea{{font-size:12px;font-weight:800;color:#07894f;margin-bottom:14px}}.punto{{width:8px;height:8px;background:#12b76a;border-radius:50%;display:inline-block;margin-right:7px}}.servicio-mini{{display:flex;justify-content:space-between;align-items:center;font-size:11px;color:var(--muted);margin-top:10px}}.chip{{background:#eef2ff;color:#4f46e5;border-radius:999px;padding:4px 8px}}.contenido{{padding:34px 36px 44px;min-width:0}}.cabecera-top{{display:flex;justify-content:space-between;gap:20px;align-items:flex-start;margin-bottom:24px}}.cabecera h1{{font-size:29px;margin:0 0 7px}}.cabecera p{{margin:0;color:var(--muted);font-size:14px;line-height:1.5}}.contador{{background:#f1efff;color:#5548e8;border-radius:999px;padding:8px 12px;font-size:12px;font-weight:800;white-space:nowrap}}.lista{{display:grid;gap:16px}}.carga-card{{background:#fff;border:1px solid var(--borde);border-radius:17px;box-shadow:var(--sombra);padding:22px;display:grid;grid-template-columns:minmax(245px,1.15fr) 170px minmax(320px,1.25fr) 185px;gap:22px;align-items:center}}.curso-bloque{{display:flex;gap:14px;align-items:center;min-width:0}}.curso-icono{{width:58px;height:58px;border-radius:15px;background:#f1efff;color:#5548e8;display:grid;place-items:center;font-size:24px;flex:0 0 58px}}.curso-texto{{min-width:0}}.curso-texto h2{{font-size:16px;margin:0 0 5px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.curso-texto p{{font-size:12px;color:var(--muted);margin:0 0 10px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.estado-chip{{display:inline-flex;align-items:center;gap:6px;background:#eef6ff;color:#1473e6;border-radius:999px;padding:5px 9px;font-size:11px;font-weight:800}}.estado-chip span{{width:6px;height:6px;background:#2e90fa;border-radius:50%}}.inicio-bloque small{{display:block;color:var(--muted);font-size:11px;margin-bottom:7px}}.inicio-bloque strong{{font-size:12px;color:#475467}}.avance-cab{{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;color:var(--muted);font-size:11px}}.avance-cab strong{{font-size:21px;color:#5548e8}}.barra{{height:10px;border-radius:999px;background:#eeecff;overflow:hidden}}.barra-interna{{height:100%;background:linear-gradient(90deg,#6759f5,#4f46e5);border-radius:999px}}.metricas{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-top:12px}}.metricas div{{display:grid;grid-template-columns:8px 1fr;column-gap:5px;align-items:center}}.metricas i{{width:7px;height:7px;border-radius:50%}}.metricas i.verde{{background:#12b76a}}.metricas i.rojo{{background:#ef4444}}.metricas small{{font-size:10px;color:var(--muted)}}.metricas strong{{grid-column:2;font-size:14px;margin-top:2px}}.accion-bloque{{border-left:1px solid var(--borde);padding-left:20px;text-align:center}}.accion-bloque a{{display:flex;justify-content:center;align-items:center;gap:8px;background:linear-gradient(90deg,#5548e8,#6546e8);color:#fff;text-decoration:none;font-size:12px;font-weight:800;border-radius:10px;padding:12px 13px}}.accion-bloque b{{font-size:18px}}.accion-bloque small{{display:block;color:var(--muted);font-size:10px;margin-top:8px}}.nota{{margin-top:18px;border:1px solid #ddd8ff;background:#faf9ff;border-radius:14px;padding:14px 17px;color:#475467;font-size:12px}}@media(max-width:1200px){{.carga-card{{grid-template-columns:1fr 1fr}}.accion-bloque{{border-left:0;padding-left:0}}}}@media(max-width:900px){{.app{{grid-template-columns:1fr}}.sidebar{{display:none}}.contenido{{padding:20px}}.carga-card{{grid-template-columns:1fr}}}}
</style><link rel="stylesheet" href="/estilos-responsive.css"></head><body><div class="app"><aside class="sidebar"><div class="marca"><div class="logo"><svg viewBox="0 0 48 48"><path d="M9 35.5 20.5 8.5c.8-1.9 3.4-1.9 4.2 0l4.1 9.6-5.4 12.7-3.1-7.4-5.2 12.1z"/><path d="M26.4 14.5 39 35.5h-8.2l-8.5-14.2z"/></svg></div><div><strong>Auto Prizma Pro</strong><span>Automatización PRIZMA</span></div></div><nav class="nav"><a class="nav-item" href="/">⌂ <span>Inicio</span></a><a class="nav-item activo" href="/cargue-actual">⇧ <span>Cargue actual</span></a><a class="nav-item" href="/historial">◷ <span>Historial</span></a><a class="nav-item" href="/reportes">▥ <span>Reportes</span></a><a class="nav-item" href="/cambiar-clave">✎ <span>Mi contraseña</span></a><a class="nav-item" href="/salir">⏻ <span>Salir</span></a></nav><div class="estado-servicio"><div class="servicio-linea"><span class="punto"></span> Servicio activo</div><div class="servicio-mini"><span>Navegador</span><span class="chip">Chromium</span></div><div class="servicio-mini"><span>Conexión</span><span class="chip">Estable</span></div></div></aside><main class="contenido"><section class="cabecera"><div class="cabecera-top"><div><h1>Cargue actual</h1><p>Aquí puedes ver todos los cursos que están siendo procesados actualmente.<br>Entra a un proceso para revisar el detalle de sus actividades.</p></div><div class="contador">{cantidad} proceso{plural} activo{plural}</div></div></section><section class="lista">{cuerpo}</section><div class="nota"><strong>Vista global del equipo.</strong> Cada tarjeta representa un cargue activo. Para ver la lista detallada de actividades, entra al proceso correspondiente.</div></main></div><script>setTimeout(function(){{window.location.reload();}},3000);</script></body></html>''')


@app.get("/cargue-actual", response_class=HTMLResponse)
def cargue_actual():
    usuario = _usuario_actual()

    activos = [
        trabajo for trabajo in TRABAJOS.values()
        if not trabajo.get("terminado")
        and _es_de(trabajo, usuario)
        and trabajo.get("etapa") in {"en_cola", "iniciando", "login", "preparando", "procesando", "validando_login"}
    ]
    activos.sort(key=lambda item: item.get("encolado_en") or item.get("iniciado_en") or item.get("creado_en", ""))
    if not activos:
        return _pagina_sin_cargue_actual()
    return _pagina_cargues_activos(activos)


@app.post("/cancelar/{trabajo_id}", response_class=HTMLResponse)
def cancelar_trabajo(trabajo_id: str):
    with COLA_CONDICION:
        trabajo = TRABAJOS.get(trabajo_id)
        if not trabajo or not _es_de(trabajo, _usuario_actual()):
            return HTMLResponse("<h2>Proceso no encontrado.</h2>", status_code=404)

        if trabajo.get("terminado"):
            return HTMLResponse(
                "<script>window.location='/cargue-actual';</script>"
            )

        trabajo["cancelar_solicitado"] = True

        if trabajo.get("etapa") == "en_cola":
            try:
                COLA_CARGUES.remove(trabajo_id)
            except ValueError:
                pass
            trabajo.pop("usuario_prizma_temporal", None)
            trabajo.pop("contrasena_prizma_temporal", None)
            trabajo["etapa"] = "cancelado"
            trabajo["mensaje"] = "Proceso cancelado antes de iniciar."
            trabajo["terminado"] = True
            trabajo["finalizado_en"] = _ahora_colombia_iso()
        else:
            trabajo["mensaje"] = (
                "Cancelación solicitada. El proceso se detendrá de forma segura "
                "al terminar la actividad actual."
            )

        COLA_CONDICION.notify_all()

    return HTMLResponse("<script>window.location='/cargue-actual';</script>")


@app.get("/proceso/{trabajo_id}", response_class=HTMLResponse)
def ver_proceso(trabajo_id: str):
    trabajo = TRABAJOS.get(trabajo_id)
    if not trabajo or not _es_de(trabajo, _usuario_actual()):
        return HTMLResponse("<h2>Proceso no encontrado.</h2><p><a href='/cargue-actual'>Volver a Cargue actual</a></p>", status_code=404)
    pagina = trabajo.get("pagina_progreso")
    if pagina:
        return HTMLResponse(pagina)
    return HTMLResponse("<h2>El proceso todavía se está preparando.</h2><p><a href='/cargue-actual'>Volver a Cargue actual</a></p>", status_code=202)


# ============================================================
# ESTADO
# ============================================================

@app.get(
    "/estado/{trabajo_id}",
)
def estado_trabajo(
    trabajo_id: str,
):

    trabajo = TRABAJOS.get(
        trabajo_id
    )

    if not trabajo or not _es_de(trabajo, _usuario_actual()):

        return JSONResponse(
            {
                "error":
                    "Trabajo no encontrado"
            },
            status_code=404,
        )

    return {
        "etapa":
            trabajo["etapa"],

        "mensaje":
            trabajo["mensaje"],

        "total":
            trabajo["total"],

        "procesadas":
            trabajo["procesadas"],

        "exitosas":
            trabajo["exitosas"],

        "errores":
            trabajo["errores"],

        "terminado":
            trabajo["terminado"],

        "detalle_actividades":
            trabajo.get(
                "detalle_actividades",
                [],
            ),

        "reporte_disponible":
            os.path.isfile(
                trabajo["ruta_reporte"]
            ),
        "posicion_cola": _posicion_en_cola(trabajo_id),
        "cancelar_solicitado": bool(trabajo.get("cancelar_solicitado")),
    }


# ============================================================
# HISTORIAL Y REPORTES
# ============================================================

def _formatear_fecha_hora(fecha_iso):
    try:
        fecha = datetime.fromisoformat(fecha_iso)
        if fecha.tzinfo is None:
            fecha = fecha.replace(tzinfo=ZONA_HORARIA_COLOMBIA)
        else:
            fecha = fecha.astimezone(ZONA_HORARIA_COLOMBIA)
    except Exception:
        return "—", "—"

    fecha_texto = fecha.strftime("%d/%m/%Y")
    hora = fecha.strftime("%I:%M").lstrip("0") or "0:00"
    sufijo = "a. m." if fecha.hour < 12 else "p. m."
    return fecha_texto, f"{hora} {sufijo}"


def _historial_del_usuario(usuario=None):
    """Solo los registros de esa persona.

    Los registros viejos (anteriores al login) no tienen dueno y no se
    muestran a nadie, para no filtrar reportes de otros.
    """
    usuario = usuario or _usuario_actual()

    return [
        registro for registro in _cargar_historial()
        if _normalizar_usuario(
            registro.get("usuario_app")
        ) == _normalizar_usuario(usuario)
    ]


def _pagina_registros(tipo="historial"):
    registros = list(reversed(_historial_del_usuario()))
    es_historial = tipo == "historial"
    activo_historial = "activo" if es_historial else ""
    activo_reportes = "" if es_historial else "activo"

    filas = []
    programas = set()

    if es_historial:
        for registro in registros:
            fecha, hora = _formatear_fecha_hora(registro.get("fecha_iso", ""))
            cursos = registro.get("cursos") or [{"curso": "Curso sin nombre", "programa": "Programa sin nombre"}]

            for item in cursos:
                curso = str(item.get("curso") or "Curso sin nombre")
                programa = str(item.get("programa") or "Programa sin nombre")
                programas.add(programa)
                busqueda = html.escape(f"{curso} {programa}".lower(), quote=True)
                programa_attr = html.escape(programa, quote=True)
                registro_id = html.escape(str(registro.get("id") or ""), quote=True)

                filas.append(
                    f'<tr class="fila-registro" data-busqueda="{busqueda}" data-programa="{programa_attr}">'
                    f'<td class="fecha-celda"><span class="icono-fecha">▣</span><div><strong>{html.escape(fecha)}</strong><small>{html.escape(hora)}</small></div></td>'
                    f'<td>{html.escape(curso)}</td>'
                    f'<td>{html.escape(programa)}</td>'
                    f'<td class="accion-celda"><a class="boton-reporte" href="/reporte-guardado/{registro_id}">◉ <span>Ver reporte</span></a></td>'
                    '</tr>'
                )
    else:
        for registro in registros:
            fecha, hora = _formatear_fecha_hora(registro.get("fecha_iso", ""))
            cursos = registro.get("cursos") or [{"curso": "Curso sin nombre", "programa": "Programa sin nombre"}]
            principal = cursos[0]
            curso = str(principal.get("curso") or "Curso sin nombre")
            programa = str(principal.get("programa") or "Programa sin nombre")
            archivo = str(registro.get("archivo_reporte") or "reporte.csv")
            programas.add(programa)
            busqueda = html.escape(f"{archivo} {curso} {programa}".lower(), quote=True)
            programa_attr = html.escape(programa, quote=True)
            registro_id = html.escape(str(registro.get("id") or ""), quote=True)

            filas.append(
                f'<tr class="fila-registro" data-busqueda="{busqueda}" data-programa="{programa_attr}">'
                f'<td class="fecha-celda"><span class="icono-fecha">▣</span><div><strong>{html.escape(fecha)}</strong><small>{html.escape(hora)}</small></div></td>'
                f'<td class="archivo-reporte" title="{html.escape(archivo, quote=True)}">{html.escape(archivo)}</td>'
                f'<td>{html.escape(curso)}</td>'
                f'<td>{html.escape(programa)}</td>'
                f'<td class="accion-celda"><a class="boton-reporte" href="/reporte-guardado/{registro_id}">⇩ <span>Descargar</span></a></td>'
                '</tr>'
            )

    opciones_programa = ''.join(
        f'<option value="{html.escape(p, quote=True)}">{html.escape(p)}</option>'
        for p in sorted(programas, key=str.casefold)
    )

    if es_historial:
        titulo = "Historial de cargues"
        subtitulo = "Consulta los cursos procesados anteriormente y accede a su reporte correspondiente."
        contador = sum(len(r.get("cursos") or [1]) for r in registros)
        contador_texto = f"{contador} registros"
        placeholder = "Buscar curso o programa..."
        tabla_cabecera = "<th>Fecha y hora</th><th>Curso</th><th>Programa</th><th>Reporte</th>"
        min_table = "900px"
        vacio = "Todavía no hay cargues registrados."
    else:
        titulo = "Reportes generados"
        subtitulo = "Todos los reportes se guardan automáticamente con el nombre del curso y el programa correspondiente."
        contador = len(registros)
        contador_texto = f"{contador} reportes"
        placeholder = "Buscar reporte, curso o programa..."
        tabla_cabecera = "<th>Fecha</th><th>Nombre del reporte</th><th>Curso</th><th>Programa</th><th>Acción</th>"
        min_table = "1120px"
        vacio = "Todavía no hay reportes guardados."

    filas_html = ''.join(filas)
    if not filas_html:
        filas_html = f'<tr><td colspan="5" class="vacio">{html.escape(vacio)}</td></tr>'

    pagina = '''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>__TITULO__ - Auto Prizma Pro</title>
    <style>
        :root{--fondo:#f7f8fc;--panel:#fff;--texto:#101828;--muted:#667085;--borde:#e5e7ef;--morado:#5548e8;--sombra:0 12px 34px rgba(29,41,57,.06)}
        *{box-sizing:border-box} body{margin:0;font-family:Inter,ui-sans-serif,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--fondo);color:var(--texto)}
        .app{min-height:100vh;display:grid;grid-template-columns:235px 1fr}.sidebar{position:sticky;top:0;height:100vh;background:#fff;border-right:1px solid var(--borde);padding:28px 20px;display:flex;flex-direction:column}
        .marca{display:flex;align-items:center;gap:12px;margin-bottom:34px}.logo{width:44px;height:44px;border-radius:13px;display:grid;place-items:center;background:linear-gradient(145deg,#6d5dfc,#4338ca);box-shadow:0 8px 20px rgba(79,70,229,.25)}
        .logo svg{width:29px;height:29px;overflow:visible}.logo svg path:first-child{fill:#fff}.logo svg path:last-child{fill:#c7d2fe}.marca strong{display:block;font-size:18px}.marca span{display:block;color:var(--muted);font-size:12px;margin-top:3px}
        .nav{display:grid;gap:8px}.nav-item{padding:12px 14px;border-radius:11px;color:#475467;font-size:14px;display:flex;gap:11px;align-items:center;text-decoration:none;transition:.18s ease}.nav-item:hover{background:#f7f5ff;color:#4f46e5}.nav-item.activo{background:#f1efff;color:#4f46e5;font-weight:700}
        .estado-servicio{margin-top:auto;border:1px solid var(--borde);border-radius:14px;padding:15px}.servicio-linea{font-size:12px;font-weight:800;color:#07894f;margin-bottom:14px}.punto{width:8px;height:8px;background:#12b76a;border-radius:50%;display:inline-block;margin-right:7px}.servicio-mini{display:flex;justify-content:space-between;align-items:center;font-size:11px;color:var(--muted);margin-top:10px}.chip{background:#eef2ff;color:#4f46e5;border-radius:999px;padding:4px 8px}
        .contenido{padding:30px 34px}.panel{background:#fff;border:1px solid var(--borde);border-radius:18px;box-shadow:var(--sombra);padding:28px;max-width:1320px;margin:0 auto}.cabecera{display:flex;align-items:flex-start;justify-content:space-between;gap:20px}h1{font-size:28px;margin:0 0 7px}.subtitulo{margin:0;color:var(--muted);font-size:14px}.contador{background:#f1efff;color:#4f46e5;border-radius:10px;padding:9px 12px;font-size:12px;font-weight:800;white-space:nowrap}
        .herramientas{display:grid;grid-template-columns:minmax(260px,1fr) 290px;gap:18px;margin:28px 0 22px}.buscador,.filtro{height:52px;border:1px solid #d8dce6;border-radius:11px;background:#fff;display:flex;align-items:center;gap:10px;padding:0 15px;color:#667085}.buscador input,.filtro select{width:100%;border:0;outline:0;background:transparent;font-size:14px;color:#475467}.filtro select{cursor:pointer}
        .tabla-wrap{border:1px solid var(--borde);border-radius:14px;overflow:auto}table{width:100%;border-collapse:collapse;font-size:13px;min-width:__MIN_TABLE__}thead{background:#faf9ff}th{text-align:left;padding:15px 16px;color:#4338ca;font-size:12px;border-bottom:1px solid var(--borde)}td{padding:15px 16px;border-bottom:1px solid #eef0f4;vertical-align:middle;color:#344054}tbody tr:last-child td{border-bottom:0}.fecha-celda{display:flex;align-items:center;gap:11px;white-space:nowrap}.fecha-celda strong{display:block;font-weight:500}.fecha-celda small{display:block;color:#98a2b3;margin-top:4px}.icono-fecha{width:34px;height:34px;border-radius:9px;background:#f4f1ff;color:#6d5dfc;display:grid;place-items:center}.accion-celda{white-space:nowrap}.boton-reporte{display:inline-flex;align-items:center;gap:7px;border:1px solid #d7d0ff;color:#5b4ce8;text-decoration:none;border-radius:9px;padding:9px 12px;font-weight:700;font-size:12px;background:#fff}.boton-reporte:hover{background:#f7f5ff}.archivo-reporte{max-width:390px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.vacio{text-align:center;color:#98a2b3;padding:50px 20px!important}.oculta{display:none}
        @media(max-width:1050px){.app{grid-template-columns:1fr}.sidebar{display:none}.contenido{padding:20px}}@media(max-width:700px){.contenido{padding:12px}.panel{padding:18px}.cabecera{flex-direction:column}.herramientas{grid-template-columns:1fr}}
    </style><link rel="stylesheet" href="/estilos-responsive.css">
</head>
<body>
<div class="app">
    <aside class="sidebar">
        <div class="marca"><div class="logo"><svg viewBox="0 0 48 48" aria-hidden="true"><path d="M9 35.5 20.5 8.5c.8-1.9 3.4-1.9 4.2 0l4.1 9.6-5.4 12.7-3.1-7.4-5.2 12.1z"/><path d="M26.4 14.5 39 35.5h-8.2l-8.5-14.2z"/></svg></div><div><strong>Auto Prizma Pro</strong><span>Automatización PRIZMA</span></div></div>
        <nav class="nav">
            <a class="nav-item" href="/">⌂ <span>Inicio</span></a>
            <a class="nav-item" href="/cargue-actual">⇧ <span>Cargue actual</span></a>
            <a class="nav-item __ACTIVO_HISTORIAL__" href="/historial">◷ <span>Historial</span></a>
            <a class="nav-item __ACTIVO_REPORTES__" href="/reportes">▥ <span>Reportes</span></a>
            <a class="nav-item" href="/cambiar-clave">✎ <span>Mi contraseña</span></a>
            <a class="nav-item" href="/salir">⏻ <span>Salir</span></a>
        </nav>
        <div class="estado-servicio"><div class="servicio-linea"><span class="punto"></span> Servicio activo</div><div class="servicio-mini"><span>Navegador</span><span class="chip">Chromium</span></div><div class="servicio-mini"><span>Conexión</span><span class="chip">Estable</span></div></div>
    </aside>
    <main class="contenido"><section class="panel"><div class="cabecera"><div><h1>__TITULO__</h1><p class="subtitulo">__SUBTITULO__</p></div><div class="contador">▣ &nbsp;__CONTADOR__</div></div><div class="herramientas"><label class="buscador">⌕ <input id="buscar" type="search" placeholder="__PLACEHOLDER__"></label><label class="filtro">▽ <select id="programa"><option value="">Todos los programas</option>__OPCIONES__</select></label></div><div class="tabla-wrap"><table><thead><tr>__CABECERA__</tr></thead><tbody>__FILAS__</tbody></table></div></section></main>
</div>
<script>
const buscar=document.getElementById('buscar');const programa=document.getElementById('programa');function filtrar(){const texto=(buscar.value||'').trim().toLowerCase();const p=programa.value||'';document.querySelectorAll('.fila-registro').forEach(f=>{const okTexto=!texto||(f.dataset.busqueda||'').includes(texto);const okPrograma=!p||f.dataset.programa===p;f.classList.toggle('oculta',!(okTexto&&okPrograma));});}buscar.addEventListener('input',filtrar);programa.addEventListener('change',filtrar);
</script>
</body></html>'''

    pagina = (pagina
        .replace("__TITULO__", html.escape(titulo))
        .replace("__SUBTITULO__", html.escape(subtitulo))
        .replace("__CONTADOR__", html.escape(contador_texto))
        .replace("__PLACEHOLDER__", html.escape(placeholder, quote=True))
        .replace("__OPCIONES__", opciones_programa)
        .replace("__CABECERA__", tabla_cabecera)
        .replace("__FILAS__", filas_html)
        .replace("__MIN_TABLE__", min_table)
        .replace("__ACTIVO_HISTORIAL__", activo_historial)
        .replace("__ACTIVO_REPORTES__", activo_reportes)
    )
    return HTMLResponse(pagina)


@app.get("/historial", response_class=HTMLResponse)
def historial_cargues():
    return _pagina_registros("historial")


@app.get("/reportes", response_class=HTMLResponse)
def reportes_generados():
    return _pagina_registros("reportes")


@app.get("/reporte-guardado/{registro_id}")
def descargar_reporte_guardado(registro_id: str):
    registro = next(
        (
            r for r in _historial_del_usuario()
            if r.get("id") == registro_id
        ),
        None,
    )

    if not registro:
        return JSONResponse({"error": "Reporte no encontrado"}, status_code=404)

    nombre = os.path.basename(str(registro.get("archivo_reporte") or ""))
    ruta = os.path.join(RESULTADOS_DIR, nombre)

    if not nombre or not os.path.isfile(ruta):
        return JSONResponse({"error": "El archivo del reporte ya no está disponible"}, status_code=404)

    return FileResponse(ruta, media_type="text/csv", filename=nombre)


# ============================================================
# DESCARGAR REPORTE
# ============================================================

@app.get(
    "/reporte/{trabajo_id}",
)
def descargar_reporte(
    trabajo_id: str,
):

    trabajo = TRABAJOS.get(
        trabajo_id
    )

    if not trabajo or not _es_de(trabajo, _usuario_actual()):

        return JSONResponse(
            {
                "error":
                    "Trabajo no encontrado"
            },
            status_code=404,
        )

    ruta = trabajo[
        "ruta_reporte"
    ]

    if not os.path.isfile(
        ruta
    ):

        return JSONResponse(
            {
                "error":
                    "Reporte todavía no disponible"
            },
            status_code=404,
        )

    return FileResponse(
        ruta,
        media_type="text/csv",
        filename=os.path.basename(ruta),
    )